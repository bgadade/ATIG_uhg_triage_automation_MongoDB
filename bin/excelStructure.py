# -*- coding: utf-8 -*-

#import packages and libraries
import openpyxl
import numpy as np
import warnings
import json
import xlrd
#import utils

warnings.simplefilter("ignore")

#load the workwook from the path
def loadWorkBook(path):
    """loads the excel file in memory"""
    workbook = openpyxl.load_workbook(path)
    # ws = workbook.active
    # for index, row in enumerate(ws):
    # # for index,row in workbook.iter_rows():
    #     for cell in row:
    #         print cell
    return workbook


#get the sheetnames of the workbook
def getWorkbookSheetNames(workbook):
    """returns the list of sheets in the workbook"""
    return workbook.get_sheet_names()
   
#get the specific sheet of the workbook by name
def getWorkbookSheetByName(workbook,sheetname):
    """returns the specific sheet by name"""
    return workbook.get_sheet_by_name(sheetname)
    
#get the total number of sheets in the workbook
def getWorkbookSheetCount(workbook):
    """returns the total number of sheets in the workbook"""
    return len(getWorkbookSheetNames(workbook))


def getIndexRealValues(workbook, index):
    sheet = workbook.worksheets[index]
    rowIndex = 0
    for row in sheet.iter_rows():
        rowIndex += 1
        for cell in row:
              if np.isreal(cell.value) and cell.value is not None:
                return rowIndex
    return None

# def getIndexIntegerValues(workbook):
#     ws = workbook.active
#     for index, row in enumerate(ws):
#         for cell in row:
#             print cell
#             if cell.value_is_int():
#                 return index


def getRowMaximumColumnCount(workbook, index):
    sheet = workbook.worksheets[index]
    maxCols = 0
    for row in sheet.iter_rows():
        col_count = 0
        for columns in row:
            if columns.value is None:
                continue
            col_count += 1
            if maxCols < col_count:
                maxCols = col_count
    return maxCols

def getRowMinimumColumnCount(workbook, index):
    sheet=workbook.worksheets[index]
    minCols = 0
    for row in sheet.iter_rows():
        col_count = 0
        for columns in row:
            if columns.value is None:
                continue
            col_count += 1
            if minCols < col_count:
                minCols = col_count
        return minCols

def exportExcelToJSON(workbook):
    """returns JSON object for workbook meta data"""

    #get the sheetnames of workbook as List
    sheetnames = getWorkbookSheetNames(workbook)

    #excelJSON
    excelJSON = {}
    excelJSON['sheetcount'] = getWorkbookSheetCount(workbook)

    excelJSONTabsDictionary = {}
    excelJSON["tabDict"] = []


    for index, sheet in enumerate(sheetnames):
        # for sheet in sheetnames :
        # print sheet
        excelJSONTabsDictionary = {}
        excelJSONTabsDictionary['sheetname'] = sheet
        excelJSONTabsDictionary['RowMinimumColumnCount'] = getRowMinimumColumnCount(workbook, index)
        excelJSONTabsDictionary['maximumColumnCount'] = getRowMaximumColumnCount(workbook,index)
        excelJSONTabsDictionary['startIndex'] = getIndexRealValues(workbook,index)
        # print excelJSONTabsDictionary
        excelJSON["tabDict"].append(excelJSONTabsDictionary)

    return excelJSON


