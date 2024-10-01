import random
import utils as utils
import FetchNPIregistry as fnpi
import pandas as pd
import excelStructureXlrd as es
# import functionsForUI as funcUI
import nlpMapping as nlp
from copy import deepcopy
import json
import constants
import os
from os import listdir, rmdir, remove
import shutil
import copy
import datetime
import re
from collections import OrderedDict
import xlrd
import pickle as pickle
# import csv
import csv
import elastic
import multiprocessing
from functools import partial
import math
from xlutils.copy import copy as xlutilscopy
import openpyxl
from os import listdir
from os.path import join as osjoin
import importlib
import numpy as np
import operator
from dateutil.parser import parse
import traceback
from commonFiles import diImport
from classify_request_type import classify
from bson.binary import Binary
import hashlib
import requests
import io
from concurrent.futures import ProcessPoolExecutor
from time import perf_counter
from pymongo import MongoClient
from gridfs import GridFS
import openpyxl
from multiprocessing import cpu_count
import concurrent.futures
import asyncio
from bson import json_util
import aiohttp
from aiohttp import ClientSession, ClientResponseError
tDrivers = copy.deepcopy(constants.templateDrivers)
for k, v in list(constants.templateDrivers.items()):
    # configPath='../config/'
    for typ, filename in list(v.items()):
        if typ == 'mv':
            tDrivers[k][typ] = importlib.import_module(filename)
        else:
            #filePath = constants.configPath + filename + '.json'
            #tDrivers[k][typ] = utils.readFile(filePath, type="json")
            tDrivers[k][typ] = constants.handler.find_one_document(filename)[filename]

def plmi_columns(fdfNM,delType,provName):
    # lob_collection = db.get_collection('LobData')
    lob_collection = constants.handler.db.get_collection('LobData')
    lob_data = json.loads(json_util.dumps(lob_collection.find({})))
    lob_df = pd.DataFrame(lob_data)
    state= provName[:2] if provName else None
    #import pdb;pdb.set_trace()
    if 'Import_Del_Segment_C&S' in fdfNM.columns:
        #state_unique = fdfNM["Prac_State"].unique()
        #for state in state_unique:
            #check_state = lob_collection.find_one({'state_short_name':state})
        if lob_df[lob_df['state_short_name']==state].empty:
            return fdfNM
        check_state_index =lob_df[lob_df['state_short_name']==state].index.item()
        check_state = lob_df.loc[check_state_index]
        
        if check_state_index==0 or check_state_index: # and fdfNM['Import_Del_Segment_C&S'] == 'TRUE':
            # dsnp = True if "dsnp" in check_state.get('lob').lower() else False
            # ifp = True if "ifp" in check_state.get('lob').lower() else False
            # medicaid = True if "medicaid" in check_state.get('lob').lower() else False
            dsnp = check_state.get('DSNP')
            ifp = check_state.get('IFP')
            medicaid = check_state.get('MEDICAID')
            fdfNM['Import_Del_Product_IFP'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_Del_Segment_C&S'] == 'TRUE' and ifp else ('' if x['Import_Del_Segment_C&S'] == '' else 'FALSE'), axis = 1)
            fdfNM['Import_Del_Product_MEDICAID'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_Del_Segment_C&S'] == 'TRUE' and medicaid else ('' if x['Import_Del_Segment_C&S'] == '' else 'FALSE'), axis = 1)
            fdfNM['Import_Del_Product_DSNP'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_Del_Segment_C&S'] == 'TRUE' and dsnp else ('' if x['Import_Del_Segment_C&S'] == '' else 'FALSE'), axis = 1)
        else:
            fdfNM['Import_Del_Product_IFP'] = fdfNM['Import_Del_Product_MEDICAID'] = fdfNM['Import_Del_Product_DSNP'] = ''
        #return fdfNM

    if 'Import_NonDel_Segment_C&S' in fdfNM.columns:
        #state_unique = fdfNM["Prac_State"].unique()
        #for state in state_unique:
            #check_state = lob_collection.find_one({'state_short_name':state})
        if lob_df[lob_df['state_short_name']==state].empty:
            return fdfNM
        check_state_index =lob_df[lob_df['state_short_name']==state].index.item()
        check_state = lob_df.loc[check_state_index]
        #check_state = {'lob': 'medicaid/ifp'}
        if check_state_index: # and fdfNM['Import_NonDel_Segment_C&S'] == 'TRUE':
            dsnp = True if "dsnp" in check_state.get('lob').lower() else False
            ifp = True if "ifp" in check_state.get('lob').lower() else False
            medicaid = True if "medicaid" in check_state.get('lob').lower() else False
            fdfNM['Import_NonDel_Product_IFP'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_NonDel_Segment_C&S'] == 'TRUE' and ifp else ('' if x['Import_NonDel_Segment_C&S'] == '' else 'FALSE'), axis = 1)
            fdfNM['Import_NonDel_Product_MEDICAID'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_NonDel_Segment_C&S'] == 'TRUE' and medicaid else ('' if x['Import_NonDel_Segment_C&S'] == '' else 'FALSE'), axis = 1)
            fdfNM['Import_NonDel_Product_DSNP'] = fdfNM.apply(lambda x: 'TRUE' if x['Import_NonDel_Segment_C&S'] == 'TRUE' and dsnp else ('' if x['Import_NonDel_Segment_C&S'] == '' else 'FALSE'), axis = 1)
        else:
            fdfNM['Import_NonDel_Product_IFP'] = fdfNM['Import_NonDel_Product_MEDICAID'] = fdfNM['Import_NonDel_Product_DSNP'] = ''
        return fdfNM
    else:
        return fdfNM
    return fdfNM

# -------------------------- Checking Credentials -------------------------------------#

# def validateCredentials(_name, _password):
#     role = ''
#     for i in constants.credentialsData:
#         user = i['userName']
#         password = i['password']
#         role = i['role']
#         if user == _name and password == _password:
#             if role == 'admin':
#                 msg = 'Successfully logged in as admin'
#                 return role
#             else:
#                 msg = 'Successfully logged in as user'
#                 return role
#     msg = 'Invalid Credentials'
#     return msg

def validateCredentials(_name,_password):
    success = 0
    credentialsJson = constants.credentialsData
    if credentialsJson.get(_name) :
        di = credentialsJson[_name]
        if di['password'] != _password:
           if di['accountStatus'] == 'Locked':
               msg = "Account has been locked due to previous failed Attempts, please contact Administrator"
           else:
                di['failedAttemptCount'] +=1
                if di['failedAttemptCount'] >= 3:
                    di['accountStatus'] = "Locked"
                    msg = "Your Account is Locked after three failed Attempts"
                else:
                    msg = "Failed Attempt to login "
        else:
            if di['accountStatus'] != 'Locked':
                di['failedAttemptCount'] = 0
                role = di['role']
                if di['role'] == 'admin' and di['logStatus']=="Inactive":
                    di['logStatus'] = 'Active'
                    success = 1
                    msg = "Successfully logged in as admin"
                elif di['role'] == 'admin' and di['logStatus'] == "Active":
                    msg = "You have already logged in as admin"
                else:
                    success = 1
                    msg = "Successfully logged in as user"
            else:
                msg = "Account Locked, please contact Administrator"
        credentialsJson[_name] = di
    else:
        msg = 'User not found'
    # file write
    replaceData = {'credentials':credentialsJson}
    try:
        constants.handler.replace_one_document('credentials', replaceData)
    except Exception as e:
        print(e)
    else:
        setattr(constants,"credentialsData",credentialsJson)
    return role if success else msg

def addCredentials(data):
    originalData = constants.credentialsData
    data['accountStatus'] = 'Unlocked'
    data['failedAttemptCount'] = 0
    originalData[data['userName']] = data
    #print("originalData:", originalData)
    toinsertData = {'credentials': originalData}
    try:
        constants.handler.replace_one_document('credentials', toinsertData)
    except Exception as e:
        print(e)
    else:
        setattr(constants,"credentialsData",originalData)


def setJSONData(data, filemapping):
    try:
        constants.handler.replace_one_document(filemapping, {filemapping: data})
    except Exception as e:
        print(e)
    else:
        if filemapping == 'credentials':
            setattr(constants, "credentialsData", data)
        elif filemapping == "providers":
            setattr(constants, "provAdmScrJson", data)


# -------------------------- Reading File -------------------------------------#

def getFileJSON(filename):
    data = constants.handler.find_one_document(filename.split('.')[0])[filename.split('.')[0]]
    return data


# -------------------------- Format detection -------------------------------------#

def exportExcelToJSON(fObj):
    workbook=xlrd.open_workbook(file_contents=fObj.getvalue(),on_demand=True)
    res=esExportExcelToJSON(workbook)
    return res,workbook


def getAllProviders():
    providers = constants.provAdmScrJson
    return providers


def getTaxIDColumnValues(inputDF, taxIDColumnName):
    taxIDColumn = list(inputDF[taxIDColumnName].fillna('').astype(str))
    taxID = list(set(taxIDColumn))

    taxID = [value.replace("-", "").replace("/", "") for value in taxID if
             value != '' and not re.search('[a-z]', value, re.IGNORECASE)]
    taxID = list(map(int, taxID))

    return taxID


def readFrameWrapper(filename, excelJSON, Token,mongo,fileObj,_diDFs):
    st=perf_counter()
    sheet_count = excelJSON['sheetcount']
    sheets_dict = excelJSON['tabDict']

    diDFs = {}
    diHeaders = {}

    sheet_with_header_detail = {}
    for sheet in range(0, sheet_count):
        xlData=sheets_dict[sheet]['data']
        sheet_name = sheets_dict[sheet]['sheetname']
        header_row = sheets_dict[sheet]['startIndex']
        maximumColumnCount = sheets_dict[sheet]['maximumColumnCount']
        sheet_with_header_detail[sheet_name] = {}
        if not maximumColumnCount:  # so the condition should be based on second half only
            diDFs[sheet_name] = pd.DataFrame({'placeholder': ['']})
            diHeaders[sheet_name] = ['placeholder']
            sheet_with_header_detail[sheet_name]['diheader'] = ['placeholder']
            sheet_with_header_detail[sheet_name]['format'] = ['placeholder']
            continue
        xlrdHeaderIx=getXlrdHeaderIx(header_row)
        xlrdHeaderRow=xlData[xlrdHeaderIx]
        inputDF, cleaned_header = readFrame(filename,sheet, header_row,xlrdHeaderRow,fileObj,_diDFs)

        # print(inputDF)
        # original_header = xlData[int(header_row - 1)]
        inputDF = inputDF.dropna(how="all")
        diDFs[sheet_name] = inputDF
        diHeaders[sheet_name] = xlrdHeaderRow
        sheet_with_header_detail[sheet_name]['diheader'] = xlrdHeaderRow
        sheet_with_header_detail[sheet_name]['format'] = cleaned_header
        sheet_with_header_detail[sheet_name]['dfLen'] = len(inputDF)
        sheet_with_header_detail[sheet_name]['dfSnapshot'] = inputDF.head(constants.nrowsSnapshotMapping).to_json()

    # inserting into mongoDB Supressed - To be enabled in tandem with readPickle
    # input_DF = pickle.dumps(diDFs)
    # sot_header = pickle.dumps(diHeaders)
    # pickle_inpdf_to_insert = {"Token": Token, "Type": "InpDF", "Pickle_Data": Binary(input_DF)}
    # constants.handler.insert_one_document(pickle_inpdf_to_insert['Type'], pickle_inpdf_to_insert)
    # pickle_sotheader_to_insert = {"Token": Token, "Type": "sotHeader", "Pickle_Data": Binary(sot_header)}
    # constants.handler.insert_one_document(pickle_sotheader_to_insert['Type'], pickle_sotheader_to_insert)

    diDFsToMongo(diDFs,Token,mongo)
    diDFsToMongo(diDFs, "copy_" + Token, mongo)
    # diHeadersToMongo(diHeaders, Token)
    logPerf(st,'readFrameWrapper')
    return sheet_with_header_detail

def writeChunksToMongo(strFormat,token,type):
    lstChunks = re.findall('.{1,%s}' % (constants.chunkLimit), strFormat)
    parts=len(lstChunks)
    for ix, chk in enumerate(lstChunks):
        data = {'Token': token, 'Pickle_Data': chk, 'Type': type, 'part': ix+1,'parts':parts}
        constants.handler.insert_one_document(type, data)

def readChunksFromMongo(token,type):
    find_data={"Token": token, "Type": type}
    lstDocs = constants.handler.find_documents(type, find_data)
    lstDocsSrt = sorted(lstDocs, key=lambda doc: doc['part'])
    strFormat = ''.join([doc['Pickle_Data'] for doc in lstDocsSrt])
    return strFormat

def diDFsToMongo(diDFs,token,mongo):
    stMain = perf_counter()
    for k, v in diDFs.items():
        fp = io.BytesIO()
        v.to_parquet(fp)
        fp.seek(0)
        saveToGridFS(mongo, token + 'InpDF', fp, 'InpDF',tab=k)
        fp.close()
    logPerf(stMain,'diDFsToMongo')

def diHeadersToMongo(diHeaders,token):
    strFormat=str(diHeaders)
    data={'Token':token,'Pickle_Data':strFormat,'Type': 'sotHeader'}
    # re.findall('.{1,100')
    constants.handler.insert_one_document('sotHeader', data)


def diDFsFromMongo(token,mongo,fs):
    st = perf_counter()
    strFormat=retrieveFileFromMongo(token+"InpDF",mongo.db,fs,multiple=True,collectby='tab')
    diDFs={k: pd.read_parquet(io.BytesIO(v)) for k, v in strFormat}
    logPerf(st,'diDFsFromMongo')
    return diDFs

def diHeadersFromMongo(token):
    find_data = {"Token": token, "Type": "sotHeader"}
    strFormat = constants.handler.find_one_document('sotHeader', find_data)['Pickle_Data']
    diHeaders = eval(strFormat)
    return diHeaders

def readFrame(filename,sheet, headerIx,xlrdHeaderRow,fileObj,_diDFs,taxIDColumnName=None):
    globFileLoc=constants.loc+filename
    # inputDF = utils.readFile(fileObj, type="xlsx", sheetIndex=int(sheet), headerIndx=int(headerIx), colAsStr=True,_diDFs=_diDFs)
    inputDF = _diDFs[int(sheet)]
    xlrdHeaderRow = xlrdHeaderRow[:len(inputDF.columns)]
    inputDF.columns = [col.strip().lower() for col in map(str, inputDF.columns)]
    inputDF = utils.renameDuplicateColumns(inputDF)
    # inputDF = utils.dropEmptyRows(inputDF)
    inputDF=inputDF.applymap(lambda cel: np.nan if (isinstance(cel, str) and not cel.strip()) else cel.strip() if isinstance(cel,str) else cel)
    cleaned_header, inputDF = renameSOTCols(globFileLoc, int(sheet), int(headerIx), inputDF,xlrdHeaderRow,taxIDColumnName)


    return inputDF,cleaned_header


def newReadPickle(token, type):
    if type == ".pkl":
        dfname = token + "InpDF" + '.pkl'
        df = pd.read_pickle("../tmp/" + dfname)
    else:
        dfname = token + ".sotHeader"
        df = pd.read_pickle("../tmp/" + dfname)

    return df

#MongoDB Code Supressed - can be revived as is
# def readPickle(token, **kwargs):
#     if kwargs.get('type'):
#         find_data = {"Token": token, "Type": kwargs.get('type')}
#         pickledObj = pickle.loads(constants.handler.find_one_document(kwargs.get('type'), find_data)['Pickle_Data'])
#         return pickledObj
#     find_data = {"Token": token, "Type": "InpDF"}
#     inputDF = pickle.loads(constants.handler.find_one_document('InpDF', find_data)['Pickle_Data'])
#     return inputDF

def readPickle(token,**kwargs):
    if kwargs.get('type'):
        with open("../tmp/"+token+'.'+kwargs.get('type'), 'rb') as handle:
            pickledObj = pickle.load(handle)
        return pickledObj
    dfname = token + "InpDF" + '.pkl'
    inputDF=pd.read_pickle("../tmp/" + dfname)
    return inputDF

def getSingleValueMap(inputDF, provName=None):
    # print sheet
    # print globFileLoc
    # global globSheet
    # globSheet = sheet
    # inputDF = utils.readFile(globFileLoc, type="xlsx", sheetIndex = int(sheet)) #Todo hit a func and get a inp-df
    cols = inputDF.columns
    if provName:
        svMappedOPColsDict, mappedOPColsList, unmappedOPCols, newInpCols, ioFeildMappings = provMapSv(provName, cols,
                                                                                                      constants.mappings)
    else:
        svMappedOPColsDict, mappedOPColsList, unmappedOPCols, newInpCols, ioFeildMappings = svMapColumns(cols,
                                                                                                         constants.mappings)
    svFeildMappings = getSvFeildMappings(svMappedOPColsDict, inputDF, cols, mappedOPColsList, unmappedOPCols,
                                         newInpCols, ioFeildMappings, constants.mappings)
    return svFeildMappings


def getMultiValueMap(inputDF, svUserMapping,
                     provName=None):  # svUserMapping is {"sv":{"inpCol1":"outputCol1","inpCol2":"outputCol2"}
    # print globFileLoc
    # print globSheet
    # inputDF = utils.readFile(globFileLoc, type="xlsx", sheetIndex = int(globSheet))
    availableInputCols = [col for col in inputDF.columns if col not in list(svUserMapping['sv'].keys())]
    asIsAssoc = {}
    if provName:
        dictMVMappings, asIsAssoc = provMapMv(provName, availableInputCols, constants.mvMappings)
    else:
        dictMVMappings = utils.mvMapColumns(availableInputCols, constants.mvMappings, constants.functionMappings)
    mvMappedOPColsDict, mvMappedOPColsList, mvUnmappedOPCols, mvUnmappedInpCols = utils.combineMVMapping(dictMVMappings)
    mvFeildMappings = getMvFeildMappings(dictMVMappings, mvUnmappedOPCols, constants.mvMappings, mvUnmappedInpCols,
                                         'mv', svUserMapping, mvMappedOPColsDict, list(inputDF.columns),
                                         asIsAssoc=asIsAssoc)
    return mvFeildMappings


def svMapColumns(cols, mappings):
    return utils.mapColumns(cols, constants.mappings)


def getInFields(inputDF, cols, svMappedOPColsDict, lstBestMatch):
    inputFields = []
    lstBestMatch1 = list(set([item['bestMatch'] for item in lstBestMatch if item['bestMatch']]))
    # inputFields["inFields"] = []
    for col in cols:
        inFields = {}
        inFields["inputField"] = col
        inFields["sampleInput"] = list(map(str, list(inputDF.head(3)[col])))
        mapped = col in list(svMappedOPColsDict.keys())
        if col in lstBestMatch1:
            mapped = 'nlp'
        inFields["mapped"] = mapped
        inputFields.append(inFields)
    return inputFields


def getBestMatch(outputCols, inpCols, mappings, fieldType, loadPickle=True):
    # TODO: NLP logic for determining best match
    svLoopBreakCutOff = 500
    mvLoopBreakCutOff = 300
    if fieldType == 'sv':
        excluded = []
        nlpMap = {}
        if loadPickle:
            with open(constants.svNlpPicklePath, 'rb') as fp:
                diLookup = pickle.load(fp)
        else:
            diLookup = {key: val['Input_Column'] for key, val in list(mappings.items())}
        print('diLookup in sv:', diLookup)
        for inpCol in inpCols:
            di = {key: val for key, val in list(diLookup.items()) if key not in excluded and key in outputCols}
            maxScore = 0
            maxScoreDict = {}
            for outputCol, lookup in list(di.items()):
                if not lookup:
                    continue
                scoreDict, maxScoreLst = nlp.get_nearest_matching_word(lookup, inpCol)
                try:
                    if maxScoreLst[1] > maxScore:
                        maxScore = maxScoreLst[1]
                        maxScoreDict['outputCol'] = outputCol
                        maxScoreDict['score'] = maxScore
                except:
                    continue
            try:
                if maxScoreDict['score'] > 0.75:
                    nlpMap.update({inpCol: maxScoreDict['outputCol']})
                    excluded.append(maxScoreDict['outputCol'])
            except:
                continue
            # print '*****************nlpMap:',nlpMap
            if len(
                    list(nlpMap.keys())) > svLoopBreakCutOff:  # runs nlp logic only to some extent in order to reduce processing time #todo to removed later
                break
        nlpMapReverse = {val: key for key, val in list(nlpMap.items())}
        return nlpMapReverse
    elif fieldType == 'mv':
        excluded = []
        nlpMap = {}
        if loadPickle:
            with open(constants.mvNlpPicklePath, 'rb') as fp:
                diParentCat = pickle.load(fp)
        else:
            diParentCat = getMvLookup()

        for inpCol in inpCols:
            # di = {key['PARENT_CATEGORY']: [val['Column_Type']['Input_Column'] for key, val in mappings.items() if
            #       key not in excluded and key in outputCols}
            maxScore = 0
            maxScoreDict = {}
            diMatchCount = {}
            for outputCol, lookup in list(diParentCat.items()):
                if outputCol in excluded:
                    continue
                if not lookup:
                    continue
                scoreDict, maxScoreLst = nlp.get_nearest_matching_word(lookup, inpCol)
                try:
                    if maxScoreLst[1] > maxScore:
                        maxScore = maxScoreLst[1]
                        maxScoreDict['outputCol'] = outputCol
                        maxScoreDict['score'] = maxScore
                except:
                    continue
            # excluded.append(maxScoreDict['outputCol'])
            try:
                if maxScoreDict['score'] > 0.75:
                    nlpMap.update({inpCol: maxScoreDict['outputCol']})
            except:
                pass

            for k, v in list(nlpMap.items()):
                # diMatchCount[v]=diMatchCount.get(v,0)+1
                if v not in diMatchCount:
                    diMatchCount[v] = []
                diMatchCount[v].append(k)
                if len(diMatchCount[v]) == mvLoopBreakCutOff:
                    excluded.append(v)
            # print '\n\n*********************** diMatchCount:',diMatchCount

        return nlpMap
        # nlpMapReverse = {val: key for key, val in nlpMap.items()}
        # return nlpMapReverse'


def getUnmapOtFields(cols, mappings, unmappedInpCols, field_type):
    unmapOtFields = []
    # bestMatchDict=getBestMatch(cols,unmappedInpCols,mappings,field_type,loadPickle=True)
    bestMatchDict = getBestMatchElastic(cols, unmappedInpCols, mappings, field_type)
    # bestMatchDict={}
    if field_type == 'mv':
        return bestMatchDict
    for col in cols:
        unmapOtField = {}
        unmapOtField["outputField"] = col
        unmapOtField["bestMatch"] = bestMatchDict.get(col)
        unmapOtFields.append(unmapOtField)
    return unmapOtFields


def getSvFeildMappings(svMappedOPColsDict, inputDF, inputcols, mappedOPColsList, unmappedOPCols, newInpCols,
                       ioFeildMappings, mappings):
    mapping = {}
    mapping["unmapOtFields"] = getUnmapOtFields(unmappedOPCols, mappings, newInpCols, 'sv')
    # mapping["inFields"] = getInFields(svMappedOPColsDict,inputDF,inputcols,mapping['unmapOtFields'])
    mapping["inFields"] = getInFields(inputDF, inputcols, svMappedOPColsDict, mapping['unmapOtFields'])
    mapping["mappedOtFields"] = ioFeildMappings
    return mapping


def filterMvUnmappedInpCols(mvUnmappedInpCols, svMappedInp):
    # svMappedInp = []
    # for di in svFieldMappings['inFields']:
    #     if di['mapped'] and di['mapped']!='nlp':
    #         svMappedInp.append(di['inputField'])
    mvUnmappedInpCols = [col for col in mvUnmappedInpCols if col not in svMappedInp]
    return svMappedInp, mvUnmappedInpCols


def computeUnmappedInpCols(svMappedInp, mvMappedInp, mvNlpMappedInp, sotColumns):
    unmappedInpCols = [col for col in sotColumns if col not in svMappedInp + mvMappedInp + mvNlpMappedInp]
    return unmappedInpCols


def getMvFeildMappings(dictMVMappings, mvUnmappedOPCols, mvMappings, mvUnmappedInpCols, fieldType, svUserMapping,
                       mvMappedOPColsDict, sotColumns, asIsAssoc={}):
    diParentCat = {}
    asisTag = 'AsIs'
    # mvMappings=orderMapping(mvMappings)
    for key, val in list(mvMappings.items()):
        mapElement = {'order': val['order'], 'outputField': key, 'inputField': None}
        if val['PARENT_CATEGORY'] not in diParentCat:
            diParentCat[val['PARENT_CATEGORY']] = {}

        if 'Column_Type' in val:
            tags = [di['Tag'] for di in val['Column_Type']]
            if "GROUP_DETAILS" in val:
                grpName = val["GROUP_DETAILS"]["GROUP_NAME"]
                orderInGrp = val["GROUP_DETAILS"]["ORDER_WITHIN_GROUP"]
                mapElm = copy.deepcopy(mapElement)
                mapElm.update({"groupName": grpName, "orderWithinGroup": orderInGrp, "groupOrder": 0})
                for tag in tags:
                    if tag not in diParentCat[val['PARENT_CATEGORY']]:
                        diParentCat[val['PARENT_CATEGORY']].update({tag: [mapElm]})
                    else:
                        diParentCat[val['PARENT_CATEGORY']][tag].append(mapElm)
            else:
                for tag in tags:
                    if tag not in diParentCat[val['PARENT_CATEGORY']]:
                        diParentCat[val['PARENT_CATEGORY']].update({tag: [mapElement]})
                    else:
                        diParentCat[val['PARENT_CATEGORY']][tag].append(mapElement)

        if val['DATAFRAME_TYPE'] == asisTag:
            if asisTag not in diParentCat[val['PARENT_CATEGORY']]:
                diParentCat[val['PARENT_CATEGORY']].update({asisTag: [mapElement]})
            else:
                diParentCat[val['PARENT_CATEGORY']][asisTag].append(mapElement)

    diAll = {}
    for key, val in list(diParentCat.items()):
        for k, v in list(val.items()):
            di = {'parentCategory': key, 'type': k, 'map': []}
            v.sort(key=lambda x: x['order'])
            v1 = copy.deepcopy(v)
            for elm in v1:
                del elm['order']
            di['map'] = v1
            # for outputCol in [vl for ki,vl in sorted(v)]:
            #     di['map'].append({'outputField': outputCol, 'inputField': None})
            diAll.update({key + k: di})
    # print 'diAll:',json.dumps(diAll)
    # exit(0)
    mvFieldMappings = []
    mapped = []
    suggested = []
    covered = []
    for parentCategory in dictMVMappings:
        filterdDict = dictMVMappings[parentCategory]
        dictReverse = {val: key for key, val in list(filterdDict[0].items())}
        for lst in filterdDict[1]:
            type = parentCategory + lst[0].split('@')[0]
            di = deepcopy(diAll[type])
            lstMappedCols = [(item, item.split('@')[1]) for item in lst]
            di = updateDiWithGroups(di, lstMappedCols)
            for tup in lstMappedCols:
                spltOutFld = tup[1].split('#')
                for map in di['map']:
                    if map['outputField'] == spltOutFld[0]:
                        if len(spltOutFld) == 2:
                            if map['groupOrder'] == int(spltOutFld[1]):
                                map['inputField'] = dictReverse[tup[0]]
                                break
                        else:
                            map['inputField'] = dictReverse[tup[0]]
                            break
            di['order'] = lst[0].split('@')[-1]
            # if lst[0].split('@')[0] == 'GENERAL':
            #     for asIs in lstAsIs[0:2]:
            #         di['map'].append({'outputField': asIs, 'inputField': dictReverse[asIs]})

            mapped.append(deepcopy(di))
            covered.append(type)

        asisType = parentCategory + asisTag
        if asisType not in diAll:
            continue
        asisDi = deepcopy(diAll[asisType])
        for d in asisDi['map']:
            d['inputField'] = dictReverse.get(d['outputField'])
            if d['inputField']:
                if parentCategory in asIsAssoc:
                    d['assocSet'] = asIsAssoc[parentCategory].get(d['inputField'])
                else:
                    d['assocSet'] = None
        mapped.append(asisDi)
        covered.append(asisType)

    for k, v in list(diAll.items()):
        if k not in covered:
            mapped.append(v)
    mvFieldMappings.append({'mapped': mapped})
    svMappedInp, mvUnmappedInpCols = filterMvUnmappedInpCols(mvUnmappedInpCols, list(svUserMapping['sv'].keys()))
    bestMatchDict = getUnmapOtFields(mvUnmappedOPCols, mvMappings, mvUnmappedInpCols, fieldType)
    mvNlpMappedInp = []
    for inpCol, nlpMatch in bestMatchDict.items():
        suggested.append({'parentCategory': nlpMatch, 'inputField': inpCol})
        mvNlpMappedInp.append(inpCol)
    mvFieldMappings[0].update({'suggested': suggested})
    unmappedInpCols = computeUnmappedInpCols(svMappedInp, list(mvMappedOPColsDict.keys()), mvNlpMappedInp, sotColumns)
    mvFieldMappings[0].update({'unmappedInpCols': unmappedInpCols})
    return mvFieldMappings


class NestedDict(dict):
    def __getitem__(self, key):
        if key in self: return self.get(key)
        return self.setdefault(key, NestedDict())


def transformUserMapping(userMapping, type):  # todo we need to print userMaping
    # print("usermapping:", userMapping)
    if type == 'sv':
        lstTransformed = []
        lstTransformed.insert(0, userMapping)
        lstTransformed.insert(1, list(userMapping.values()))
        lstTransformed.insert(2, [key for key in list(constants.mappings.keys()) if key not in list(userMapping.values())])
        lstTransformed.insert(3, [])
        return lstTransformed
    elif type == 'mv':
        diTransformed = utils.NestedDict()
        for parentCat, di in list(userMapping.items()):
            filterdMappingJson = {key: value for key, value in list(constants.mvMappings.items()) if
                                  value['PARENT_CATEGORY'] == parentCat}
            diTransformed[parentCat] = []
            diTemp = utils.NestedDict()
            mappedOPColsList = []
            for outputFieldString in list(di.values()):
                groupId = outputFieldString.split('@')[2]
                mappedOPColsList.append(outputFieldString.split('@')[1].split('#')[0])
                if groupId not in diTemp:
                    diTemp.update({groupId: []})
                diTemp[groupId].append(outputFieldString)
            diTransformed[parentCat].insert(0, di)
            diTransformed[parentCat].insert(1, [v for k, v in sorted(diTemp.items())])
            unmappedOPCols = [col for col in list(filterdMappingJson.keys()) if col not in mappedOPColsList]
            #             unmappedInpCols=[col for col in inputColumns  if col not in di.keys()]
            unmappedInpCols = []
            diTransformed[parentCat].insert(2, unmappedOPCols)
            diTransformed[parentCat].insert(3, unmappedInpCols)
        return diTransformed


def renameSOTCols(ExcelLocation, sheetIdx, HeaderIdx, inputDF, xlrdHeaderRow,taxIDColumnName=None):  # mk taxIDColumnName=None
    def getrow(ws, hdidx):
        if hdidx != 0:
            return ws.row(hdidx - 1)
        return ws.row(hdidx)

    def CleanColumnName(row, dflength, taxIDColumnName=None):
        col = []
        dstnctCol = []
        dupCol = []
        # mk updatedTaxIDColumnName = taxIDColumnName

        for i, vlu in enumerate(row):
            vlu = str(vlu)
            if (vlu is None or vlu.strip() == '') and i < dflength:
                col.append("empty_column")
                dupCol.append("empty_column")
            elif vlu is not None and vlu.strip() != '':
                val = vlu

                # temp_val = val
                # temp_val = standardHeaderColumn([temp_val])[0]
                # mk taxIDColumnName = standardHeaderColumn([taxIDColumnName])[0]

                val = re.sub("tricare.*only[\s.:;,]*\n", '', val, flags=re.IGNORECASE)
                val = val.replace('\n', '|').replace('\r', '|')
                # print vlu.value, val
                val = " ".join([x.strip() for x in val.split("|") if x])  # [x for x in val.split("|") if x][0]
                val = val.strip().lower()
                col.append(val)
                if val not in dstnctCol:
                    dstnctCol.append(val)
                else:
                    dupCol.append(val)
            else:
                pass
        return col, dupCol

    def getFinalRowHeader(rowCol, dupCol):
        process = []
        for i in dupCol:
            count = 1
            for j, x in enumerate(rowCol):
                if x == i:
                    if x in process:
                        count = count + 1
                    else:
                        process.append(x)
                    v = x + str(count)
                    rowCol[j] = v
        return rowCol

    rowCol, dupCol = CleanColumnName(xlrdHeaderRow, len(inputDF.columns))
    cleanedHeader = getFinalRowHeader(rowCol, dupCol)
    inputDF.columns = cleanedHeader
    return cleanedHeader, inputDF

def renameDfCols(inputDF,mapping,tab,filename,sotHeader,validatedActionColumn):
    inputDF[validatedActionColumn]=constants.defaultTTyp
    inputDF[constants.combinedInpActionColNm]=""
    svMappedOPColsDict, mappedOPColsList, unmappedOPCols, unmappedInpCols = transformUserMapping(mapping['sv'],
                                                                                                 'sv')
    # print 'unmappedOPCols:',unmappedOPCols
    dictMVUserMapping = transformUserMapping(mapping['mv'], 'mv')
    mvMappedOPColsDict, mvMappedOPColsList, mvUnmappedOPCols, mvUnmappedInpCols = utils.combineMVMapping(
        dictMVUserMapping)
    # print 'mvUnmappedOPCols:',mvUnmappedOPCols

    dataframe = utils.renameColumns(inputDF, svMappedOPColsDict)  # ToDo  rename it for single value
    dataframe = utils.renameColumns(dataframe, mvMappedOPColsDict)  # ToDo then rename it for multi value
    diColMapping = dict(list(zip(dataframe.columns.tolist(), sotHeader)))
    relevantCols = list(svMappedOPColsDict.values()) + list(mvMappedOPColsDict.values())
    dataframe = dataframe[relevantCols + [constants.validatedActionColNm,constants.combinedInpActionColNm]]
    dataframe = dataframe.loc[dataframe.applymap(lambda cell: str(cell).strip() != '').any(axis=1)]

    # dataframe = dataframe.dropna(subset=relevantCols, how="all")
    # print 'diColMapping:',diColMapping

    if svMappedOPColsDict.get('NA'):
        unmappedOPCols.append(svMappedOPColsDict['NA'])

    dataframe[constants.fileNmCol] = filename
    dataframe[constants.tabNmCol] = tab
    missingCols = unmappedOPCols + mvUnmappedOPCols + list(constants.variablesCommon[
        "npiReg"].keys())  # Todo: if NPI reg fails #Todo: sv_unamapped and camel casing everywhere.
    groupCols = [k for k, v in list(constants.mvMappings.items()) if v.get('GROUP_DETAILS')]
    missingCols = [col for col in missingCols if col not in groupCols]
    dataframe = utils.addMissingColumns(dataframe, missingCols)  # Todo: Rohan to get back on This
    dataframe = dataframe.fillna('')
    st = perf_counter()
    dataframe = utils.fillFinalNPIs(dataframe, "FINAL_NPI")
    logPerf(st, 'fillFinalNPIs')
    #dataframe['TAX_ID'] = dataframe['TAX_ID'].map(lambda a: ''.join(re.findall(r'\b\d+\b', a.strip())))
    dataframe['TAX_ID'] = dataframe['TAX_ID'].map(lambda a: ''.join(re.findall(r'\b\d+\b', a.strip())) if not bool(re.search("reject\:|error\:|clarify\:", a, re.IGNORECASE)) else a)

    return dataframe,diColMapping,dictMVUserMapping


def modelActions(dataframe):
    lstTabDataDi=[]
    if constants.storeActions:
        lstTabDataDi=storeActions(dataframe)
    if constants.predictActions:
        dataframe=predictActions(dataframe)
    return lstTabDataDi,dataframe
# userMapping={'s1':{'sv':{},'mv':{}},'s2':{}, 's3'}

def createActionForHash(dataframe):
    dataframe[constants.actionColForHash] = dataframe.apply(
        lambda row: row[constants.combinedInpActionColNm] if row[constants.combinedInpActionColNm] else row[
            constants.tabNmCol], axis=1)
    return dataframe

def uiDriverFunction(userInterimFile, token, filename, userID,delType,mongo,fs,fileType=None, ff=False, provName=None, validatedActions=None):
    stMain=perf_counter()
    logData = constants.handler.find_one_document('logging', {'token':token})
    t1 = datetime.datetime.now()
    print(t1)
    userID = userID
    if not ff:
        writeUserMappingToDisk(userInterimFile, userID, "UserMapping", provName,fileType)
    masterDictListSot = []

    fNm, fileExt = os.path.splitext(filename)
    if fileExt == '.pdf':
        diDFs = readPickle(token)
    else:
        diDFs = diDFsFromMongo(token,mongo,fs)
    # diHeaders = readPickle(token, type='sotHeader')
    diHeaders = {k: v['diheader'] for k, v in userInterimFile.items()}
    splittedFrames = []
    argsMPW = []
    tabCounter = 0
    weakHashElmsForActionData=[]
    strongHashElmsForActionData=[]
    validatedActionData=[]
    lstData=[]
    lstSheet=[]
    lstNPI=[]
    tabsTtyp = []
    for tmplt, tmplVal in validatedActions['data'].items():
        for ttyp in tmplVal:
            if set(ttyp['TransactionType']).issubset(validatedActions['AllTransactionTypes']):
                tabsTtyp.append(''.join(ttyp['TransactionType']))
            else:
                tabsTtyp.append(constants.defaultTTyp)
    setDefaultTTyp=True if not tabsTtyp else True if set(tabsTtyp).issubset(set(['NON MASS','DEFAULT'])) else False
    for tab, tab_info in sorted(list(userInterimFile.items()),key=lambda tup:tup[0]):
        mapping = tab_info['mappings']
        if isinstance(mapping, str) and mapping.lower() == "to be ignored":
            ############## logging #############################
            for ix,di in enumerate(logData['tabInfo']):
                if di['tabName'] == tab:
                    logData['tabInfo'][ix]['tabProcessed'] = False
            ####################################################
            continue

        elif isinstance(mapping, dict):
            inputDF = diDFs[tab]
            sotHeader = diHeaders[tab]
            validatedActionColumn = constants.validatedActionColNm
            dataframe,diColMapping,dictMVUserMapping=renameDfCols(inputDF,mapping,tab,filename,sotHeader,validatedActionColumn)

            ################## Added for Validated Actions by User ##################

            validatedData = validatedActions["data"][tab] if validatedActions else []




            for record in validatedData:
                # record['rowIndex'] = record['updatedrowIndex']
                record['rowIndex'] = record.get('updatedrowIndex')
                dataframe.iloc[record['rowIndex'], dataframe.columns.get_loc(validatedActionColumn)] = \
                    record['TransactionType'][0] if determineTemplate(
                        record['TransactionType'][0]) else constants.defaultTTyp
                dataframe.iloc[record['rowIndex'],dataframe.columns.get_loc(constants.combinedInpActionColNm)]=" ".join(sorted(record['colValues']))
                ###################changes all demo all rows to nonmass in case of only (demo other and nonmass)transaction type############
            # if len(dataframe['VALIDATED_ACTION'].unique()) <= 2 and set((dataframe['VALIDATED_ACTION'])).issubset(set(['NON MASS', 'DEMO OTHER'])):
            if setDefaultTTyp:
                dataframe['VALIDATED_ACTION'] = 'NON MASS'
            ############ logging purpose ########################################
            for ix, di in enumerate(logData['tabInfo']):
                if di['tabName'] == tab:
                    logData['tabInfo'][ix]['tabProcessed'] = True
                    logData['tabInfo'][ix]['rowCount'] = dataframe.shape[0]
                    logData['tabInfo'][ix]['nUniqueIds'] = len(list(dataframe.set_index(['FINAL_NPI', 'TAX_ID'], drop=False).index.unique()))
                    mappedTTypes = dataframe['VALIDATED_ACTION'].value_counts().to_dict()
                    logData['tabInfo'][ix]['transactionInfo'] = [{"transactionType":ttype, "rowCount":int(num)} for ttype, num in mappedTTypes.items()]
            #######################################################################

            #######################ADDED FOR THE CONTRACT PRAC ADD TEMPLATE#############################
            dataframe = getDerivedTemplateData(dataframe,constants.templates,delType)
            ############################################################################################
            if len(list(dataframe.index)) == 0:
                tabCounter += 1
                continue
            dataframe=createActionForHash(dataframe)
            validatedActionData.extend(list(dataframe[constants.validatedActionColNm]))
            weakHashElmsForActionData.extend(collectHashElms(copy.deepcopy(dataframe),constants.weakHashColsForActionsData))
            strongHashElmsForActionData.extend(collectHashElms(copy.deepcopy(dataframe),constants.strongHashColsForActionsData))
            st = perf_counter()
            lstTabDataDi,dataframe=modelActions(copy.deepcopy(dataframe))
            logPerf(st, 'modelActions')
            lstData.extend(lstTabDataDi)

            # Then Do this for multi value

            sv_splitFrameCols = utils.merge_two_dicts(constants.mappings, constants.variablesCommon["mandatory"])
            sv_splitFrameCols = utils.merge_two_dicts(sv_splitFrameCols, constants.variablesCommon["npiReg"])
            mv_splitFrameCols = utils.merge_two_dicts(constants.mvMappings, constants.variablesCommon["mandatory"])

            idxLvls = constants.uidLevels
            dataframe['idxCol'] = dataframe.apply(lambda row: 'pi3.141'.join([row[lvl] for lvl in idxLvls]), axis=1)
            lstNPI.extend(dataframe[constants.finalNpiCol].unique())
            dataframeMI = dataframe.set_index('idxCol')
            # if len(list(dataframeMI.index))==0 or (len(list(dataframeMI.index.unique()))==1 and not any(list(list(dataframeMI.index.unique())[0]))):
            #     continue

            svCols = list(sv_splitFrameCols.keys())
            mvCols = getMvCols(mv_splitFrameCols, dataframe.columns)
            uniqueIds = list(dataframeMI.index.unique())
            stFrm = perf_counter()
            for ix,uid in enumerate(uniqueIds):
                frm=dataframeMI.loc[[uid]].reset_index(drop=True)
                frmNm=uid.split('pi3.141')
                frmToJson=frm.to_json()
                #frmToJson=frm
                frmTup=[frmNm,frmToJson]
                argsMPW.append([frmTup, ix, svCols, mvCols, dictMVUserMapping, diColMapping, tabCounter])
            logPerf(stFrm, 'Creating Frames', ovrRide=True)
            # st=perf_counter()
            # splittedFramesLocal, rejectedFrames = utils.splitFrames1(frames, sv_splitFrameCols,
            #                                                          mv_splitFrameCols, dictMVUserMapping, diColMapping,
            #                                                          tabCounter)  # splittedFrames --> list of single valued frames, multi valued frames for every unique NPI + Tax_ID combination.
            # logPerf(st, 'splitFrames1')


            # argsMPW.extend([[[ix,frm],svCols,mvCols,dictMVUserMapping, diColMapping,tabCounter] for ix,frm in enumerate(frames)])
            # splittedFramesLocal, rejectedFrames = utils.splitFrames2(frames, svCols,mvCols, dictMVUserMapping, diColMapping,tabCounter)
            # splittedFrames --> list of single valued frames, multi valued frames for every unique NPI + Tax_ID combination.
            '''Above function should return the list of Tuples (Single value dataframe, multivalue dataframe)
            pass two parameters to this function one is single value mapping another one is multivalue mapping'''

            # todo for every tuple process Single value and multivalue data frame
            # npiCols = constants.variables["npiReg"]
            # singleValuedFrames = []
            # multiValuedFrames = []
            # splittedFrames.extend(splittedFramesLocal)
            tabCounter += 1

    weakHash=createHash([elm.encode('utf-8') for elm in weakHashElmsForActionData])
    strongHash = createHash([elm.encode('utf-8') for elm in strongHashElmsForActionData])
    # strongHash=createHash(strongHashElmsForActionData)
    actionDocument={"weakHash":weakHash,"strongHash":strongHash,"data":lstData,"time":datetime.datetime.now().__str__(),"validatedActions":validatedActionData}
    if constants.storeActions:
        strFormat=str(actionDocument)
        fileObj = io.BytesIO(strFormat.encode('utf-8'))
        saveToGridFS(mongo, token + 'ActionData', fileObj, 'ActionData', col='ActionsGFS')
        fileObj.close()
    # with open(constants.actionModelDataFilePath,'wb') as fp:
    #     fp.write(json.dumps({"hash":hash,"data":lstData,"time":datetime.datetime.now().__str__(),"va":[]}))

    masterDictList = []
    # VariableCols = outputCols
    # splittedFrames=reGroupTTyp(splittedFrames)
    argsMPW=reGroupTTyp(argsMPW)

    lstNPI = list(set(lstNPI))
    st = perf_counter()
    argsMPW = npiLookupAsync(lstNPI, argsMPW)
    logPerf(st, 'npiLookupAsync', ovrRide=True)
    logPerf(stMain,"uiDriverFunction before multiProcWork",ovrRide=True)
    st=perf_counter()
    if constants.debug:
        batches = [argsMPW]
        for batch in batches:
            res=multiProcWorkWrapper(batch)
            masterDictList.extend(res)
    else:
        masterDictList=getMasterDictList(argsMPW)
        # with multiprocessing.Manager() as manager:
        #     L = manager.list()
        #     number_processes = int(max(min(multiprocessing.cpu_count(), math.ceil(len(argsMPW) / 10.0)), 1))
        #     print('number_processes',number_processes)
        #     pool = multiprocessing.Pool(2)
        #     workFunc = partial(multiProcWorkWrapper, L)
        #     results = pool.map_async(workFunc, argsMPW)
        #     pool.close()
        #     pool.join()
        #     # print L
        #     masterDictList = [x for x in L]
        #     # masterDictList.sort()
        #     masterDictList.sort(
        #         key=lambda x: (x['sheetIdx'], x['recordIdx']))  # sorting on tab index (x[2]) and record index (x[2])
    # masterDictListSot=masterDictList
    templateDataDi = {}
    for di in masterDictList:
        templateName = di.get('templateName')
        cdi = {k: v for k, v in list(di.items()) if k in ['sheetIdx', 'recordIdx']}
        diNonMass = dict(list(cdi.items()) + [('masterDict', di['masterDictNonMass'])])
        nmTName = constants.nmTName
        if nmTName not in templateDataDi:
            templateDataDi.update({nmTName: []})
        templateDataDi[nmTName].append(diNonMass)
        if templateName:
            diTemplate = dict(list(cdi.items()) + [('masterDict', di['masterDictTemplate'])])
            if templateName not in templateDataDi:
                templateDataDi.update({templateName: []})
            templateDataDi[templateName].append(diTemplate)

    for template, templateData in list(templateDataDi.items()):
        masterDictListLayer = filterVar(template, val='MasterDictListLayer')
        pcats = list({pcat for pcat, di in list(masterDictListLayer[0].items()) if di['derivations']})
        if templateData:
            diPcatDf = {}
            for pcat in pcats:
                pcatDf = pd.concat([di['masterDict'][pcat].assign(srno=idx) for idx, di in
                                    enumerate(templateData)])  # joins all address dataframes into one
                diPcatDf.update({pcat: pcatDf})
            results = cleanseDF(tDrivers[template]['mv'], diPcatDf,
                                masterDictListLayer)  # runs operations on the unified address df
            for pcat in pcats:  # decouples the updated unified addres df into individual dfs
                for i in range(len(templateData)):
                    tdf = results[pcat].loc[results[pcat]['srno'] == i, :].drop('srno', axis=1)
                    templateData[i]['masterDict'][pcat] = tdf
            templateDataDi[template] = templateData
    diFdf={}
    #import pdb;pdb.set_trace()
    plmiData=fetchPlmiData(provName)
    for template, templateData in list(templateDataDi.items()):
        fdfLayer = filterVar(template, val='fdfLayer')
        fdf = getFinalDf(templateData, template)
        fdf = fdf.reset_index(drop="True")
        fdf = cleanseDF(tDrivers[template]['mv'], {"fdf": fdf,"plmiData":plmiData}, fdfLayer)["fdf"]
        diFdf.update({template: fdf})

    if constants.runInterTmpltLayer:
        allTmp=constants.allTmpNm
        miscLayer = filterVar(allTmp, val='misc')
        diFdf=interTemplateDerivation(diFdf, miscLayer)

    fdfOutColDi = {}
    for template, templateFdf in list(diFdf.items()):
        fdfOutCol = getfdfOutCol(templateFdf.fillna(''), tDrivers[template]['output'],delType)

        fdfOutColDi.update({template: fdfOutCol})
    fdfNMDi = prependNonMass(fdfOutColDi)
    # tmp = fdfNMDi['prac']
    # tmp.columns = map(lambda val: re.sub('[^\w]', '_', val), list(tmp.columns)) #renames columns to view tmp dataframe in pycharm
    fdfNMDi=hilightCmnTTyp(fdfNMDi)
    if constants.fetchPlmi:
        plmiData=plmiWrapper(plmiData)
        fdfNMDi=assignPlmiData(fdfNMDi,plmiData)
    outputTemplatesLog = []    # logging
    templateOutputDi = {}
    for template, fdfNM in list(fdfNMDi.items()):
        if template in constants.templates:
            outputTemplatesLog.append({'template':constants.templates[template]['Screenname'][0],
                                   'rowCount':fdfNM.shape[0],
                                   'nUniqueIds':len(list(fdfNM.set_index(['Prac_TIN', 'Prac_NPI'], drop=False).index.unique()))})
        if template == str('demo-addr-inact'):
            bs = fdfNM.duplicated(keep='first')
            fdfNM = fdfNM[~bs]
            fdfNM = fdfNM.reset_index(drop=True)
        if delType == 'delegate':
            plmi_columns(fdfNM,delType,provName)
        outputJson = getOutputJson(fdfNM)
        templateOutputDi.update({template: outputJson})

    # outputJsonPsi = getOutputJsonNew(fdfPsi.fillna(''), constants.outputColsPsi)

    # masterDictListPSI = mv.cleanseDF(masterDictListPSI, constants.psiLayerDf)
    #
    # # for tupple in splittedFrames:
    # masterDictListCopy.sort()
    #
    # # print masterDictList
    # fdfSot = getFinalDf([x[1] for x in masterDictListCopy])
    # outputJsonSot = getOutputJson(fdfSot.fillna(''),constants.outputCols)
    templates = list(constants.templates.keys())
    outputJson = {tmp: templateOutputDi.get(tmp, None) for tmp in templates}
    t2 = datetime.datetime.now()
    ############### logging ###################################
    logData['nTabsProcessed'] = tabCounter
    logData['outputTemplatesInfo'] = outputTemplatesLog
    logData['businessRuleStartTimestamp'] = str(t1)
    logData['businessRuleEndTimestamp'] = str(t2)
    startProcessTimestamp = datetime.datetime.strptime(logData['startProcessTimestamp'],"%Y-%m-%d %H:%M:%S.%f")
    logData['totalFileProcessingTime'] = str(t2 - startProcessTimestamp)
    logData['processCompleteFlag'] = True
    constants.handler.replace_one_document('logging',logData)
    ###########################################################
    print(t2)
    print("***********************Time Taken to run : " + str(t2 - t1))
    return outputJson

def applyVariables(masterDict, varFile, mvFile):
    for layer in varFile:
        if layer['ProcessStage'] != 0:
            continue
        if layer['LayerType'] == 'ROW':
            masterDict = deriveMVCols(mvFile, masterDict, layer['Transformation'])
        elif layer['LayerType'] == 'DF':
            masterDict = cleanseDF(mvFile, masterDict, layer['Transformation'])
    return masterDict

def multiProcWorkWrapper(lstArgs):
    waste=[]
    lstRes=[]
    for arg in lstArgs:
        recordTup=arg[0]
        recordNm=recordTup[0]
        recordBytes=recordTup[1]
        record=pd.read_json(recordBytes,dtype=str)
        #record=recordBytes
        recordIx = arg[1]
        svCols, mvCols=arg[2],arg[3]
        dictMVUserMapping, diColMapping,tabCounter=arg[4],arg[5],arg[6]
        ttype = recordNm[-1]
        svDf,mvDf = utils.splitFrames3(record, svCols, mvCols)
        svDf=svDf.assign(**arg[-1])
        params=[svDf,mvDf,recordIx,dictMVUserMapping, diColMapping,tabCounter,ttype,recordNm]
        res=multiProcWork(waste,params)
        lstRes.append(res)
    return lstRes


def multiProcWork(masterDictList, tupple):
    st=perf_counter()
    recordNm=tupple[7]
    recordIx=tupple[2]
    sheetIx=tupple[5]
    nmTName = constants.nmTName
    ttype = tupple[6]
    template = determineTemplate(ttype)

    dictMVUserMapping = tupple[3]
    diColMapping = tupple[4]
    # Todo get NPI supplementary data.
    # enrichedDataFrame, emptyNpiDFs, failedNpis = fnpi.enrichDataFrame(tupple[0], npiCols) #ToDo this should be part of single value data frame operation
    enrichedDataFrame = tupple[0]
    # MultiValueOP = DeriveMultiValue(tupple[0], VariableCols, mv_mappings) # returns a dictionary {"Address": dataframe} (both notrmalized and with all associated derivations completed)
    # ndf=utils.mv_address_normalization(tupple[1], mv_mappedOPColsList)
    mvFile = None
    varFile = None
    if tDrivers.get(template):
        mvFile = tDrivers[template].get('mv')
        varFile = tDrivers[template].get('var')
    # Todo Normalization of Multi value dataframe to be done here.
    masterDict = utils.mvNormalizeColumns(tupple[1], constants.functionMappings, dictMVUserMapping,
                                          constants.mvMappings)  # TOdo: recheck

    # Todo Field wise derivation for Single value
    # singlevalueOP, failedSNpis = fnpi.deriveSingleValue(enrichedDataFrame,outputCols)  # returns a single output dataframe
    masterDict = utils.merge_two_dicts(masterDict, {"singleValue": enrichedDataFrame})  # Todo: check with rohan**

    masterDict.update({"diColMapping": diColMapping})

    # Todo Field wise derivation for Multi value
    # masterDict = mv.deriveMVCols(masterDict, constants.multivaluedCols, constants.ndb_p,constants.midlevNdBSpec)
    # masterDict = mv.deriveMVCols(masterDict, cleansingCols, constants.midlevNdBSpec)
    # masterDict = mv.deriveMVCols(masterDict, multivaluedCols, constants.midlevNdBSpec)

    nonMassTempDict = copy.deepcopy(masterDict)

    masterDictNonMass = applyVariables(nonMassTempDict, tDrivers[nmTName]['var'], tDrivers[nmTName]['mv'])
    if mvFile and varFile:
        masterDictTemplate = applyVariables(masterDict, varFile, mvFile)
        resultDi={'recordIdx': recordIx, 'masterDictNonMass': masterDictNonMass, 'masterDictTemplate': masterDictTemplate,
             'sheetIdx': sheetIx, 'templateName': template}
    else:
        resultDi={'recordIdx': recordIx, 'masterDictNonMass': masterDictNonMass, 'sheetIdx': sheetIx,
                               'templateName': template}
    npi,tin=recordNm[0],recordNm[1]
    print("processed ===> pid:({})|recordIx:({})|sheetIx:({})|template:({})|NPI:({})|TIN:({}) | Time Taken:({:.2f}s)".format(os.getpid(),recordIx,sheetIx,template,npi,tin,perf_counter()-st))
    # masterDictList.append(resultDi)
    return resultDi
    # masterDict = mv.cleanseDF(masterDict, var.cleansingDf)
    # masterDict = mv.deriveMVCols(masterDict, var.cleansingCols)
    # masterDict = mv.deriveMVCols(masterDict, var.multivaluedCols)
    # masterDict = mv.cleanseDF(masterDict, var.finalCleansingDf)
    # masterDict = mv.cleanseDF(masterDict, var.psiLayerDf)

    # multivalueOP, failedMNpis = fnpi.deriveMultiValue(enrichedDataFrame, normalizedMVFrame, outputCols)  # returns a single output dataframe
    # normalized_mv_frames.append(normalizedMVFrame)

    # singleValuedFrames.append(singlevalueOP)
    # multiValuedFrames.append(multivalueOP)


def writeUserMappingToDisk(userMapping, userID, fileMapping, provName=None,fileType=None):
    if fileType=='attachmentFile':
        userData_insert = {"UserName": userID, "Type": fileMapping, "Doc": userMapping}
        constants.handler.insert_one_document(fileMapping, userData_insert)
        if provName:
            providerInterim_insertData = {"UserName": userID, "Provider": provName, "Type": "ProviderInterim",
                                          "Doc": userMapping}
            constants.handler.replace_one_document(providerInterim_insertData['Type'], providerInterim_insertData)


def identifyMappingDiff(userMappings, diDiff):
    #print('userMappings:', userMappings)
    # exit(0)
    for sheet, map in list(userMappings.items()):
        if isinstance(map, dict):
            for key, val in list(map['sv'].items()):
                if key not in constants.mappings[val]['Input_Column']:
                    if val not in diDiff['sv']:
                        diDiff['sv'].update({val: []})
                    if key not in diDiff['sv'][val]:
                        diDiff['sv'][val].append(key)
            for parentCat, di in list(map['mv'].items()):
                for key, val in list(di.items()):
                    tag = val.split('@')[0]
                    outputField = val.split('@')[1].split('#')[0]
                    filterdDi = constants.mvMappings[outputField]
                    tags = sorted(
                        [elem['Tag'] for elem in filterdDi['Column_Type']]) if 'Column_Type' in filterdDi else None
                    if filterdDi['PARENT_CATEGORY'] == parentCat:
                        flag = 0
                        if filterdDi.get('Column_Type'):
                            for diMap in filterdDi['Column_Type']:
                                if key in diMap['Input_Column']:
                                    flag = 1
                                    break
                        else:
                            if key in filterdDi.get('Input_Column'):
                                flag = 1
                                break

                        if not flag:
                            if tags is not None and 'tags' not in diDiff['mv'][parentCat][outputField]:
                                diDiff['mv'][parentCat][outputField].update({"tags": tags})
                            if tag not in diDiff['mv'][parentCat][outputField]:
                                diDiff['mv'][parentCat][outputField].update({tag: []})
                            if key not in diDiff['mv'][parentCat][outputField][tag]:
                                diDiff['mv'][parentCat][outputField][tag].append(key)
    return diDiff


def commitAdminMapping(adminMapping,userMapIds):
    print('adminMapping:', adminMapping)
    svMappingsCopy = copy.deepcopy(constants.mappings)
    mvMappingsCopy = copy.deepcopy(constants.mvMappings)
    if adminMapping.get('sv'):
        for key, val in list(adminMapping['sv'].items()):
            svMappingsCopy[key]['Input_Column'] += val
    if adminMapping.get('mv'):
        for parentCat, di in list(adminMapping['mv'].items()):
            for outputField, di1 in list(di.items()):
                filtereddDi = mvMappingsCopy[outputField]
                for tag, lstInpCols in list(di1.items()):
                    if filtereddDi['PARENT_CATEGORY'] == parentCat:
                        try:
                            for diMap in filtereddDi['Column_Type']:
                                if diMap['Tag'] == tag:
                                    diMap['Input_Column'] += [list(d.values())[0] for d in lstInpCols]

                        except:
                            if filtereddDi['DATAFRAME_TYPE'] == 'AsIs':
                                filtereddDi['Input_Column'] += [list(d.values())[0] for d in lstInpCols]
                            pass
    #print('updatedMvMappings:', mvMappingsCopy)
    #print('updatedSvMappings:', svMappingsCopy)
    mvMappings_data = {'mvMappings': mvMappingsCopy}
    try:
        constants.handler.replace_one_document('mvMappings', mvMappings_data)
    except Exception as e:
        print(e)
    else:
        setattr(constants, 'mvMappings', mvMappingsCopy)
    mappings_data = {'mappings': svMappingsCopy}
    try:
        constants.handler.replace_one_document('mappings', mappings_data)
    except Exception as e:
        print(e)
    else:
        setattr(constants, 'mappings', svMappingsCopy)
    backupOneUserMapping("UserMapping",userMapIds)
    elastic.createMappingIndexes()


def adminScreenDriverFunction():
    diDiff = utils.NestedDict()
    lstFiles,user_mapping_ids,totalCount = constants.handler.find_n_documents('UserMapping')
    lstFiles = [file['Doc'] for file in lstFiles]
    for file in lstFiles:
        userMappings = {k: v['mappings'] for k, v in file.items()}
        diDiff = identifyMappingDiff(userMappings, diDiff)
    return diDiff,user_mapping_ids,totalCount


def dfToJson(inputDf, toJson=1):
    dataJson = {}
    colSet = inputDf.columns.tolist()
    dataSet = inputDf.values.tolist()
    dataJson.update({'colSet': colSet, 'dataSet': dataSet})
    if toJson:
        return json.dumps(dataJson)
    else:
        return dataJson


def sotDataToJson(inputDf):
    return dfToJson(inputDf)


def getFinalDf(templateData, templateName):
    fdf = pd.DataFrame()
    templateData = copy.deepcopy(templateData)
    for di in templateData:
        shtIdx = di['sheetIdx']
        recIdx = di['recordIdx']
        masterDict = di['masterDict']
        confLayer = filterVar(templateName, val='conflictRemovalLayer')
        pcats = list({k for d in confLayer for k in d if k!='singleValue'})
        if pcats:
            masterDict = {k: v for k, v in list(masterDict.items()) if k in pcats}
            dfs = [v for k, v in list(masterDict.items())]
            multiValDf = pd.concat(dfs, axis=1)
            multiValDf['tmp'] = 1
            singleValDf = di['masterDict']['singleValue']
            singleValDf['tmp']=1
            record = pd.merge(singleValDf, multiValDf, on='tmp')
        else:
            record = di['masterDict']['singleValue']
        record['sheetIdx'] = shtIdx
        record['recordIdx'] = recIdx
        record = record.sort_index(axis=1)
        fdf = fdf._append(record)
    return fdf


def getOutputJsonNew(df, outputTemplate):
    dataJson = []

    outputJson = deepcopy(outputTemplate)
    for key, val in list(outputJson.items()):
        colSet = []
        dataSetCols = []
        for di in val['finalDfColNames']:
            if di['dfColName']:
                colSet.append(di['outputColName'])
                dataSetCols.append(di['dfColName'])

        outputDF = df[dataSetCols]

        taxIDColumn = outputDF['OUT_TAX_ID'].values.tolist()
        npiColumn = outputDF['OUT_NPI'].values.tolist()

        for tax_id, npi in zip(taxIDColumn, npiColumn):
            if tax_id != "" and npi != "":
                tax_id = str(tax_id)
                npi = str(npi)

                try:
                    data_row = outputDF.loc[(outputDF['OUT_TAX_ID'].astype(str) == tax_id)
                                            & (outputDF['OUT_NPI'].astype(str) == npi)]
                except KeyError:
                    print("Rows not found for " + tax_id + " and " + npi)

            elif tax_id != "" and npi == "":
                tax_id = str(tax_id)

                try:
                    data_row = outputDF.loc[(outputDF['OUT_TAX_ID'].astype(str) == tax_id)]
                except KeyError:
                    print("Rows not found for " + tax_id + " and " + npi)

            elif tax_id == "" and npi != "":
                npi = str(npi)

                try:
                    data_row = outputDF.loc[(outputDF['OUT_NPI'].astype(str) == npi)]
                except KeyError:
                    print("Rows not found for " + tax_id + " and " + npi)

            transactionType = data_row.iloc[0]['OUT_TRANSACTION_TYPE']
            row = data_row.values.tolist()[0]

            data = {}
            data['NPI'] = npi
            data['TAX ID'] = tax_id
            # data['PSI ROW'] = [colSet, row]
            data['PSI ROW'] = {
                "colSet": colSet,
                "dataSet": row
            }
            data['TRANSACTION TYPE'] = transactionType.split(',')

            dataJson.append(data)

    return dataJson

## optimizing for both zip and individual excel file downloading
def exportReviewedData(reviewedJsons, token, downloadFlag,logData):
    outputFiles = []
    lstFileObj = []
    consolidated_meta = {"TemplateName": [], "nLines": [], "nUniqueNPI": []}
    ConsolidatedFileObj = io.BytesIO()
    ConsolidatedWriter = pd.ExcelWriter(ConsolidatedFileObj, engine='openpyxl')

    transactions = [] ## transactions list to make PSI sheet names
    for idx, transaction in enumerate(reviewedJsons):
        for trastype, transdata in list(transaction.items()):
            transactions.append(trastype)

    for idx, reviewedJson in enumerate(reviewedJsons):
        lstDf = []
        for cat, jsn in list(reviewedJson.items()):
            if jsn == None:
                lstDf.append(pd.DataFrame())
                continue
            df = pd.DataFrame(jsn['data']['dataSet'],
                              columns=jsn['data']['colSet'])
            lstDf.append(df)
        fdf = pd.DataFrame([di for df in lstDf for di in df.to_dict('records')])

        ##outputFile = cat  # "Output_" + cat + "_" + token

        ##-----------------------for PSI Sheet Names ----------------------##
        if cat == 'contract-prac-add-to-grp' and 'tax-id' in transactions and \
                len(transactions) == 2:
            cont_grp_cat = 'taxid_cont_grp'
            outputFile = constants.psi_sheet_names.get(cont_grp_cat)
        elif cat == 'contract-prac-add-to-grp' and \
                'demo-non-prac-add' in transactions \
                and len(transactions) == 2:
            cont_grp_cat = 'np_prac_add_cont_grp'
            outputFile = constants.psi_sheet_names.get(cont_grp_cat)
        else:
            outputFile = constants.psi_sheet_names.get(
                cat, cat.replace('-', ' ').upper()
            )

        fdf.to_excel(ConsolidatedWriter, sheet_name=outputFile, index=False)
        ws = ConsolidatedWriter.sheets[outputFile]

        IndividualFileObj = io.BytesIO()
        IndividualWriter = pd.ExcelWriter(IndividualFileObj, engine='openpyxl')
        fdf.to_excel(IndividualWriter, sheet_name='Sheet1', index=False)

        fdfMeta = pd.DataFrame()

        for tName, tData in list(reviewedJson.items()):
            if tData == None:
                continue
            if tName == 'non-mass':
                outputColsList = tDrivers[tName]['output']["finalDfColNames"]
            else:
                outputColsList = tDrivers['non-mass']['output']["finalDfColNames"] + \
                                 tDrivers[tName]['output']["finalDfColNames"]

            numeric_cols = {}
            for ele in outputColsList:
                if ele["dataType"] == "numeric":
                    numeric_cols[ele["outputColName"]] = ele["format"]

            for col, format in list(numeric_cols.items()):
                excel_col_idx = fdf.columns.get_loc(col)
                for i in range(2, len(fdf) + 2):  # Ignore header - Start at index 2, Add 2 to include last row
                    num_val = ws.cell(row=i, column=excel_col_idx + 1)  # Internally points to WS location and any changes to num_val will change the worksheet as well
                    if (num_val.value).isnumeric():  # Values that are non numeric will remain as it is
                        num_val.value = int(num_val.value)  # Change the data type before changing the format
                        num_val.number_format = format
            fdfMeta = pd.DataFrame([{'nLines': len(fdf.index),'nUniqueNPI': len(fdf['Prac_NPI'].unique())}])
            consolidated_meta["nLines"].append(fdfMeta.at[0, 'nLines'])
            consolidated_meta["nUniqueNPI"].append(fdfMeta.at[0, 'nUniqueNPI'])

        consolidated_meta["TemplateName"].append(cat)

        fdfMeta.to_excel(IndividualWriter, sheet_name='Meta', index=False)
        IndividualWriter.close()#IndividualWriter.save()
        IndividualFileObj.seek(0)
        lstFileObj.append(IndividualFileObj)
        if not downloadFlag:
            outputFile = "Output_" + cat + "_" + token
        if downloadFlag:
            outputFile = "Output_" + cat
        outputFiles.append({list(reviewedJson.keys())[0]: outputFile})

    if not downloadFlag: ## returning file object for individual file downloading
        return outputFiles, lstFileObj
    if downloadFlag:
        if logData['inputNpiData']:
            meta2 = reconsolation(logData, reviewedJsons, downloadFlag, token)
            meta2.to_excel(ConsolidatedWriter, sheet_name='RECONCILIATION REPORT', index=False)

        consdfMeta2 = pd.DataFrame(consolidated_meta)
        consdfMeta2.to_excel(ConsolidatedWriter, sheet_name='META', index=False)
        ConsolidatedWriter.close()
        ConsolidatedFileObj.seek(0)
        lstFileObj.append(ConsolidatedFileObj)
        outputFile = "Consolidated_" + token  ## Excel File Name for consolidated sheet only
        outputFiles.append({list(reviewedJson.keys())[0]: outputFile})
        return outputFiles, lstFileObj  ## returning file object for zip downloading

def inputsotnpi(userMapping,token,filename,mongo,fs):
    try:
        #fetching the input sheets data from mongodb with the new token name
        diDFs = diDFsFromMongo("copy_"+token, mongo, fs)
        new_diDFs = diDFs.copy()

        #list stores the complete npi data of input sheets which are processed in UI
        complete_input_npi = []
        for each_sheet, value in userMapping.items():
            sheet_data = value['mappings']
            if sheet_data != 'to be ignored':
                #extracting the single_value mappings from usermapping which we got from UI
                sv_mappings = sheet_data['sv']
                for dataframe in new_diDFs:
                    if dataframe == each_sheet:
                        extract_df = new_diDFs[dataframe]

                        #renaming the input sheet df with standard headers
                        extract_df.rename(columns=sv_mappings, inplace=True)

                        if 'INDIVIDUAL_NPI' in extract_df.columns:
                            extract_npi = extract_df.get('INDIVIDUAL_NPI',pd.Series())
                            extract_npi = extract_npi.to_list()
                            complete_input_npi.extend(extract_npi)
                        else:
                            none_list = []
                            none_list = none_list + [None] * len(extract_df)
                            complete_input_npi.extend(none_list)


        #removing duplicates from total npi list
        complete_input_npi = [None if isinstance(item, float) and math.isnan(item) else item for item in complete_input_npi]   # New changes 
        extract_unique_npi = list(OrderedDict.fromkeys(complete_input_npi))

        #list contains the count of unique npi's repeated in input sot
        count_repeating_lines = []
        for each_npi in extract_unique_npi:
            lines = complete_input_npi.count(each_npi)
            count_repeating_lines.append(lines)
        json_column = {'Provider NPI': extract_unique_npi, 'lines in input sot': count_repeating_lines}
        return json_column
    except Exception as e:
        return {}

def reconsolation(logData,reviewedJsons,downloadFlag,token):
    try:
        if downloadFlag:
            new_list = []
            inputFileData = logData['inputNpiData']
            extract_unique_npi = inputFileData['Provider NPI']

            #creating the df with input unique npi data and how many times it repeated
            meta_df = pd.DataFrame(inputFileData)
            npi_dict = {}
            #fetch the npi's from transactions and store in dictionary
            for idx, reviewedJson in enumerate(reviewedJsons):
                for temp_name, jsn in list(reviewedJson.items()):
                    df = pd.DataFrame(jsn['data']['dataSet'], columns=jsn['data']['colSet'])
                    transaction_npi = df['Prac_NPI']
                    transaction_npi = transaction_npi.to_list()
                    npi_dict[temp_name] = transaction_npi

            #cleaning the npi's which are stored in dictionary and updating to same dic
            for tmplt_name,value in npi_dict.items():
                transaction_npi = value
                clean_npi = []
                for item in transaction_npi:
                    if item.startswith('C'):
                        item = item.strip()
                        new_list = item.split(' ')
                        clean_npi.append(new_list[-1])
                    else:
                        clean_npi.append(item.strip())
                npi_dict[tmplt_name] = clean_npi

            #counting the repetation of unique npi's from dictionary and adding to the df with new column
            for trans_name, trans_value in npi_dict.items():
                clean_npi = trans_value
                transaction_count_updation = []
                for each_npi in extract_unique_npi:
                    lines = clean_npi.count(each_npi)
                    transaction_count_updation.append(lines)

                meta_df[trans_name] = transaction_count_updation
        return meta_df
    except Exception as e:
        return e


def backupUserMappings(filter):
    allUserData = constants.handler.find_documents(filter)
    constants.handler.insert_many_documents('Bkp', allUserData)
    constants.handler.delete_many_documents(filter)

def backupOneUserMapping(doc,userMapIds):
    for userMapId in userMapIds:
        userData = constants.handler.find_one_document(doc)
        constants.handler.insert_one_document('Bkp', userData)
        constants.handler.delete_by_id(doc,userMapId)


def getMvLookup():
    lstParentCat = []
    diParentCat = {}
    for key, val in list(constants.mvMappings.items()):
        # if key not in excluded and key in outputCols:
        if val['PARENT_CATEGORY'] not in lstParentCat:
            lstParentCat.append(val['PARENT_CATEGORY'])
        for lookup in val.get('Column_Type', []):
            if val['PARENT_CATEGORY'] not in diParentCat:
                diParentCat.update({val['PARENT_CATEGORY']: []})
            if lookup['Input_Column'] != [""]:
                diParentCat[val['PARENT_CATEGORY']] += lookup['Input_Column']
    return diParentCat


def pickleNlpLookup():
    for fieldType in ['sv', 'mv']:
        if fieldType == 'sv':
            file_mapping = "pickledSvNlpLookup"
            diLookup = {key: val['Input_Column'] for key, val in list(constants.mappings.items())}
        elif fieldType == 'mv':
            file_mapping = "pickledMvNlpLookup"
            diLookup = getMvLookup()
        diTokenizedLookup = {}
        for pCat, looup in list(diLookup.items()):
            diTokens = {}
            for word in looup:
                dict_word_combination = nlp.make_word_combination(word)
                diTokens[word] = dict_word_combination
            diTokenizedLookup.update({pCat: diTokens})
        pickledData_to_replace = {file_mapping: str(diTokenizedLookup)}
        constants.handler.replace_one_document(file_mapping, pickledData_to_replace)


def pickleSotHeader(filename, headerData, token):
    print('filename in pickling:', filename)
    sotHeader = es.loadWorkBook(globFileLoc).sheet_by_index(headerData['sheetIndex']).row_values(
        headerData['headerIndx'] - 1)
    with open("../tmp/" + token + ".sotHeader", 'wb') as handle:
        pickle.dump(sotHeader, handle)


def svAdminMapping():
    mappings = constants.mappings
    return mappings


def mvAdminMapping():
    mvMappings = constants.mvMappings
    return mvMappings


def inputfunction(lists):
    tmpdict = {}
    for i, element_index in enumerate(lists):
        tmpdict[lists[i]] = lists[i + 1]
        return tmpdict


def commitSvAdminMapping(svAdminMapping):
    try:
        constants.handler.replace_one_document('mappings', {'mappings': svAdminMapping})
    except Exception as e:
        print(e)
    else:
        setattr(constants, 'mappings', svAdminMapping)


def commitMvAdminMapping(mvAdminMapping):
    mvMappingsCopy = copy.deepcopy(constants.mvMappings)
    if mvAdminMapping:
        for parentCat, di in list(mvAdminMapping.items()):
            for outputField, di1 in list(di.items()):
                filtereddDi = mvMappingsCopy[outputField]
                for tag, lstInpCols in list(di1.items()):
                    if filtereddDi['PARENT_CATEGORY'] == parentCat:
                        try:
                            for diMap in filtereddDi['Column_Type']:
                                if diMap['Tag'] == tag:
                                    diMap['Input_Column'] = [list(d.values())[0] for d in lstInpCols]
                                    # diMap['Input_Column']=lstInpCols
                        except:
                            if filtereddDi['DATAFRAME_TYPE'] == 'AsIs':
                                filtereddDi['Input_Column'] += lstInpCols
                            pass
    print('mvAdminMappings:', mvMappingsCopy)
    try:
        constants.handler.replace_one_document('mvMappings', {'mvMappings': mvMappingsCopy})
    except Exception as e:
        print(e)
    else:
        setattr(constants, 'mvMappings', mvMappingsCopy)


def getBestMatchElastic(outputCols, inpCols, mappings, fieldType):
    if fieldType == 'sv':
        excluded = []
        nlpMap = {}
        st = datetime.datetime.now()
        for inpCol in inpCols:
            outputCol = elastic.elasticMapping('sv_nlp', inpCol, excluded=excluded)
            if outputCol:
                nlpMap.update({inpCol: outputCol})
                excluded.append(outputCol)
        nlpMapReverse = {val: key for key, val in list(nlpMap.items())}
        et = datetime.datetime.now()
        print('***************sv_elasticmap time taken:', (et - st).total_seconds())
        return nlpMapReverse

    elif fieldType == 'mv':
        excluded = []
        nlpMap = {}
        st = datetime.datetime.now()
        for inpCol in inpCols:
            pCat = elastic.elasticMapping('mv_nlp', inpCol)
            if pCat:
                nlpMap.update({inpCol: pCat})
                excluded.append(pCat)
        et = datetime.datetime.now()
        print('***************mv_elasticmap time taken:', (et - st).total_seconds())
        return nlpMap


def zeroPadding(string, lenth):
    if len(string) > lenth:
        string = string[-lenth:]
    return '0' * (lenth - len(string)) + string


def processNDBFile(filename):
    elasticConfig = utils.readFile('../config/elasticConfig.json', type="json")["index_input_files"]
    with open('../input/' + filename, 'rb') as fp:
        temp = pd.read_excel(fp, skiprows=1)
    with open('../input/' + filename, 'rb') as fp:
        ndb_grid = pd.read_excel(fp, skiprows=1, converters={i: str for i in range(len(temp.columns))})
    ndb_grid = ndb_grid.loc[((-ndb_grid['Prov Type Category'].isin(['ALLIED HLT PROF GRP', 'PHYSICIAN GRP'])) & (
        -ndb_grid['NDB DEG'].isnull()))]
    ndb_grid = ndb_grid[['Prov Type Name', 'NDB DEG', 'NDB SPEC', 'COS PROV', 'COS CRED', 'NDB REC', 'UR Ind']]
    ndb_grid['NDB SPEC'] = ndb_grid['NDB SPEC'].map(lambda x: zeroPadding(str(x), 3))
    ndb_grid['COS PROV'] = ndb_grid['COS PROV'].map(lambda x: zeroPadding(str(x), 4))
    ndb_grid = ndb_grid.fillna('')
    ndb_grid.columns = [re.sub('[ ]+', '_', col.lower()) for col in ndb_grid.columns]
    ndb_grid.to_csv('../config/' + str(filename) + '.csv', sep=',', encoding='utf-8', index=False)
    indexName = elasticConfig[filename]
    elastic.index_data_bulk(indexName, doc_type='my_type', recreate=1, csv_file_path='../config/',
                            csv_file_name=filename)
    # elastic.index_data_bulk('ndb_taxonomy',doc_type='my_type',recreate=1,csv_file_path='../config/',csv_file_name=None)


def processGridFile(filename, type, skipRows):
    elasticConfig = utils.readFile('../config/elasticConfig.json', type="json")["index_input_files"]
    with open('../input/' + filename, 'rb') as fp:
        temp = pd.read_excel(fp, skiprows=skipRows)
    with open('../input/' + filename, 'rb') as fp:
        ndb_grid = pd.read_excel(fp, skiprows=skipRows, converters={i: str for i in range(len(temp.columns))})

    ndb_grid.columns = [re.sub('[ ]+', '_', col.lower()) for col in ndb_grid.columns.str.strip()]
    ndb_grid.to_csv('../config/' + type + '.csv', sep=',', encoding='utf-8', index=False)
    indexName = elasticConfig[type]
    elastic.index_data_bulk(indexName, doc_type='my_type', recreate=1, csv_file_path='../config/', csv_file_name=type)


def getCsvToJson(filemapping):
    outFile = json.dumps(constants.handler.find_one_document(filemapping)[filemapping])
    return outFile


def getJsonToCsv(filemapping, data, colNames):
    df = pd.DataFrame(data)
    df = df.fillna('NA')
    jsonData = df.to_json(orient='records')
    jsonData = json.loads(jsonData)
    data_to_replace = {filemapping: jsonData}
    constants.handler.replace_one_document(filemapping, data_to_replace)
    if filemapping == 'specialty_exceptions':
        backupUserMappings('SpecialtyException')
        reIndexing(filemapping)
    else:
        reIndexing(filemapping)


def adminSpecExceptDriverFunction():
    diDiff = []
    lstFiles = constants.handler.find_documents('SpecialtyException')
    lstFiles = [file['Doc'] for file in lstFiles]
    for file in lstFiles:
        diDiff = identifySpecialityDiff(file, diDiff)
        print("diDiff", diDiff)
    return diDiff


def identifySpecialityDiff(userSpeciality, diDiff):
    for item in userSpeciality:
        flag = 0
        if not diDiff:
            diDiff.append(item)
        for x in diDiff:
            if set(item.values()) == set(x.values()):
                flag = 1
                pass
        if flag == 0:
            diDiff.append(item)
    return diDiff


def reIndexing(idxName):
    elastic.index_data_bulk(idxName, doc_type='my_type', recreate=1, csv_file_path='../config/', csv_file_name=None)
    print('indexing done for index:', idxName)


def readGridFile(filename, skipRows):
    with open('../input/' + filename, 'rb') as fp:
        temp = pd.read_excel(fp, skiprows=skipRows)
    with open('../input/' + filename, 'rb') as fp:
        grid = pd.read_excel(fp, skiprows=skipRows, converters={i: str for i in range(len(temp.columns))})

    grid.columns = [col.lower() for col in grid.columns.str.strip()]
    return grid


def getStandardGridCols(gridCols, gridTemplateCols):
    stndrdOutCols = []
    for outCol in gridTemplateCols:
        outColFound = False
        for idx, inCol in enumerate(gridCols):
            if inCol.lower() in (col.lower() for col in outCol["psbInCols"]):
                outColFound = True
                gridCols.values[idx] = outCol[
                    "stndrdOutCol"]  # [outCol["stndrdOutCol"] if idx == idx2 else x for idx2, x in enumerate(gridCols)] #[idx] = outCol["stndrdOutCol"]
                stndrdOutCols.append(str(outCol["stndrdOutCol"]))
                break
        if not outColFound:
            raise Exception('Invalid Template')
    return gridCols, stndrdOutCols


def saveAsCsv(grid, type):
    try:
        elasticConfig = utils.readFile('../config/elasticConfig.json', type="json")["index_input_files"]
        grid.to_csv('../config/' + type + '.csv', sep=',', encoding='utf-8', index=False)
        indexName = elasticConfig[type]
        elastic.index_data_bulk(indexName, doc_type='my_type', recreate=1, csv_file_path='../config/',
                                csv_file_name=type)
    except:
        return False

    return True


def applyTransformations(grid, gridName):
    if gridName == "NDB_Taxonomy":
        grid = grid.loc[
            ((-grid['prov_type_name'].isin(['ALLIED HLT PROF GRP', 'PHYSICIAN GRP'])) & (-grid['ndb_deg'].isnull()))]
        grid['ndb_spec'] = grid['ndb_spec'].map(lambda x: zeroPadding(str(x), 3))
        grid['cos_prov'] = grid['cos_prov'].map(lambda x: zeroPadding(str(x), 4))
    return grid.fillna('')


def applyGridConfig(gridName):
    gridConfig = constants.standardsDict["gridConfigs"][gridName]
    gridStartIdx = gridConfig["headerStartIdx"]
    gridTemplateCols = gridConfig["cols"]

    grid = readGridFile(gridName, gridStartIdx)
    grid.columns, stndrdOutCols = getStandardGridCols(grid.columns, gridTemplateCols)
    grid = grid[stndrdOutCols]
    grid = applyTransformations(grid, gridName)
    return saveAsCsv(grid, gridName)


def updateFirstRow(excelPath):
    # file_format = excelPath.split(".")[1]
    filename, file_extension = os.path.splitext(excelPath)

    if file_extension == ".xls":
        workbook = xlrd.open_workbook(excelPath, formatting_info=True)
        wb = xlutilscopy(workbook)

        for idx, sheet_name in enumerate(workbook._sheet_names):
            sheet = workbook.sheet_by_name(sheet_name)
            row = sheet.row(0) if sheet.nrows else []
            isEmpty = True

            for i in range(len(row)):
                if row[0] == "":
                    break
                if row[i].value != "":
                    isEmpty = False
                    break

            if isEmpty == True:
                w_sheet = wb.get_sheet(idx)
                w_sheet.write(0, 0, "Empty")
                w_sheet.write(0, 1, "Empty")

        wb.save(excelPath)

    elif file_extension == ".xlsx":
        workbook = openpyxl.load_workbook(excelPath)

        for idx, sheet in enumerate(workbook.worksheets):
            for row in sheet.iter_rows():

                isEmpty = True

                for cell in row:
                    if not row[0].value:
                        break
                    if cell.value:
                        isEmpty = False
                        break

                if isEmpty == True:
                    sheet.cell(row=1, column=1).value = 'Empty'
                    sheet.cell(row=1, column=2).value = 'Empty'
                break

        workbook.save(excelPath)

def displayInputNew(fObj, excelJSON,sheetNames):
    dataDict = {}

    for idx, sheet in enumerate(sheetNames):
        df = pd.read_excel(fObj.getvalue(), sheet_name=idx, dtype=str).fillna('')
        headerIndex = excelJSON['tabDict'][idx]['startIndex'] - 1
        if len(df.index) > 10:
            df = df.iloc[:headerIndex + 10]
        lst = [[str(elm) for elm in df.columns]] + df.values.tolist()
        lst[headerIndex] = standardHeaderColumn(list(map(str, lst[headerIndex])))
        dataDict[sheet] = [[x.replace('\n', ' ') if idx != headerIndex else x for x in l] for idx, l in enumerate(lst)]
        # print json.dumps(dataDict)
    return dataDict


def displayInput(fileName, excelJSON):
    ExcelLocation = constants.loc + fileName
    dataDict = {}
    workbook = xlrd.open_workbook(ExcelLocation)

    updateFirstRow(ExcelLocation)

    for idx, sheet in enumerate(workbook._sheet_names):
        collen = len(pd.read_excel(open(ExcelLocation, "rb"), sheet_name=idx).columns)
        diConverter = {col: str for col in range(collen)}
        df = pd.read_excel(open(ExcelLocation, "rb"), sheet_name=idx, converters=diConverter).fillna('')
        headerIndex = excelJSON['tabDict'][idx]['startIndex'] - 1
        if len(df.index) > 10:
            df = df.iloc[:headerIndex + 10]
        lst = [[str(elm) for elm in df.columns]] + df.values.tolist()
        lst[headerIndex] = standardHeaderColumn(list(map(str, lst[headerIndex])))
        dataDict[sheet] = [[x.replace('\n', ' ') if idx != headerIndex else x for x in l] for idx, l in enumerate(lst)]
        # print json.dumps(dataDict)
    return dataDict


def provMapSv(provMapping, cols, mappings):
    # provMapping = readProvMapping(provName)
    provMappingSv = copy.deepcopy(provMapping)
    for k, v in list(provMappingSv.items()):
        if k not in cols:
            provMappingSv.pop(k)
    unmappedOPCols = [opCol for opCol in list(mappings.keys()) if opCol not in list(provMappingSv.values())]
    mappedOPColsDict = provMappingSv
    mappedOPColsList = list(provMappingSv.values())
    unmappedInpCols = [inpCol for inpCol in cols if inpCol not in list(provMappingSv.keys())]
    ioFeildMappings = [{"inputField": k, "outputField": v} for k, v in list(provMappingSv.items())]
    return mappedOPColsDict, mappedOPColsList, unmappedOPCols, unmappedInpCols, ioFeildMappings


def provMapMv(provMapping, cols, mvMappings):
    asIsAssoc = {}
    dictMVMappings = {}
    # provMapping = readProvMapping(provName)
    provMappingMv = copy.deepcopy(provMapping)
    for pcat, di in list(provMappingMv.items()):
        filterdMapping = {key: value for key, value in list(mvMappings.items()) if
                          value['PARENT_CATEGORY'] == pcat and value['DATAFRAME_TYPE'] == "Normalize"}
        asIsMapping = {key: value for key, value in list(mvMappings.items()) if
                       value['PARENT_CATEGORY'] == pcat and value['DATAFRAME_TYPE'] == "AsIs"}
        lstItems = []
        mvMappedOPColsDict = {}
        mvMappedOPCols = []
        for k, v in list(di.items()):
            if k not in cols:
                di.pop(k)
        for k, v in list(di.items()):
            opCol = v.split('@')[1].split('#')[0]
            if opCol in list(asIsMapping.keys()):
                if pcat not in asIsAssoc:
                    asIsAssoc.update({pcat: {}})
                asIsAssoc[pcat].update({k: v.split('@')[2]})
                mvMappedOPColsDict.update({k: opCol})
                mvMappedOPCols.append(opCol)
                di.pop(k)
        mvMappedOPColsDict.update(di)
        setIndices = list(set([map.split("@")[2] for map in list(di.values())]))
        lol = []
        for setIndex in sorted(setIndices):
            lol.append([map for map in list(di.values()) if map.split("@")[-1] == setIndex])
        mvMappedOPColsList = lol

        mvMappedOPCols += list({col.split('@')[1].split('#')[0] for col in list(di.values())})
        mvUnmappedOPCols = [col for col in list(filterdMapping.keys()) if col not in mvMappedOPCols]
        mvUnmappedInpCols = [col for col in cols if col not in list(mvMappedOPColsDict.keys())]
        dictMVMappings[pcat] = [mvMappedOPColsDict, mvMappedOPColsList, mvUnmappedOPCols, mvUnmappedInpCols]
    return dictMVMappings, asIsAssoc


def readProvMapping(provName):
    provData = {"Provider": provName}
    provMapping = constants.handler.find_one_document('Provider', provData)['Doc']
    return provMapping


def fastForward(inputDF, provName, token, filename, userID):
    # if not validateTemplate(sotHeader,provName):
    #     print 'not validated'
    #     return False
    provMapping = readProvMapping(provName)

    status, sheetsFailedValidation = validateTemplate(inputDF, provMapping)
    if not status:
        print('not validated')
        return False
    return uiDriverFunction(provMapping, token, filename, userID, ff=True)


def validateTemplate(inputDF, provMapping):
    # provMapping=readProvMapping(provName)

    sheetsFailedValidation = {}

    validationFailedFlag = False

    # new_sheets = []
    # removed_sheets = []

    # temp_InputDF = inputDF
    # temp_provMapping = provMapping
    #
    # sheetsInputDF = len(inputDF)
    # sheetsprovMapping = len(provMapping)

    # if sheetsInputDF != sheetsprovMapping:

    new_sheets = list(set(inputDF.keys()) - set(provMapping.keys()))
    removed_sheets = list(set(provMapping.keys()) - set(inputDF.keys()))

    for sheet in new_sheets:
        sheetsFailedValidation[sheet] = "new"
    for sheet in removed_sheets:
        sheetsFailedValidation[sheet] = "removed"

    if new_sheets or removed_sheets:
        validationFailedFlag = True

    for sheet, df in list(inputDF.items()):

        if sheet in new_sheets:
            continue

        sotHeader = df.columns
        provMappingSheet = provMapping[sheet]

        updatedFlag = False

        if isinstance(provMappingSheet, str) and provMappingSheet.lower() == "to be ignored":
            continue

        elif isinstance(provMappingSheet, dict):
            lstCols = [item for sublist in [list(v.keys()) for k, v in list(provMappingSheet['mv'].items())] for item in sublist] \
                      + list(provMappingSheet['sv'].keys())
            for col in lstCols:
                if col not in sotHeader:
                    sheetsFailedValidation[sheet] = "updated"
                    updatedFlag = True
                    validationFailedFlag = True
                    break

            if not updatedFlag:
                sheetsFailedValidation[sheet] = "unchanged"

    if validationFailedFlag:
        return False, sheetsFailedValidation
    return True, sheetsFailedValidation


def updateProviderStatus(updatedJson, diffScreenJson=None):
    if diffScreenJson:
        providerJson = constants.provAdmScrJson
        for provName in diffScreenJson:
            providerJson[provName]['status'] = 'existing'
        try:
            constants.handler.replace_one_document('providers', {'providers': providerJson})
        except Exception as e:
            print(e)
        else:
            setattr(constants, 'provAdmScrJson', providerJson)
        return

    for code, providers in list(updatedJson.items()):
        if providers['status'] == 'deleted':
            if constants.handler.db['Provider'].count_documents(utils.fetchProviderCondition(*(code,'ProviderFinal'))):
                constants.handler.delete_one_document('ProviderFinal', utils.fetchProviderCondition(*(code,'ProviderFinal')))
    for code, providers in list(updatedJson.items()):
        if providers['status'] == 'deleted':
            updatedJson.pop(code)
        elif providers['status'] == 'created':
            updatedJson[code]['status'] = 'new'

    try:
        constants.handler.replace_one_document('providers', {"providers": updatedJson})
    except Exception as e:
        print(e)
    else:
        setattr(constants, 'provAdmScrJson', updatedJson)


def diffFinalInterim(final, interim):
    inpColsSvFinal = list(final['sv'].keys())
    diffSv = {'diff': {}, 'additions': {}, 'deletions': {}}
    for inpCol, outCol in list(final['sv'].items()):
        if interim['sv'].get(inpCol):
            if outCol == interim['sv'][inpCol]:
                continue
            elif outCol != interim['sv'][inpCol]:
                diffSv['diff'].update({inpCol: outCol})
        elif interim['sv'].get(inpCol) is None:
            diffSv['deletions'].update({inpCol: outCol})
    for inpCol, outCol in list(interim['sv'].items()):
        if inpCol not in inpColsSvFinal:
            diffSv['additions'].update({inpCol: outCol})
    diffMv = {'diff': {}, 'additions': {}, 'deletions': {}}

    for pCat, di in list(final['mv'].items()):
        if pCat not in interim['mv']:
            diffMv['deletions'].update({pCat: di})
            continue
        for inpCol, outCol in list(di.items()):
            if interim['mv'][pCat].get(inpCol):
                if outCol.split('@')[0:2] == interim['mv'][pCat][inpCol].split('@')[0:2]:
                    continue
                elif outCol.split('@')[0:2] != interim['mv'][pCat][inpCol].split('@')[0:2]:
                    if pCat not in diffMv['diff']:
                        diffMv['diff'].update({pCat: {}})
                    diffMv['diff'][pCat].update({inpCol: interim['mv'][pCat][inpCol]})
            elif interim['mv'][pCat].get(inpCol) is None:
                if pCat not in diffMv['deletions']:
                    diffMv['deletions'].update({pCat: {}})
                diffMv['deletions'][pCat].update({inpCol: outCol})
    for pCat, di in list(interim['mv'].items()):
        if pCat not in final['mv']:
            diffMv['additions'].update({pCat: di})
            continue
        inpColsMvFinal = list(final['mv'][pCat].keys())
        for inpCol, outCol in list(di.items()):
            if inpCol not in inpColsMvFinal:
                if pCat not in diffMv['additions']:
                    diffMv['additions'].update({pCat: {}})
                diffMv['additions'][pCat].update({inpCol: outCol})
    #print('diffFinalInterim:', {'sv': diffSv, 'mv': diffMv})
    return {'sv': diffSv, 'mv': diffMv}


# def diffProvMapping():
#     diDiff={"new":{},"existing":{}}
#     provAdmScrJson=constants.provAdmScrJson
#     for provName,status in provAdmScrJson.items():
#         if status=='existing':
#             provFinalMapping=readProvMapping(provName)
#             provPath=constants.provMappingsPath+'/'+provName
#             lstProvUsrMapping=[filename for filename in next(os.walk(provPath))[2] if '.interim' in filename]
#             diffForProv={}
#             for userFileName in lstProvUsrMapping:
#                 userFilePath=provPath+'/'+userFileName
#                 userName=userFileName.split('.interim')[0]
#                 with open(userFilePath, 'rb') as fp:
#                     usrMap = json.load(fp)
#                 diff=diffFinalInterim(provFinalMapping,usrMap)
#                 diffForProv.update({userName:diff})
#             diDiff['existing'].update({provName:diffForProv})
#         if status=='new':
#             provPath = constants.provMappingsPath + '/' + provName
#             lstProvUsrMapping = [filename for filename in next(os.walk(provPath))[2] if '.interim' in filename]
#             diffForProv = {}
#             for userFileName in lstProvUsrMapping:
#                 userFilePath = provPath + '/' + userFileName
#                 userName = userFileName.split('.interim')[0]
#                 with open(userFilePath, 'rb') as fp:
#                     usrMap = json.load(fp)
#                 diffForProv.update({userName: usrMap})
#             diDiff['new'].update({provName: diffForProv})
#     return diDiff


def changedMapping(diffForSheets):
    change = False
    for sheet, sheetInfo in diffForSheets.items():
        if isinstance(sheetInfo, str) and sheetInfo.lower() == "to be ignored":
            continue
        if isinstance(sheetInfo, dict):
            for mapingType, details in sheetInfo.items():
                for changeType, changeDetatil in details.items():
                    if changeDetatil:
                        change = True
    return change


def removeEmpty(diDiff):
    diDiff_copy = deepcopy(diDiff)
    for delgStatus, details in diDiff.items():
        for delgName, delgDetails in details.items():
            if not delgDetails:
                del diDiff_copy[delgStatus][delgName]
    return diDiff_copy


def autoUpdateFinalFile(autoDiDiff):
    for delgCode, updateDetail in autoDiDiff.items():
        finalFile = constants.handler.find_one_document('ProviderFinal',utils.fetchProviderCondition(*(delgCode,'ProviderFinal')))['Doc']
        for format, timeStamp in updateDetail.items():
            if format != 0:
                finalFile[format]['timeStampUsed'] = str(sorted([parse(i) for i in timeStamp])[-1])
        replacedData = {"Type": "ProviderFinal", "Provider": delgCode, "Doc": finalFile}
        constants.handler.replace_one_document('ProviderFinal',replacedData)


def diffProvMapping():
    diDiff = {"new": {}, "existing": {}}
    provAdmScrJson = constants.provAdmScrJson
    autoDiDiff = {}

    for delegateCode, delegateDetails in list(provAdmScrJson.items()):
        status = delegateDetails['status']
        if status == 'existing':
            provFinalFile = readProvFinal(delegateCode)
            allProvFiles = constants.handler.find_documents('ProviderInterim', {"Provider": delegateCode})
            lstProvUsrMapping = [{file['UserName']: file['Doc']} for file in allProvFiles]
            diffForProv = {}
            noChangeFormatDetail = {}

            for file in lstProvUsrMapping:
                for userName, usrMapInterim in file.items():
                    diffForSheets = {}

                    new_format_sheets = [k for k, v in usrMapInterim.items() if v['formatExisting'] == 0]

                    for sheet in new_format_sheets:
                        if isinstance(usrMapInterim[sheet]['mappings'], str) and usrMapInterim[sheet][
                            'mappings'].lower() == "to be ignored":
                            diffForSheets[sheet] = "to be Ignored"
                            continue

                        diffForSheets[sheet] = {}
                        diffForSheets[sheet] = {}
                        diffForSheets[sheet]['sv'] = {}
                        diffForSheets[sheet]['sv']['deletions'] = {}
                        diffForSheets[sheet]['sv']['diff'] = {}
                        diffForSheets[sheet]['mv'] = {}
                        diffForSheets[sheet]['mv']['deletions'] = {}
                        diffForSheets[sheet]['mv']['diff'] = {}

                        diffForSheets[sheet]['sv']['additions'] = usrMapInterim[sheet]['mappings']['sv']
                        diffForSheets[sheet]['mv']['additions'] = usrMapInterim[sheet]['mappings']['mv']

                    for sheet, usrMapDetail in list(usrMapInterim.items()):
                        if sheet in new_format_sheets:
                            continue

                        if isinstance(usrMapDetail['mappings'], str) and usrMapDetail[
                            'mappings'].lower() == "to be ignored":
                            diffForSheets[sheet] = "to be Ignored"
                            continue

                        provMappingFormatTemp = provFinalFile[usrMapDetail['formatExisting']]

                        diff = diffFinalInterim(provMappingFormatTemp['mappings'], usrMapDetail['mappings'])
                        diffForSheets[sheet] = {}
                        diffForSheets[sheet] = diff

                    if changedMapping(diffForSheets):
                        diffForProv.update({userName: diffForSheets})
                    else:
                        for sheet, usrMapDetail in list(usrMapInterim.items()):
                            if usrMapDetail['formatExisting'] in list(noChangeFormatDetail.keys()):
                                noChangeFormatDetail[usrMapDetail['formatExisting']].append(
                                    usrMapDetail['timeStampUsed'])
                            else:
                                noChangeFormatDetail[usrMapDetail['formatExisting']] = [usrMapDetail['timeStampUsed']]

            diDiff['existing'].update({delegateCode: diffForProv})
            autoDiDiff.update({delegateCode: noChangeFormatDetail})
        if status == 'new':
            #provPath = constants.provMappingsPath + '/' + delegateCode
            #lstProvUsrMapping = [filename for filename in next(os.walk(provPath))[2] if '.interim' in filename]
            lstProvUsrMapping = constants.handler.find_documents('ProviderInterim',{"Provider":delegateCode})
            diffForProv = {}
            #for userFileName in lstProvUsrMapping:
            #    userFilePath = provPath + '/' + userFileName
            #    userName = userFileName.split('.interim')[0]
            #    with open(userFilePath, 'rb') as fp:
            #        usrMapInterim = json.load(fp)
            #    usrMap = {k: v['mappings'] for k, v in usrMapInterim.iteritems()}
            #    diffForProv.update({userName: usrMap})
            for data in lstProvUsrMapping:
                userName = data['UserName']
                usrMapInterim = data['Doc']
                usrMap = {k: v['mappings'] for k, v in usrMapInterim.items()}
                diffForProv.update({userName: usrMap})
            diDiff['new'].update({delegateCode: diffForProv})  # shift within loop

    diDiff = removeEmpty(diDiff)
    autoUpdateFinalFile(autoDiDiff)
    return diDiff


def readProvFinal(provName):
    provFinal = constants.handler.find_one_document('ProviderFinal', utils.fetchProviderCondition(*(provName,'ProviderFinal')))['Doc']
    return provFinal


def updateFile(provName, action, provStatus):
    if provStatus == "existing":
        finalFile = constants.handler.find_one_document('ProviderFinal', utils.fetchProviderCondition(*(provName,'ProviderFinal')))['Doc']
    else:
        finalFile = {}

    interimFile = constants.handler.find_one_document('ProviderInterim', {"Provider": provName, "UserName": action})[
        'Doc']

    keys_to_be_updated_from_interim = ['format', 'diheader', 'mappings', 'timeStampUsed']
    for sheet, sheetInfo in interimFile.items():
        if isinstance(sheetInfo['mappings'], str) and sheetInfo['mappings'].lower() == "to be ignored":
            continue
        if sheetInfo['formatExisting'] == 0:
            format_in_final = list(finalFile.keys())
            format_in_final = [int(i) for i in format_in_final]
            format_in_final.sort()
            if format_in_final:
                next_key = str(format_in_final[-1] + 1)
            else:
                next_key = str(1)
            finalFile[next_key] = {}
            for key in keys_to_be_updated_from_interim:
                finalFile[next_key][key] = sheetInfo[key]

        else:  # only applicable for existing provider
            for key, val in finalFile[sheetInfo['formatExisting']].items():
                if key in keys_to_be_updated_from_interim:
                    finalFile[sheetInfo['formatExisting']][key] = sheetInfo[key]

    replacewithData = {"Type": "ProviderFinal", "Provider": provName, "Doc": finalFile}
    constants.handler.replace_one_document('ProviderFinal', replacewithData)


def deleteFilesbyExtension(filemapping, provName):
    constants.handler.delete_many_documents(filemapping, {"Provider": provName})


def replaceFile(provName, action):
    srcFile = constants.handler.find_one_document('ProviderInterim', {'Provider': provName, 'UserName': action})['Doc']
    insertData = {'Type': 'ProviderInterim', 'Provider': provName, 'UserName': action, 'Doc': srcFile}
    constants.handler.insert_one_document('Bkp', insertData)


def commitAdminActionsProvMapping(adminActions):
    provToBeUpdated = []
    for provName, action in list(adminActions.items()):
        allInterim = constants.handler.find_documents('ProviderInterim', {"Provider": provName})
        users = [file['UserName'] for file in allInterim]
        if action == 'ignore':
            deleteFilesbyExtension('ProviderInterim', provName)
        elif action in users:
            provToBeUpdated.append(provName)
            provAdmScrJson = constants.provAdmScrJson
            provStatus = provAdmScrJson[provName]['status']
            updateFile(provName, action, provStatus)
            replaceFile(provName, action)
            deleteFilesbyExtension('ProviderInterim', provName)
    if provToBeUpdated:
        updateProviderStatus({}, diffScreenJson=provToBeUpdated)


def viewMapping(inpJson):
    inpJson["mv"] = formatLogicMv(inpJson["mv"])
    # for sheet,map in inpJson.items():
    #     if isinstance(map,dict):
    #         inpJson[sheet]["mv"] = formatLogicMv(map["mv"])
    return inpJson


def viewDiff(inpJson):
    for provider, di in list(inpJson['existing'].items()):
        for userName, sheets in list(di.items()):
            for sheet, maps in list(sheets.items()):
                if isinstance(maps, dict):
                    maps["mv"]["additions"] = formatLogicMv(maps["mv"]["additions"])
                    maps["mv"]["deletions"] = formatLogicMv(maps["mv"]["deletions"])
                    maps["mv"]["diff"] = formatLogicMv(maps["mv"]["diff"])
    for provider, di in list(inpJson['new'].items()):
        for userName, sheets in list(di.items()):
            for sheet, maps in list(sheets.items()):
                if isinstance(maps, dict):
                    maps["mv"] = formatLogicMv(maps["mv"])
    return inpJson


def formatLogicMv(mvJson):
    result = {}
    for pCat, di in list(mvJson.items()):
        result[pCat] = {}
        for inCol, outCol in list(di.items()):
            output = outCol.split('@')
            output[2] = int(output[2]) + 1
            if str(output[2]) + " " + output[0] not in result[pCat]:
                result[pCat][str(output[2]) + " " + output[0]] = {}
            result[pCat][str(output[2]) + " " + output[0]][output[1]] = inCol
    return result


def getDelegates(taxIDList):
    delegates_dict = constants.delegates_dict

    delegate_info = {}

    for tax_id in taxIDList:
        for delegate_code in list(delegates_dict.keys()):

            if tax_id in delegates_dict[delegate_code]["TAX ID"]:
                delegate_info["code"] = delegate_code
                delegate_info["name"] = delegates_dict[delegate_code]["Delegate Name"]
                delegate_info["status"] = constants.provAdmScrJson[delegate_code]["status"]

                return delegate_info

    return delegate_info


def standardHeaderColumn(headerRow):
    head = []
    for val in headerRow:
        val = val.replace("\n", "").strip()
        head.append(val)

    return head


def getFinalDfPsi(masterDictList):
    fdf = pd.DataFrame()
    for masterDict in masterDictList:
        colsToBeRemovedOthers = ['index', 'ROW_COUNT', 'ROW_NUM', 'level_0']
        colsToBeRemovedFD = []
        colsToBeRemovedSV = ['TABNAME', 'FILENAME']
        colsToBeRemovedBrd = colsToBeRemovedOthers + ['degForLkp']
        masterDict = {k: v for k, v in list(masterDict.items()) if
                      k in ['FinalDegree', 'singleValue', 'Address', 'Speciality', 'ChangeType']}
        dfs = [v.drop(colsToBeRemovedFD, axis=1) if k == 'FinalDegree'
               else v.drop(colsToBeRemovedSV, axis=1) if k == 'singleValue'
        else v.drop(colsToBeRemovedBrd, axis=1, errors='ignore') if k == 'Board'
        else v.drop(colsToBeRemovedOthers, axis=1, errors='ignore') for k, v in list(masterDict.items())]
        record = pd.concat(dfs, axis=1)
        record = record.sort_index(axis=1)
        fdf = fdf._append(record)
    return fdf


def getTaxIDColumn(workbook, sheetIndex, header_row):
    sheet = workbook.sheet_by_index(sheetIndex)
    num_columns = sheet.ncols
    tax_id_mappings = constants.mappings["TAX_ID"]["Input_Column"]

    for column in range(num_columns):
        column_name = str(sheet.cell(header_row, column).value)
        temp_column = column_name

        temp_column = standardHeaderColumn([temp_column])[0]

        column_name = column_name.split("\n")[0].lower().strip()

        if column_name in tax_id_mappings:
            return temp_column

    return


def esExportExcelToJSON(workbook):
    """returns JSON object for workbook meta data"""

    # get the sheetnames of workbook as List
    sheetnames = es.getWorkbookSheetNames(workbook)

    # excelJSON
    excelJSON = {}
    excelJSON['sheetcount'] = es.getWorkbookSheetCount(workbook)

    excelJSONTabsDictionary = {}
    excelJSON["tabDict"] = []

    for index, sheet in enumerate(sheetnames):
        # for sheet in sheetnames :
        # print sheet
        excelJSONTabsDictionary = {}
        excelJSONTabsDictionary['sheetname'] = sheet
        excelJSONTabsDictionary['RowMinimumColumnCount'] = es.getRowMinimumColumnCount(workbook, index)
        excelJSONTabsDictionary['maximumColumnCount'] = es.getRowMaximumColumnCount(workbook, index)
        # excelJSONTabsDictionary['startIndex'] = getIndexRealValues(workbook, index)
        header_row = es.getHeaderRowIndex(workbook, index)
        excelJSONTabsDictionary['startIndex'] = header_row + 1
        excelJSONTabsDictionary['TaxIDColumnName'] = getTaxIDColumn(workbook, index, header_row)
        # print(excelJSONTabsDictionary)
        excelJSONTabsDictionary['data']=[[cl.value for cl in row] for row in workbook.sheet_by_index(index).get_rows()]
        excelJSON["tabDict"].append(excelJSONTabsDictionary)

    return excelJSON


def determineTemplate(action):
    for k, v in list(constants.templates.items()):
        if action in v['Screenname']:
            return k


def filterVar(templateName, by='LayerName', val=None):
    for layer in tDrivers[templateName]['var']:
        if layer[by] == val:
            return layer['Transformation']

def collectHashElms(df,hashCols):
    df=df.sort_values(by=hashCols)
    lstHashElm=[]
    tmp=[tuple(arr) for arr in df[hashCols].values]
    lstHashElm.extend([elm for tup in tmp for elm in tup])
    return lstHashElm

def createHash(hashElms):
    return hashlib.md5(b"".join(hashElms)).hexdigest()

def storeActions(df):
    templateNm = 'actions'
    mvFile = tDrivers[templateNm]['mv']
    masterDict = {"data": df}
    cmnLayer = filterVar(templateNm, val='common')
    masterDict = cleanseDF(mvFile, masterDict, cmnLayer)
    lstDi=masterDict['data'].fillna('').to_dict(orient='records')
    return lstDi


def predictActions(df):
    templateNm = 'actions'
    mvFile = tDrivers[templateNm]['mv']
    masterDict={"data":df}
    cmnLayer=filterVar(templateNm, val='common')
    masterDict=cleanseDF(mvFile,masterDict,cmnLayer)
    predictLayer=filterVar(templateNm, val='predict')
    masterDict = cleanseDF(mvFile, masterDict, predictLayer)
    #print(masterDict["data"])


def compileActions(userMapInterim, token, filename,mongo,fs):
    st=perf_counter()
    userMapping = {k: v['mappings'] for k, v in userMapInterim.items()}

    def predict(diDFs, userMapInterim):
        if constants.useNlpModel:
            actionFile = classify(filename)
            return dataDictionary(diDFs, actionFile)
        else:
            diHeaders = {k: v['diheader'] for k, v in userMapInterim.items()}
            predictions = {}
            for tab, tab_info in list(userMapInterim.items()):
                mapping = tab_info['mappings']
                if isinstance(mapping, str) and mapping.lower() == "to be ignored":
                    continue
                elif isinstance(mapping, dict):
                    inputDF = diDFs[tab]
                    sotHeader = diHeaders[tab]
                    validatedActionColumn = str(constants.validatedActionColNm, "utf-8")
                    dataframe, diColMapping, dictMVUserMapping = renameDfCols(inputDF,mapping,tab,filename,validatedActionColumn)

                    dataframe = predictActions(dataframe)
                    predictions[tab] = [
                        {"colValues": [], "rowIndex": idx, "TransactionType": row[constants.predictedActionColNm]} for
                        idx, row in dataframe.reset_index().iterrows()]
            return predictions

    def transactionTypes():

        completeTransactionTypes = ["DEFAULT"]

        templatesDict = constants.templates

        for key, value in templatesDict.items():
            if value['status'].lower()=='hidden':
                continue
            completeTransactionTypes.extend(value['Screenname'])

        return completeTransactionTypes

    def dataDictionary(diDFs, actionFile):

        data = {}

        for tab, mapping in list(userMapping.items()):

            if isinstance(mapping, str) and mapping.lower() == "to be ignored":
                continue
            actionColumnNames=list(mapping['mv'].get('ChangeType',{}).keys())
            npi_col_name = ''
            for k,v in mapping['sv'].items():
                if v == "INDIVIDUAL_NPI":
                    npi_col_name+=k
                    actionColumnNames.append(k)
            inputDF = diDFs[tab].fillna("")

            actionTab = classify(tab)

            lstRecords=classifyData(inputDF, actionColumnNames, actionTab, actionFile,npi_col_name)
            data[tab] = lstRecords

        return data

    inputJson = {}
    fNm, fileExt = os.path.splitext(filename)
    if fileExt =='.pdf':
        diDFs = readPickle(token)
    else:
        diDFs = diDFsFromMongo(token,mongo,fs)
    inputJson["filename"] = filename
    inputJson["AllTransactionTypes"] = transactionTypes()
    inputJson["data"] =predict(diDFs, userMapInterim)
    logPerf(st,'compileActions')
    return inputJson


def getOutputJson(fdf):
    dataKey = "data"
    dataJson = {dataKey: None}
    dataJson.update({dataKey: dfToJson(fdf.fillna(''), toJson=0)})
    dataJson[dataKey].update({'colSet': list(fdf.columns)})
    dataJson[dataKey].update({'order': 0})
    return dataJson


def getfdfOutCol(fdf, outputTemplate,delType):
    outputJson = deepcopy(outputTemplate)
    # print 'outputJson:',outputJson
    dataJson = {}

    colSet = []
    dataSetCols = []
    for di in outputJson['finalDfColNames']:
        if di['dfColName']:
            if di.get('outputScreen','Both')=='Both':
                colSet.append(di['outputColName'])
                dataSetCols.append(di['dfColName'])
                continue
            elif delType=='delegate' and di['outputScreen']==delType:
                colSet.append(di['outputColName'])
                dataSetCols.append(di['dfColName'])
                continue
            elif delType=='nonDelegate' and di['outputScreen']==delType:
                colSet.append(di['outputColName'])
                dataSetCols.append(di['dfColName'])

    # if fdf.empty:
    #     fdf = pd.DataFrame([{k: "" for k in dataSetCols}])
    dataSetCols = dataSetCols + constants.uidLevels
    colSet = colSet + constants.uidLevels
    fdfOutCol = fdf[dataSetCols].reset_index(drop=True)
    fdfOutCol.columns = colSet
    # fdfOutCol = pd.concat([fdfOutCol, fdf[constants.recIdLevels].reset_index(drop=True)], axis=1)
    # fdfUID = fdf[constants.uidLevels + constants.recIdLevels].reset_index(drop=True).drop_duplicates(
    #     subset=constants.recIdLevels)
    # fdfOutCol = fdfOutCol.merge(fdfUID, on=constants.recIdLevels).drop(constants.recIdLevels, axis=1)
    return fdfOutCol


def prependNonMass(fdfOutColDi):
    if not fdfOutColDi:
        return {}
    idxForRowRemoval = []
    outDi = {}
    nmTName = constants.nmTName

    nmCoreCols = filterOutput(nmTName, val="Prac", returnKey="outputColName") + constants.uidLevels  # todo
    flag=0
    for template, fdfOutCol in list(fdfOutColDi.items()):
        if template == nmTName:
            continue
        if fdfOutCol.empty:
            continue
        fdfOutCol = fdfOutColDi[nmTName][nmCoreCols].merge(fdfOutCol, on=constants.uidLevels, how='right')  # todo
        # fdfOutCol = fdfOutColDi[nmTName].merge(fdfOutCol, on=constants.uidLevels, how='right')
        # fdfOutCol = fdfOutCol.reset_index()
        if flag==0:
            fdfOutCol['Prac_PrimaryTin'] = pd.Series(["TRUE" if x == 0 else "FALSE" for x in fdfOutCol.reset_index().index])
            flag=1
        elif flag==1:
            fdfOutCol['Prac_PrimaryTin'] = "FALSE"

        if constants.runInterTmpltLayer and template == str('demo-addr-change'):
            fdfOutCol = fdfOutCol.sort_values(constants.uidLevels)
            fdfOutColMI = fdfOutCol.set_index(constants.uidLevels)
        else:
            fdfOutColMI = fdfOutCol.set_index(constants.uidLevels)
        
        outDi.update({template: fdfOutColMI})
        idxForRowRemoval.extend(list(fdfOutColMI.index.unique()))
    nmFdfMI = fdfOutColDi[nmTName].set_index(constants.uidLevels)
    # if idxForRowRemoval:
    #     nmFdfMI = nmFdfMI.loc[~(nmFdfMI.index.isin(idxForRowRemoval))]
    outDi[nmTName] = nmFdfMI
    return outDi
    
def hilightCmnTTyp(fdfNMDi):
    cmnColNm="Common Across "+",".join(constants.cmnTTyp)
    if not all([tTyp in list(fdfNMDi.keys()) for tTyp in constants.cmnTTyp]):
        for tTyp in constants.cmnTTyp:
            if tTyp in fdfNMDi:
                fdfNMDi[tTyp][cmnColNm] = "False"
        return fdfNMDi
    for tTyp in constants.cmnTTyp:
        fdfNMDi[tTyp].set_index(constants.commonUidLevels,drop=False,inplace=True)
    lstIx = [fdfNMDi[tTyp].index.unique() for tTyp in constants.cmnTTyp]
    cmnIx = lstIx[0]
    for ix in lstIx:
        cmnIx = cmnIx.intersection(ix)
    for tTyp in constants.cmnTTyp:
        fdfNMDi[tTyp][cmnColNm] = "False"
        if len(cmnIx)>0:
            fdfNMDi[tTyp].loc[cmnIx,cmnColNm]="True"
        fdfNMDi[tTyp]=fdfNMDi[tTyp].reset_index(drop=True)
    return fdfNMDi

def ffillColumns(df, cols):
    i = -1
    diVal = {}
    for idx, row in df.iterrows():
        i += 1
        if any(row.fillna('')):
            diVal.update(row[col])
        for col in cols:
            df.iloc[i, df.columns.get_loc(col)] = diVal[col]
    return df


def filterOutput(templateName, by='inclusionLevel', val=None, returnKey=None):
    filterResult = []
    for colDi in tDrivers[templateName]['output']['finalDfColNames']:
        if colDi[by] == val:
            if returnKey:
                filterResult.append(colDi[returnKey])
            else:
                filterResult.append(colDi)
    return filterResult


def updateDiWithGroups(di, lstMappedCols):
    def getGroup(k, maps):
        for _map in maps:
            if _map['outputField'] == k:
                return _map['groupName']

    di = copy.deepcopy(di)
    groupTracker = {}
    for tup in lstMappedCols:
        splt = tup[1].split('#')
        if len(splt) == 1:
            continue
        outCol = splt[0]
        grpOrder = int(splt[1])
        grpName = getGroup(outCol, di['map'])
        if grpName not in groupTracker:
            groupTracker.update({grpName: grpOrder})
        else:
            if grpOrder > groupTracker[grpName]:
                groupTracker[grpName] = grpOrder

    for grpName, counter in list(groupTracker.items()):
        for i in range(counter + 1):
            if i == 0:
                lstGrpElm = []
                for map in di['map']:
                    if map.get('groupName') == grpName:
                        lstGrpElm.append(copy.deepcopy(map))
                        map['groupOrder'] = i
            else:
                for elm in lstGrpElm:
                    elm['groupOrder'] = i
                di['map'].extend(copy.deepcopy(lstGrpElm))
    return di


def calculateMatchPer(cleaned_header, format_):
    """func description"""
    if np.all(cleaned_header == format_):
        return 100
    else:
        common = [i for i in cleaned_header if i in format_]
        if len(cleaned_header) > len(format_):
            match = int(math.floor((float(len(common)) / len(cleaned_header)) * 90))
        else:
            match = int(math.floor((float(len(common)) / len(format_)) * 90))
        if match < constants.mappingMatchThreshold:
            match = 0
        return match


def compareProvAndGlobDict(provScoreDict, globScoreDict, provMapping, globalMapping):
    scoreDict, bestMatch_entry, mappingUsed, prov_using_glob = None, None, None, True

    if globScoreDict:
        prov_using_glob = True
        # bestMatch_entry_glob = max(globScoreDict.iteritems(), key=operator.itemgetter(1))[0]
        bestMatch_entry_glob = maxPercentSortedKeys(globScoreDict)
        scoreDict = globScoreDict
        bestMatch_entry = bestMatch_entry_glob
        mappingUsed = globalMapping

    if provScoreDict:
        # bestMatch_entry_prov = max(provScoreDict.iteritems(), key=operator.itemgetter(1))[0]
        bestMatch_entry_prov = maxPercentSortedKeys(provScoreDict)
        if provScoreDict[bestMatch_entry_prov] > constants.mappingMatchThreshold:
            scoreDict = provScoreDict
            bestMatch_entry = bestMatch_entry_prov
            mappingUsed = provMapping
            prov_using_glob = False
        else:
            if globScoreDict:
                if provScoreDict[bestMatch_entry_prov] >= globScoreDict[bestMatch_entry_glob]:
                    scoreDict = provScoreDict
                    bestMatch_entry = bestMatch_entry_prov
                    mappingUsed = provMapping
                    prov_using_glob = False
            else:
                scoreDict = provScoreDict
                bestMatch_entry = bestMatch_entry_prov
                mappingUsed = provMapping
                prov_using_glob = False

    return scoreDict, bestMatch_entry, mappingUsed, prov_using_glob


def maxPercentSortedKeys(inpdict):
    bestMatch_entry_score, bestMatch_entry = -1, -1
    keys = [int(k) for k in inpdict.keys()]
    keys.sort()
    for k in keys:
        if inpdict[str(k)] > bestMatch_entry_score:
            bestMatch_entry_score = inpdict[str(k)]
            bestMatch_entry = str(k)
    return bestMatch_entry


def findMappingwithMatch(cleaned_header, globalMapping, provMapping, delegateName):
    # sheet wise function
    timeStampUsed = str(datetime.datetime.now())
    provScoreDict = {}
    globScoreDict = {}

    if provMapping:
        for entry, entryDetail in provMapping.items():
            provScoreDict[entry] = calculateMatchPer(cleaned_header, entryDetail["format"])
        if provScoreDict:
            bestMatch_entry_prov = maxPercentSortedKeys(provScoreDict)
            # bestMatch_entry_prov = max(provScoreDict.iteritems(), key=operator.itemgetter(1))[0]
            if provScoreDict[bestMatch_entry_prov] < constants.mappingMatchThreshold:
                if globalMapping:
                    for entry, entryDetail in globalMapping.items():
                        globScoreDict[entry] = calculateMatchPer(cleaned_header, entryDetail["format"])
    elif globalMapping:
        for entry, entryDetail in globalMapping.items():
            globScoreDict[entry] = calculateMatchPer(cleaned_header, entryDetail["format"])

    scoreDict, bestMatch_entry, mappingUsed, prov_using_glob = compareProvAndGlobDict(provScoreDict, globScoreDict,
                                                                                      provMapping, globalMapping)

    if mappingUsed:
        format_existing = bestMatch_entry
        if delegateName and prov_using_glob:
            format_existing = 0

        if scoreDict[bestMatch_entry] == 0:
            return 0, {}, scoreDict[bestMatch_entry], timeStampUsed
        elif scoreDict[bestMatch_entry] >= constants.mappingMatchThreshold and scoreDict[bestMatch_entry] < 100:
            return 0, mappingUsed[bestMatch_entry]["mappings"], scoreDict[bestMatch_entry], timeStampUsed

        return format_existing, mappingUsed[bestMatch_entry]["mappings"], scoreDict[bestMatch_entry], timeStampUsed
    else:
        return 0, {}, 0, timeStampUsed


def getMappingwithMatch(sheet_with_header_detail, delegateName, delegateStatus):
    """Return the format which best match the input file cleaned header
    return the formatNumber, mapping, matchingpercentage and timeStamp"""

    provMapping, globalMapping = None, None
    globalMapping_data = utils.fetchProviderCondition(*(constants.globalDelegateName,'ProviderFinal'))
    provMapping_data = utils.fetchProviderCondition(*(delegateName,'ProviderFinal'))
    if constants.handler.db['Provider'].count_documents(globalMapping_data):
        globalMapping = constants.handler.find_one_document('ProviderFinal', globalMapping_data)['Doc']
    if constants.handler.db['Provider'].count_documents(provMapping_data):
        provMapping = constants.handler.find_one_document('ProviderFinal', provMapping_data)['Doc']

    for sheet in list(sheet_with_header_detail.keys()):
        sheet_with_header_detail[sheet]['formatExisting'], sheet_with_header_detail[sheet]['mappings'], \
        sheet_with_header_detail[sheet]['matchPer'], sheet_with_header_detail[sheet]['timeStampUsed'] = \
            findMappingwithMatch(sheet_with_header_detail[sheet]['format'], globalMapping, provMapping, delegateName)

    return sheet_with_header_detail


def deriveMVCols(mvFile, masterDict, multivaluedCols):
    # derivemvCols = [masterDict, midlevNdBSpec]

    for obj in multivaluedCols:
        for key, value in obj.items():
            mulValColumn = key

            # Create empty DF if not exists
            if not mulValColumn in masterDict:
                masterDict[mulValColumn] = pd.DataFrame()

            df = masterDict[mulValColumn]

            for index, row in df.iterrows():
                row = row.fillna('')
                for funcIdx, func in enumerate(value["derivations"]):

                    funcName = func["name"]
                    funcInp = func["input"]
                    col = func["col"]

                    if (not func.get("type")) or func.get("type") == "var":
                        funcInp = [row] + [catch(param, masterDict) for param in funcInp]
                        try:
                            if col and col != "DYNAMIC":
                                df.loc[index, col] = None
                                row[col] = getattr(mvFile, funcName)(*funcInp)
                            else:
                                col, data = getattr(mvFile, funcName)(*funcInp)
                                if isinstance(col, list) and isinstance(data, list):
                                    diValues = dict(list(zip(col, data)))
                                    for k, v in list(diValues.items()):
                                        df.loc[index, k] = None
                                        row[k] = v
                                else:
                                    df.loc[index, col] = None
                                    row[col] = data
                        except Exception as e:
                            row[col] = utils.ErrorMessage([], ["E", "Undefined Error", ""])
                            # print "NPI and TAX IDs are "+row["FINAL_NPI"]+", "+row["TAX_ID"]
                            traceback.print_exc()
                            continue
                    if func["type"] != "var":
                        # df2 = pd.DataFrame(row).T
                        # funcInp = eval( ", ".join(funcInp))
                        funcInp = [catch(param, masterDict) for param in func["input"]]
                        df.loc[index, col] = None
                        try:
                            if func["type"] == "tree":
                                row[col] = str(getattr(diImport[func["type"]], funcName)().populate([row] + funcInp))
                            else:
                                row[col] = getattr(diImport[func["type"]], funcName)(row, *funcInp)
                        except Exception as e:
                            row[col] = utils.ErrorMessage([], ["E", "Undefined Error", ""])
                            # print "NPI and TAX IDs are " + row["FINAL_NPI"] + ", " + row["TAX_ID"]
                            traceback.print_exc()
                            continue
                try:
                    df.loc[index] = list(row)
                except Exception as e:
                    # print "NPI and TAX IDs are " + row["FINAL_NPI"] + ", " + row["TAX_ID"]
                    traceback.print_exc()
                    continue

            # updating current df in master dict
            masterDict[mulValColumn] = df
    return masterDict


def catch(tup, masterDict):
    try:
        return eval(tup)
    except Exception as e:
        return tup


def cleanseDF(mvFile, masterDict, CleanseDataFrame):
    # derivemvCols = [masterDict, midlevNdBSpec]
    
    for obj in CleanseDataFrame:
        for key, value in obj.items():
            if not value["derivations"]:
                continue
            mulValColumn = key
            df = masterDict[key]
            for funcIdx, func in enumerate(value["derivations"]):
                funcName = func["name"]
                funcInp = func["input"] if not isinstance(func["input"], str) else eval(func["input"])
                outputCol = func.get("col")
                if isinstance(funcInp, dict) and funcInp.get('diDf'):
                    diDf = funcInp['diDf']
                    diDf = {pcat: copy.deepcopy(masterDict[pcat]) for pcat in list(diDf.keys())}
                    funcInp['diDf'] = diDf
                try:
                    if not func.get("type") or func.get("type") == 'var':
                        df = getattr(mvFile, funcName)(df, funcInp)
                    else:
                        df = getattr(diImport[func["type"]], funcName)(df, funcInp)
                except:
                    traceback.print_exc()
                    continue

            # Input Variable names
            # df = Function that works on Data frame and returns a updated dataframe back

            # updating current df in master dict
            masterDict[mulValColumn] = df
    return masterDict

def interTemplateDerivation(fdfDict, layer):

    if not layer:
        return fdfDict
    for funcIdx, func in enumerate(layer):
        outTemplate=func["outTemplate"]
        funcName = func["name"]
        funcInp = func["input"] if not isinstance(func["input"], str) else eval(func["input"])
        outTemplateOutCols = {di['outputColName']:di['dfColName'] for di in tDrivers[outTemplate]['output']['finalDfColNames']}
        try:
            if not func.get("type") or func.get("type") == 'var':
                fdfDict=getattr(tDrivers[outTemplate]['mv'], funcName)(outTemplate,outTemplateOutCols,fdfDict, funcInp)
            else:
                fdfDict=getattr(diImport[func["type"]], funcName)(outTemplate,outTemplateOutCols,fdfDict, funcInp)
        except:
            traceback.print_exc()
            continue
    return fdfDict


def remapTType(ttype):
    for key, val in list(constants.standardsDict['TTypeRemap'].items()):
        if ttype in val:
            return key
    return ttype


def adminSignout(role,name):
    if role=="admin":
        credentialsJson = constants.credentialsData
        credentialsJson[name]['logStatus'] = 'Inactive'
        replaceData = {'credentials': credentialsJson}
        try:
            constants.handler.replace_one_document('credentials', replaceData)
        except Exception as e:
            print(e)
        else:
            setattr(constants, "credentialsData", credentialsJson)


def fetchPlmiData(delCode):
    plmiData = {'delId': None, 'globalData': None, 'delData': None}
    if not constants.fetchPlmi:
        return plmiData
    def getDelId(delCode, plmiGlobalData):
        #print("------------------------getDelID method----------------",lstDel)
        lstDel = plmiGlobalData['delegates']['DelegateList']
        #print("------------------------getDelID method lstDel----------------",lstDel)
        lstDelId = [diDel["DelegateId"] for diDel in lstDel if diDel["DelegateCode"] == delCode]
        #print("===================================getDelID method lstDelId=================================================",lstDelId)
        #print("++++++++++++++++++++++++++++++++++++++getDelID method return++++++++++++++++++++++++++++++++",lstDelId[0] if lstDelId else None)
        return lstDelId[0] if lstDelId else None

    plmiGlobalData = {k: utils.fetchResponse(url) for k, url in list(constants.plmiGlobalData.items())}
    plmiDelData=constants.plmiDelData
    # delCode='NY007'
    delId=getDelId(delCode,plmiGlobalData)
    if delId:
        plmiDelData = {k: utils.fetchResponse(di["url"], dict(list(zip(list(di["params"].keys()), (delId,))))) for k, di in
                       list(plmiDelData.items())}
        
    else:
        plmiDelData=None
    plmiData={'delId':delId,'globalData':plmiGlobalData,'delData':plmiDelData}
    return plmiData


def plmiWrapper(plmiData):
    #import pdb;pdb.set_trace()
    delId=plmiData['delId']
    template = 'plmi'
    plmiLayer = filterVar(template, val=template)
    plmiOutputCols = tDrivers[template]['output']
    outCols = [di['outputColName'] for di in plmiOutputCols['finalDfColNames']]
    plmiDf = utils.createOneRowDf(outCols)
    #if delId is None:
        #plmiDf=plmiDf.applymap(lambda cell:utils.ErrorMessage([], ["C", "Delegate Not Found", ""]))
        #return dict(plmiDf[outCols].iloc[0])
    plmiDf["globalData"]=str(plmiData['globalData'])
    plmiDf["delData"]=str(plmiData['delData'])
    plmiDf["delId"]=delId
    diPlmiDf={"singleValue": plmiDf}
    diPlmiDf=cleanseDF(tDrivers[template]['mv'],diPlmiDf,plmiLayer)
    return dict(diPlmiDf['singleValue'][outCols].iloc[0])

def assignPlmiData(fdfNMDi,plmiData):
    for template,df in list(fdfNMDi.items()):
        for k,v in list(plmiData.items()):
            if k in df.columns:
                df[k]=v
    return fdfNMDi

def getDerivedTemplateData(dataframe,hiddenTempDI,delType):
    def appendParentTmpRows(dataframe,hiddenTempDI,ParentTemplate):
        for newTempTtyp,lstMasterTemplateTTyp in hiddenTempDI.items():
            for masterTmpTTyp in lstMasterTemplateTTyp[ParentTemplate]:
                newTempDF = dataframe[dataframe[constants.validatedActionColNm] == masterTmpTTyp]
                newTempDF[constants.validatedActionColNm] = lstMasterTemplateTTyp['Screenname'][0]
                dataframe = dataframe._append(newTempDF).reset_index(drop=True)
        return dataframe
    if delType=='delegate':
        return appendParentTmpRows(dataframe,hiddenTempDI,'DelParentTemplates')
    return appendParentTmpRows(dataframe,hiddenTempDI,'NonDelParentTemplates')

def fileTypeDriver(token,prevData,userMapping,fileType):
    if fileType == 'pdf':
        userMapping = mimicUserInterimFile(token)
        prevData['mappings'] = userMapping
    elif fileType == 'email':
        pass
    elif fileType == 'nonDelPdf':
        createNonDelDfPkl(userMapping,token)
        userMapping = mimicUserInterimFile(token)
        prevData['mappings'] = userMapping
    # else:
    #     prevData['ff'] = False
    return prevData, userMapping

def mimicUserInterimFile(token=None, diDFs=None):
    diDFs = diDFs if diDFs else readPickle(token)
    userInterimFile={}
    for tab,df in list(diDFs.items()):
        mappings={"sv":{},"mv":{}}
        cols={elm:elm.split('@')[1].split('#')[0] if len(elm.split('@')) == 3 else elm for elm in df.columns}
        for renamedCol,stdCol in list(cols.items()):
            if stdCol in constants.mappings:
                mappings["sv"].update({renamedCol: renamedCol})
            elif stdCol in constants.mvMappings:
                pCat=constants.mvMappings[stdCol]['PARENT_CATEGORY']
                mappings["mv"].setdefault(pCat,{}).update({renamedCol:renamedCol})
        userInterimFile[tab]={"mappings":mappings,'diheader':list(df.columns)}
    return userInterimFile


def int_keys(obj):
    for key in list(obj.keys()):
        new_key = key
        try:
            new_key = int(key)
        except ValueError:
            pass
        if new_key != key:
            obj[new_key] = obj[key]
            del obj[key]
    return obj

def createNonDelDfPkl(jsonData, token):
    result = []
    for recordDi in jsonData:
        recordResDi = {}
        for pCat,pCatDi in recordDi['mv'].items():
            recordResDi.update({k:v[0] for k,v in pCatDi.items()})
        recordResDi.update({k: v[0] for k, v in recordDi['sv'].items()})
        result.append(recordResDi)
    df = pd.DataFrame(result)
    df.fillna('', inplace=True)
    diDFs = {"pdfTab": df}
    dfname = token + "InpDF" + '.pkl'
    with open("../tmp/" + dfname, 'wb') as handle:
        pickle.dump(diDFs, handle)
    return df

def callPdfToHtmlApi(token, fNm):
    fNm = constants.inFileLoc + fNm
    infilePath = '{}.pdf'.format(fNm)
    tkn = token + '_{}'.format(fNm)
    files = {'pdfFile': open(infilePath, 'rb')}
    hashVal = hashlib.sha256(tkn.encode('utf-8')).hexdigest()
    print((tkn, '\t', tkn.encode('utf-8'), '\t', hash(tkn), '\t', hashVal))
    params = {'tkn': hashVal}
    res = requests.post(url=constants.pdf2HtmlApi, files=files, data=params)
    resJsn = res.json()
    return resJsn['html']

def toTrain(jsonData):
    '''
    decides whether to put data for training or not
    :param jsonData:
             json structure: [{"sv":{"colName":["valueMapped", [], "filename", "PDF/HTML"]},"mv":{}},{"sv":{},"mv":{}}]
    :return:
         fileTrain : dict with filename as key and bool flag as value signifying to train or not
                    (True implies it has to stored for training purpose)
         fileWiseData : dict with filename as key and its corresponding data as value
    '''
    # gather filename wise data
    fileWiseData = {}
    fileTrain = {}
    for recDi in jsonData:
        recFileWiseData = {}
        recFileTrain = {}
        for tag, valLst in recDi['sv'].items():
            dataGrp = recFileWiseData.setdefault(valLst[2],{'sv':{},'mv':{}})
            recFileTrain.setdefault(valLst[2], True)
            dataGrp['sv'][tag] = valLst[:2]
            if valLst[3] == 'PDF':
                recFileTrain[valLst[2]] = False
        for pCat, pCatDi in recDi['mv'].items():
            for tag, valLst in pCatDi.items():
                dataGrp = recFileWiseData.setdefault(valLst[2], {'sv': {}, 'mv': {}})
                recFileTrain.setdefault(valLst[2], True)
                pCatGrp = dataGrp['mv'].setdefault(pCat,{})
                pCatGrp[tag] = valLst[:2]
                if valLst[3] == 'PDF':
                    recFileTrain[valLst[2]] = False
        for eachFile, data in recFileWiseData.items():
            filegrp = fileWiseData.setdefault(eachFile,[])
            filegrp.append(data)    # adding each rec data in file's array
            fileTrain.setdefault(eachFile,True)
            if fileTrain[eachFile]: # updating fileTrain only if it is True
                fileTrain[eachFile] = recFileTrain[eachFile]
    return fileTrain, fileWiseData

def createLogDf():
    logs = constants.handler.find_documents('logging')
    baseLogCols = ['token', 'filename', 'username', 'fileType', 'businessType', 'delegateCode', 'processCompleteFlag',
                   'nTabs', 'nTabsProcessed', 'startProcessTimestamp', 'businessRuleStartTimestamp',
                   'businessRuleEndTimestamp', 'totalFileProcessingTime']
    outLogCols = ['token', 'filename', 'template', 'rowCount', 'nUniqueIds']
    tabLogCols = ['token', 'filename', 'tabName', 'originalMatchPercent', 'remappedFlag', 'tabProcessed', 'rowCount',
                  'nUniqueIds']
    transationLogCols = ['token', 'filename', 'tabName', 'transactionType', 'rowCount']
    outLog = [dict(di, token=log['token'], filename=log['filename']) for log in logs if 'outputTemplatesInfo' in log for di in log['outputTemplatesInfo']]
    tabLog = [dict(di, token=log['token'], filename=log['filename']) for log in logs if 'tabInfo' in log for di in log['tabInfo']]
    transationLog = [dict(di, token=log['token'], filename=log['filename'], tabName=log['tabName']) for log in tabLog if 'transactionInfo' in log for di in log['transactionInfo']]
    baseLogDf = pd.DataFrame(logs, columns=baseLogCols)
    outLogDf = pd.DataFrame(outLog, columns=outLogCols)
    baseLogDf['date'] = [x.split()[0] for x in baseLogDf['startProcessTimestamp']]
    tabLogDf = pd.DataFrame(tabLog, columns=tabLogCols)
    transationLogDf = pd.DataFrame(transationLog, columns=transationLogCols)

    mergeDF = pd.merge(baseLogDf, outLogDf, on='token')
    colSet = ['date', 'startProcessTimestamp', 'businessRuleEndTimestamp', 'totalFileProcessingTime', 'template',
              'nUniqueIds','username','prRowCount','rowCount']
    renameCols = {'date': 'Date', "startProcessTimestamp": "Input Time", "businessRuleEndTimestamp": "Output Time",
                  "totalFileProcessingTime": "Provider Output", "template": "Output Transaction Type",
                  "nUniqueIds": "Unique NPI Count","prRowCount":"PR Row Count (Without Header)","rowCount":"Output Row Count (Without Header)"}
    mergeDF['prRowCount'] = ''
    df2 = transationLogDf[['token','transactionType','rowCount']]
    mergeDF['prRowCount'] = df2.groupby(['token', 'transactionType'], sort=False)["rowCount"].transform('sum')
    for ix, row in mergeDF.iterrows():
        mergeDF['totalFileProcessingTime'][ix] = mergeDF['totalFileProcessingTime'][ix].split('.')[0]
    uiDF = mergeDF[colSet]
    uiDF.rename(columns=renameCols, inplace=True)
    uiDF.fillna('', inplace=True)
    data  = dfToJson(uiDF)
    return data

def retrieveFileFromMongo(filename,db,fs,saveLoc=None,multiple=False,collectby=None,zipflag=None):
    st = perf_counter()
    logDb('fetching from GridFS with params filename:{}'.format(filename))
    lst=[]
    ##filelist = list(db.fs.files.find({"filename": filename}))
    ## zipflag is used excusively for the zip downloading, each time unique zip files downloading.
    if zipflag == 1:
        filelist = list(db.fs.files.find({"filename": filename}).sort([("uploadDate", -1)]))
    else:
        filelist = list(db.fs.files.find({"filename": filename}))
    if not filelist:
        print('fetching from GridFS using the condition',{"filename": filename})
    for fl in filelist:
        file_Id = fl['_id']
        fobj = fs.get(file_Id)
        if multiple:
            if collectby:
                lst.append((fl[collectby],fobj.read()))
            else:
                lst.append(fobj.read())
        else:
            if saveLoc:
                with open(saveLoc + filename, 'wb') as f:
                    f.write(fobj.read())
                    return
            else:
                return fobj
    logPerf(st,'retrieveFileFromMongo')
    return lst

def saveExlFileSnapshot(fNmUniq,fileObj):
    excelPath=constants.loc + fNmUniq
    limit=constants.nrowsForSnapshot
    filename, file_extension = os.path.splitext(excelPath)
    fObj = io.BytesIO()
    if file_extension == ".xlsx":
        if constants.snapshotInplace:
            st = perf_counter()
            workbook = openpyxl.load_workbook(fileObj)
            logPerf(st,'openpyxl.load_workbook')
            st = perf_counter()
            for idx, sheet in enumerate(workbook.worksheets):
                sheet.delete_rows(limit+1,sheet.max_row)
            workbook.save(fObj)
            logPerf(st,'deleting rows')
        else:
            st = perf_counter()
            workbookRd = openpyxl.load_workbook(fileObj,read_only=True)
            workbookWt = openpyxl.Workbook(write_only=True)
            for nm in workbookRd.sheetnames:
                workbookWt.create_sheet(nm)
            for wsRd in workbookRd.worksheets:
                wsWt=workbookWt[wsRd.title]
                for rx,r in enumerate(wsRd.iter_rows()):
                    if rx>limit:
                        break
                    wsWt.append([cel.value for cel in r])
            workbookWt.save(fObj)
            workbookRd.close()
            logPerf(st,'copying rows')
    elif file_extension=='.xls':
        st = perf_counter()
        workbookRd = xlrd.open_workbook(file_contents=fileObj.read(), on_demand=True)
        logPerf(st, 'xlrd.open_workbook')
        st = perf_counter()
        workbookWt = openpyxl.Workbook(write_only=True)
        sheetNames=workbookRd.sheet_names()
        for nm in sheetNames:
            workbookWt.create_sheet(nm)
        for sheet_name in sheetNames:
            wsRd = workbookRd.sheet_by_name(sheet_name)
            wsWt = workbookWt[sheet_name]
            for rx, r in enumerate(wsRd.get_rows()):
                if rx > limit:
                    break
                wsWt.append([xlrd.xldate.xldate_as_datetime(cel.value,workbookRd.datemode) if cel.ctype==3 else cel.value for cel in r])
        workbookWt.save(fObj)
        logPerf(st,'copying rows')
    return fObj

def saveToGridFS(mongo,fNm,fObj,typ,col='fs',**kwargs):
    logDb('uploading to GridFS with params filename:{} and objType:{} and kwargs:{}'.format(fNm,typ,kwargs))
    st=perf_counter()
    mongo.save_file(fNm, fObj,base=col,objType=typ,**kwargs)
    logPerf(st,'uploading to GridFS')

def classifyRecord(row,highLvlAction,npi_col_name):
    record = {"colValues": [], 'TransactionType': highLvlAction, 'rowIndex': '',"provider_npi":row.get(npi_col_name)}
    # record = {"colValues": [], 'TransactionType': highLvlAction, 'rowIndex': ''}
    # lst=list(row)
    lst = list(row.drop(npi_col_name)) if row.get(npi_col_name) else list(row)
    record['colValues']=lst
    actionCol = classify(''.join(lst).strip())
    if actionCol:
        record["TransactionType"] = actionCol
    record["rowIndex"] = int(row.name)
    record["TransactionType"] = list(map(remapTType, record["TransactionType"]))
    return record

def classifyData(inputDF,actionColumnNames,actionTab,actionFile,npi_col_name):
    if actionTab:
        highLvlAction=actionTab
    else:
        highLvlAction = actionFile
    if actionColumnNames:
        dfSbst = inputDF[actionColumnNames].reset_index(drop=True)
        lstRecords =classifyRecordWrapper(dfSbst,highLvlAction,npi_col_name)
    else:
        lstRecords = [{"colValues": [], 'TransactionType': list(map(remapTType, highLvlAction)), 'rowIndex': ix} for ix in range(len(inputDF))]
    return lstRecords

def classifyRecordWrapper(dfSbst,highLvlAction,npi_col_name,lazy=True):
    lstRecords=[]
    if lazy:
        for ix,row in dfSbst.iterrows():
            lstRecords.append(classifyRecord(row, highLvlAction,npi_col_name))
    else:
        lstRecords=list(dfSbst.apply(lambda row: classifyRecord(row, highLvlAction,npi_col_name), axis=1))
    return lstRecords


def getCsv(fObj,sheetIx,headerIx,headerLen):

    wb = openpyxl.load_workbook(fObj, read_only=True,data_only=True, keep_links=False)
    sheetNm=wb.sheetnames[sheetIx]
    ws = wb[sheetNm]
    lst = list(ws.iter_rows(values_only=True, max_col=headerLen))[headerIx:]
    wb.close()
    lstLen = len(lst)
    for ix in range(lstLen):
        if any(lst[(lstLen - 1) - ix]):
            break
        else:
            lst.pop()

    fp = io.StringIO()
    csv_out = csv.writer(fp)
    csv_out.writerows(lst)
    fp.seek(0)
    return fp


def parallel_worksheet(tup):
    st = perf_counter()
    fNmUniq=tup[0]
    sheetIx=tup[1]
    headerIx=tup[2]
    headerLen=tup[3]
    # print(fNmUniq,sheetDi)
    conn = MongoClient(host=constants.mongoConHost,port=constants.mongoConPort)
    db=conn.Triage
    fs=GridFS(db)
    fObj=retrieveFileFromMongo(fNmUniq,db,fs)
    fp=getCsv(fObj,sheetIx,headerIx,headerLen)
    df = pd.read_csv(fp, dtype=str, na_values=["na", "n/a", "None", "none", "Na"])
    # print('length of df',len(df))
    # print('columns of df',list(df.columns))
    logPerf(st,'parallel_worksheet for sheet: {0}'.format(sheetIx))
    conn.close()
    return (sheetIx,df)

def createDiDFs(sheets,fNmUniq,fObj):
    fNm,fExt=os.path.splitext(fNmUniq)
    st = perf_counter()
    mp=False
    if len(sheets)>1 and fExt=='.xlsx':
        wb = openpyxl.load_workbook(fObj, read_only=True,data_only=True, keep_links=False)
        wSheets=wb.worksheets
        mp=sum([ws.max_row>constants.nRowsMp and ws.max_column>constants.nColsMp for ws in wSheets])>1
        wb.close()
    args=[(fNmUniq,ix,getXlrdHeaderIx(di['startIndex']),len(di['data'][getXlrdHeaderIx(di['startIndex'])])) for ix,di in enumerate(sheets) if di['data']]
    if mp:
        # with ProcessPoolExecutor(min(len(sheets),cpu_count())) as pool:
        with ProcessPoolExecutor(min(len(sheets),constants.windows_max_cpu)) as pool:
            res = pool.map(parallel_worksheet, args)
        diDFs=dict(list(res))
    else:
        lstFNm,lstSheetIx,lstHeaderIx,lstHeadersLen=map(list,zip(*args))
        if len(set(lstHeaderIx))==1:
            diDFs=pd.read_excel(fObj,sheet_name=lstSheetIx,header=lstHeaderIx[0],dtype=str,na_values=["na", "n/a", "None", "none", "Na"])
        else:
            diDFs = pd.read_excel(fObj, header=None,sheet_name=lstSheetIx,dtype=str,na_values=["na", "n/a", "None", "none", "Na"])
            newDi={}
            for fNm,sheetIx,headerIx,headerLen in args:
                tdf=diDFs[sheetIx].iloc[headerIx+1:]
                tdf.columns=diDFs[sheetIx].iloc[headerIx]
                newDi[sheetIx]=tdf
            diDFs=newDi

    for fNm, sheetIx, headerIx, headerLen in args:
        nCols = len(diDFs[sheetIx].columns)
        newHeaderLen = min(nCols, headerLen)
        diDFs[sheetIx] = diDFs[sheetIx].iloc[:, :newHeaderLen]

    logPerf(st,'createDiDFs')
    return diDFs

def getXlrdHeaderIx(headerIx):
    return int(headerIx - 1) if headerIx != 0 else int(headerIx)

def logPerf(st,msg,ovrRide=False):
    if constants.perfLogging or ovrRide:
        print('Time taken in {0}: {1:.2f}s'.format(msg,(perf_counter()-st)))
def logDb(msg,ovrRide=False):
    if constants.dbLogging or ovrRide:
        print(msg)
def getMvCols(mvMapCols,frameCols):
    mvCols=[]
    for col in frameCols:
        colnmList = col.split("@")
        if len(colnmList) > 1:
            if colnmList[1].split('#')[0] in mvMapCols:
                mvCols.append(col)
        elif col in mvMapCols:
            mvCols.append(col)
    return mvCols




def fetchUrl(job):
    ix=job[0]
    req=job[1]
    npiRecord=[ix,'']
    try:
        res = requests.get(**req)
        npiRecord[1]=json.loads(res.content)
    except:
        #traceback.print_exc()
        pass
    return npiRecord

def fetchAsyncThread(func,lstJobs,factor=1):
    lstRes=[]
    # defaultWorkers=(cpu_count()*5)
    defaultWorkers=(constants.windows_max_cpu*5)
    workers=defaultWorkers*factor
    with concurrent.futures.ThreadPoolExecutor(workers) as executor:
        futures=[]
        for job in lstJobs:
            futures.append(executor.submit(func, job))
        for future in concurrent.futures.as_completed(futures):
            lstRes.append(future.result())
    return lstRes

def fillNpiColData(npiRecord, npiregDicts):
    if "results" in npiRecord and npiRecord["results"]:
        npiregDict = npiregDicts[npiRecord["results"][0]["enumeration_type"]]
        basicInfo = npiRecord["results"][0]["basic"]
        return json.dumps({key: (basicInfo[npiregDict[key]] if npiregDict[key] in basicInfo else "") for key in list(npiregDict.keys())}), "1",str(npiRecord)
    else:
        return json.dumps(npiregDicts["NPI"]), "0",''

def o_getNwork(lenArgs):
    initBatSize = constants.initBatSize
    nWork = max(1, lenArgs // initBatSize)
    # nWork = min(nWork, cpu_count())
    nWork = min(nWork, constants.windows_max_cpu)
    return nWork
    
def getNwork(lenArgs):
    initBatSize = constants.initBatSize
    nWork = max(1, lenArgs // initBatSize)
    # nWork = min(nWork, cpu_count())
    if lenArgs > 500:
        constants.windows_max_cpu = 30        
    nWork = min(nWork, constants.windows_max_cpu)
    return nWork

def getBatchSizeMPW(lenArgs,nWork,initBatSize):

    if lenArgs<initBatSize:
        batchSize=initBatSize
        nBatches=1
    else:
        quo,rem=lenArgs//initBatSize,lenArgs%initBatSize
        quo1,rem1=quo//nWork,quo%nWork
        quo2,rem2=((rem1*initBatSize)+rem)//(nWork*quo1),((rem1*initBatSize)+rem)%(nWork*quo1)
        batchSize=initBatSize+quo2
        nBatches=nWork*quo1
    return nBatches,batchSize


def getBatches(argsMPW,nBatches,batchSize):
    required=nBatches*batchSize
    leftOver=argsMPW[required:]
    lenLeftOver=len(leftOver)
    batches = [argsMPW[i:i + batchSize] for i in range(0, required, batchSize)]
    for ix, elm in enumerate(leftOver):
        batches[-lenLeftOver+ix].append(elm)

    return batches

def getMasterDictList(argsMPW):
    masterDictList=parallelProc(multiProcWorkWrapper,argsMPW)
    masterDictList.sort(key=lambda x: (x['sheetIdx'], x['recordIdx']))
    return masterDictList

def parallelProc(func,argsMPW,initBatSize=constants.initBatSize,toChunk=True,threshMPW=constants.threshMPW):
    st=perf_counter()
    lenArgs = len(argsMPW)
    nWork = getNwork(lenArgs)
    if nWork==1 or lenArgs<threshMPW:
        lstRes=func(argsMPW)
    else:
        nBatches, batchSize = getBatchSizeMPW(lenArgs, nWork,initBatSize)
        if toChunk:
            batches = getBatches(argsMPW, nBatches, batchSize)
        else:
            batches=argsMPW
        print('lenArgs:({}) batchSize:({}) nWork:({}) number of batches:({})'.format(lenArgs, batchSize, nWork,len(batches)))
        with ProcessPoolExecutor(int(nWork/2)) as pool:
            res = pool.map(func, batches)
            lstRes = list(res)
        if toChunk:
            lstRes = [elm for lst in lstRes for elm in lst]
    logPerf(st, func.__name__, ovrRide=True)
    return lstRes

def npiLookupAsync(lstNPI, argsMPW):
    argDi = {"inpCol": "FINAL_NPI", "npiregDict": {
        "NPI-1": {"last_name": "last_name", "first_name": "first_name", "middle_name": "middle_name",
                  "credential": "credential"},
        "NPI-2": {"last_name": "authorized_official_last_name", "first_name": "authorized_official_first_name",
                  "middle_name": "authorized_official_middle_name", "credential": "authorized_official_credential"},
        "NPI": {"last_name": "", "first_name": "", "middle_name": "", "credential": ""}
    }, "npiParseFlag": "0", "output": "npilookup", "npiData": "npiData", "maxTries": 3}
    npiParseFlagCol = "npiParseFlag"
    npiregDicts = argDi["npiregDict"]
    npiCol = argDi["inpCol"]
    npiRegDict = npiregDicts["NPI"]
    npiParseFlag = argDi[npiParseFlagCol]
    maxTries = argDi["maxTries"]
    outCol = argDi["output"]
    npiData = argDi["npiData"]

    st=perf_counter()
    print('Requesting NPI Registry...')
    # lstRes = fetchAsyncThread(fetchUrl,lstJobs,factor=5)
    diRes,npiRegErr=npiLookupAsyncIO(lstNPI,npiregDicts)
    print('Requesting NPI Registry...Done | Time Taken:({:.2f}s)'.format(perf_counter() - st))
    for ixArg,arg in enumerate(argsMPW):
        frmTup=arg[0]
        frmNm=frmTup[0]
        frm=frmTup[1]
        npi=frmNm[0]
        valTup=diRes[npi]
        diVal=dict(zip((outCol,npiParseFlagCol,npiData),valTup))
        arg[-1].update(diVal)

    if npiRegErr == len(lstNPI):
        print('NPI Registry Response: Error For All NPIs')
    elif npiRegErr > 0 and npiRegErr < len(lstNPI):
        print('NPI Registry Response: Error For {} NPIs'.format(npiRegErr))
    elif npiRegErr == 0:
        print('NPI Registry Response: No Error')
    return argsMPW


def reGroupTTyp(argsMPW):
    di = {}
    # creates dictionary with id (combinatio of npi,taxid) as key and value being list of all different ttypes appearing for the key
    for arg in argsMPW:
        frmTup=arg[0]
        frmNm=frmTup[0]
        idTup, inpTType=(frmNm[0],frmNm[1]),frmNm[-1]
        di.setdefault(idTup, []).append(inpTType)

    #loops over each tuple
    for arg in argsMPW:
        diSv={}
        frmTup = arg[0]
        frmNm = frmTup[0]
        idTup, inpTType = (frmNm[0], frmNm[1]), frmNm[-1]
        #loops over output ttypes
        for outTTyp,lstLabels in list(constants.reGroupTTyp.items()):
            if outTTyp in di[idTup]:
                # if output ttype is present then looks for the presence of input ttype in the
                # list of candidate ttypes corresponding to the output ttype
                if inpTType in lstLabels:
                    # if found then updates ttype which is the last element of the tuple with the output ttype
                    frmNm[-1] = outTTyp
                    # and updates the validated action column with the value of output ttype
                    diSv[constants.validatedActionColNm]=outTTyp
                    break

        arg.append(diSv)
    return argsMPW

async def fetchUrlAIO(session, job):

    ix = job[0]
    req = job[1]
    npiRecord = [ix, '']
    try:
        url=req['url']
        async with session.get(url) as response:
            resp = await response.read()
            npiRecord[1] = json.loads(resp)
    except:
        traceback.print_exc()
        pass
    return npiRecord


async def fetchAsyncAIO(loop,lstJobs):
    # please use url by your choice
    url = "*** some image url ****"
    tasks = []
    # try to use one client session
    async with ClientSession(connector=aiohttp.TCPConnector(ssl=None, verify_ssl=False)) as session:
        for job in lstJobs:
            task = asyncio.ensure_future(fetchUrlAIO(session, job))
            tasks.append(task)
        # await response outside the for loop
        responses = await asyncio.gather(*tasks)
    return responses

def fetchAsyncAIOWrapper(lstJobs):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    future = asyncio.ensure_future(fetchAsyncAIO(loop, lstJobs))
    loop.run_until_complete(future)
    responses = future.result()
    return responses

def npiLookupAsyncIO(lstNPI,npiregDicts):
    lstJobs = [(npi, {"url": constants.npiAPI + npi}) for npi in lstNPI]
    lstRes = fetchAsyncAIOWrapper(lstJobs)
    diRes = {}
    npiRegErr = 0
    for res in lstRes:
        npi = res[0]
        npiRecord = res[1]
        if npiRecord == '':
            npiRegErr += 1
        diRes[npi] = fillNpiColData(npiRecord, npiregDicts)
    return diRes,npiRegErr
