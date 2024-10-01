# -*- coding: utf-8 -*-

# import packages and libraries
import openpyxl
import numpy as np
import warnings
import json
import xlrd
import re
flag=1
# import utils
warnings.simplefilter("ignore")


# load the workwook from the path
def loadWorkBook(path):
    """loads the excel file in memory"""
    # workbook = openpyxl.load_workbook(path, use_iterators=True)
    workbook=xlrd.open_workbook(path)
    # ws = workbook.active
    # for index, row in enumerate(ws):
    # # for index,row in workbook.iter_rows():
    #     for cell in row:
    #         print cell
    return workbook


# get the sheetnames of the workbook
def getWorkbookSheetNames(workbook):
    """returns the list of sheets in the workbook"""
    return workbook.sheet_names()


# get the specific sheet of the workbook by name
def getWorkbookSheetByName(workbook, sheetname):
    """returns the specific sheet by name"""
    return workbook.sheet_by_name(sheetname)


# get the total number of sheets in the workbook
def getWorkbookSheetCount(workbook):
    """returns the total number of sheets in the workbook"""
    return workbook.nsheets


def getIndexRealValues(workbook, index):
    sheet = workbook.sheet_by_index(index)
    rowIndex = 1
    for rowidx in range(sheet.nrows):
        row = sheet.row(rowidx)
        rowIndex += 1
        for colidx, cell in enumerate(row):
            # print 'cell.ctype:', cell.ctype
            try:
                if (cell.ctype == 2 and cell.value is not None) or (str(cell.value).strip()!='' and len(re.sub('[^0-9]','',str(cell.value).strip()))==len(str(cell.value).strip())):
                    return rowIndex
            except:
                continue
    return rowIndex



# def getIndexIntegerValues(workbook):
#     ws = workbook.active
#     for index, row in enumerate(ws):
#         for cell in row:
#             print cell
#             if cell.value_is_int():
#                 return index


def getRowMaximumColumnCount(workbook, index):
    sheet = workbook.sheet_by_index(index)
    maxCols = 0
    for row in sheet.get_rows():
        col_count = 0
        for columns in row:
            if columns.value is None or (str(columns.value).strip()==''):
                continue
            col_count += 1
        if maxCols < col_count:
            maxCols = col_count
    return maxCols


def getRowMinimumColumnCount(workbook, index):
    sheet = workbook.sheet_by_index(index)
    minCols = 0
    first=1
    for row in sheet.get_rows():
        col_count = 0
        for columns in row:
            if columns.value is None or (str(columns.value).strip()==''):
                continue
            col_count += 1
        if first:
            minCols=col_count
            first=0
        if minCols > col_count:
            minCols = col_count
    return minCols




def getHeaderRowIndex(workbook, sheet_index):
    sheet = workbook.sheet_by_index(sheet_index)

    num_columns = sheet.ncols
    num_rows = sheet.nrows

    header_row = 0
    for row in range(num_rows):
        headers = [str(cell.value) for cell in sheet.row(row) if str(cell.value) != '']
        count = len(headers)

        if count == num_columns:
            header_row = row
            break
    else:
        return getIndexRealValues(workbook,sheet_index)-2
    return header_row


def read_delegates_file(delegateFilePath):
    workbook = xlrd.open_workbook(delegateFilePath)
    sheet = workbook.sheet_by_index(0)

    delegates_dict = {}

    num_rows = sheet.nrows

    for row_num in range(1, num_rows):
        row = sheet.row(row_num)

        delegate_code = row[0].value.strip()
        delegate_name = row[1].value.strip()
        tax_id = int(row[3].value.strip())

        if delegate_code not in delegates_dict:
            delegates_dict[delegate_code] = {}
            delegates_dict[delegate_code]["Delegate Name"] = delegate_name
            delegates_dict[delegate_code]["TAX ID"] = []
            delegates_dict[delegate_code]["TAX ID"].append(tax_id)

        else:
            delegates_dict[delegate_code]["TAX ID"].append(tax_id)

    return delegates_dict


# def getAllTaxIDs(workbook, excelJSON):
#
#     taxIDs = []
#
#     sheet_count = excelJSON['sheetcount']
#     sheets_dict = excelJSON['tabDict']
#
#     for sheet in range(1, sheet_count):
#         sheet_name = sheets_dict[sheet-1]['sheetname']
#         taxIdColName = sheets_dict[sheet-1]['TaxIDColumnName']
#         header_row = sheets_dict[sheet-1]['startIndex'] - 1
#
#         sheet = workbook.sheet_by_name(sheet_name)
#
#         num_columns = sheet.ncols
#         num_rows = sheet.nrows
#
#         for column in range(num_columns):
#             column_name = sheet.cell(header_row, column).value
#
#             if column_name == taxIdColName:
#                 column_values = sheet.col_values(column, header_row + 1)
#                 column_values = list(set(column_values))
#                 column_values = list(filter(None, column_values))
#                 column_values = [value.replace("-", "") for value in column_values]
#                 column_values = map(int, column_values)
#
#                 taxIDs.extend(column_values)
#
#     return taxIDs


def open_file(path):
    """
    Open and read an Excel file
    """
    book = xlrd.open_workbook(path)

    # print number of sheets
    print(book.nsheets)

    # print sheet names
    print(book.sheet_names())

    # get the first worksheet
    first_sheet = book.sheet_by_index(0)

    # read a row
    print(first_sheet.row_values(0))

    # read a cell
    cell = first_sheet.cell(0, 0)
    print(cell)
    print(cell.value)

    # read a row slice
    print(first_sheet.row_slice(rowx=0,
                                start_colx=0,
                                end_colx=2))


if __name__=='__main__':
    path='output/input-2.xlsx'
    workbook=loadWorkBook(path)
    # print getWorkbookSheetNames(workbook)
    # print getWorkbookSheetByName(workbook,'Sheet1')
    # print getWorkbookSheetCount(workbook)
    # print getIndexRealValues(workbook,0)
    # print getRowMaximumColumnCount(workbook,0)
    # print exportExcelToJSON(workbook)