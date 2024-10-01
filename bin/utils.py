import json
import pandas as pd
import openpyxl
import numpy as np
import xlrd
import traceback
import copy
import re
import requests
# import columnTrees as ct
import nlpMapping as nlp
class NestedDict(dict):
    def __getitem__(self, key):
        if key in self: return self.get(key)
        return self.setdefault(key, NestedDict())



def convertToStringDF(inputDF):
    inputDF = inputDF.fillna('')
    if (not inputDF.empty):
        cols = inputDF.columns
        for key in cols:
            inputDF[key] = inputDF[key].values.astype('unicode')
    return inputDF


def dropEmptyRows(df):
    # inputDF = df.replace(r'\s+', np.nan, regex=True)
    df = df.replace(r'^\s+$', np.nan, regex = True).replace(r'^\s+|\s+$','',regex=True).replace('', np.nan)
    return df.dropna(how='all')

def getOnlyHeaders(path, sheetIndex, rowNum):
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(sheetIndex)

    return worksheet.row(rowNum)

def renameDuplicateColumns(df):
    # uncomment for version 0.23 support
    # df.columns = pd.io.parsers.ParserBase({'names': df.columns})._maybe_dedup_names(df.columns)
    cols = list(df.columns)
    cols2 = cols
    duplicateColumns = df.columns[df.columns.duplicated()].unique()
    for dup in duplicateColumns:
        count = 0
        for idx,c in enumerate(cols):
            if c == dup:
                if count >0:
                    cols2[idx] = dup + "." +str(count)
                count += 1
    df.columns = cols
    return df

def readFile(path, type, sheetIndex = 0, headerIndx = 0,colAsStr=False): #Todo: include read word file functionality.
    if headerIndx != 0:
        headerIndx = headerIndx - 1

    if type == "xlsx":
        if colAsStr:
            inputFile = pd.read_excel(path, sheet_name=sheetIndex, skiprows=headerIndx,dtype=str,  na_values = ["na", "n/a", "None", "none", "Na"])
        #Todo: convert columns to lower case :Done
            inputFile.columns = [col.strip().lower() for col in map(str,inputFile.columns)]  # remove spaces from column names
        else:
            inputFile = pd.read_excel(path, sheet_name= sheetIndex, skiprows = headerIndx)

        inputFile = renameDuplicateColumns(inputFile)

        # if headerIndx and dataStrtIndx:
        #     headers = range(0, dataStrtIndx)
        #     headers.remove(headerIndx)
        #     inputFile.drop(headers)
        inputFile = convertToStringDF(inputFile)
        inputFile = dropEmptyRows(inputFile)
        return inputFile
    elif type == "json":
        with open(path) as json_data:
            configFile = json.load(json_data)
            return configFile
    else:
        raise AssertionError("Unsupported file format!", type)

def addMissingColumns(df, cols):
    for col in cols:
        df[col] = ""
    return df

def getFinalNpi(i, g):
    if i:
        return i, "I"
    else:
        if not g:
            return "", "rejected"
        return g, "G"

def fillFinalNPIs(df, col):
    def fillFinalNPI(val):
        finalNpi, type = getFinalNpi(*val.split('pi3.141'))
        return finalNpi + 'pi3.141' + type

    df[[col, "NPI_TYPE_IND"]] = (df["INDIVIDUAL_NPI"] + 'pi3.141' + df["GROUP_NPI"]).apply(lambda val: fillFinalNPI(val)).str.split('pi3.141', expand=True)
    return df

def getUniqueIDs(df):
    uniqueFinalNpis = df["FINAL_NPI"].unique()
    uniqueIds = []
    for npi in uniqueFinalNpis:
        npiDF = df[df["FINAL_NPI"] == npi]
        taxIds = npiDF.TAX_ID.unique()
        for taxId in taxIds:
            uniqueIds.append( (npi, taxId) )
    return uniqueIds

def renameColumns(inputDF, columns):
    # print "inside renameColumns"
    # inputDF.columns = inputDF.columns.str.lower()#converting inputDF columns to lowercase.
    inputDF = inputDF.rename(columns = columns) #renaming columns with filtered columns.
    return inputDF

def mapColumns(inputDFcols, svmappings): #ToDo: append unsued config file columns, store list of unused columns in input file.
    # print "inside checkColumns"
    mappedOPColsDict = {} #renameColumns expects dictionary
    mappedOPColsList = []
    unmappedOPCols = [] #missing columns in the input file after checking with columns.json
    ioFeildMappings = []
    newInpCols = inputDFcols.tolist()
    for key, value in list(svmappings.items()):
        possibleColumns = svmappings[key]['Input_Column']
        for i, column in enumerate(possibleColumns):
            column = column.lower()
            if column in newInpCols:#TOdo: remove output
                mappedOPColsDict[column] = key
                mappedOPColsList.append(key)
                try:
                    newInpCols.remove(column.lower())
                except:
                    print(column.lower(),"repeated twice in mappings.json")
                    traceback.print_exc()
                else:
                    ioFeildMappings.append({"inputField":column,"outputField":key})
                    break
            if i == len(possibleColumns) - 1:
                unmappedOPCols.append(key)
    return mappedOPColsDict, mappedOPColsList, unmappedOPCols, newInpCols, ioFeildMappings

def getFrames(df, uniqueIds):
    """returns list of dataframes for every unique id"""
    frames = []
    for uniqueId in uniqueIds:

        tdf = df[(df["FINAL_NPI"] == uniqueId[0]) & (df["TAX_ID"] == uniqueId[1])]
        frames.append(tdf)
    return frames

# def splitFrames(frames, columns):
#     splittedFrames = []
#     rejectedFrames = []
#     for frame in frames:
#         rowDF = pd.DataFrame()
#         for key, value in columns.items():
#             if columns[key]["DATAFRAME_TYPE"] == "Single value":
#                 if len(frame[key].unique()) > 1:
#                     try:
#                         raise AssertionError("<Error: >", key, "has multiple values")
#                     except Exception as e:
#                         print "<Multi Value Error: TaxId- " + str(frame.TAX_ID.unique()) + " NPI- " + str(frame.FINAL_NPI.unique()) + "has multiple values for column-" + key
#                         rejectedFrames.append( str(frame.TAX_ID.unique()) )
#                 else:
#                     rowDF[key] = frame[key].unique()
#             # else
#         # print frame["FINAL_NPI"]
#         # rowDF["FINAL_NPI"] = list(frame["FINAL_NPI"])[0]#TODO: check with rohan about assaigning FINAL_NPI
#         # rowDF["NPI_TYPE_IND"] = list(frame["NPI_TYPE_IND"])[0]#TODO: check with rohan about assaigning FINAL_NPI
#         splittedFrames.append(rowDF)
#     return splittedFrames, rejectedFrames

def merge_two_dicts(x, y):
    """Given two dicts, merge them into a new dict as a shallow copy."""
    z = x.copy()
    z.update(y)
    return z

# def fillColumnTrees(cols):
#     for key, val in cols.items():
#         funcName = val["VARIABLE_NAME"]
#         funcDef = val["Tree"]
#         setattr(ct, funcName, funcDef)


""" Multivalued definitions"""
# def checkAddressColumns(input_columns, address_json):
#     mv_mappedOPColsDict = {}  # renameColumns expects dictionary
#     missingColumns = []  # missing columns in the input file after checking with columns.json
#     identifiedColumns = []
#     mv_mappedOPColsList = []
#     mv_mapped_op_col_list=[]
#     addset = []
#     count = 0
#     flag = False
#     for col in input_columns:
#         break_innermost=0
#         for key, value in address_json.items():
#             for i in value["Column_Type"]:
#                 if col in [item.lower() for item in i["Input_Column"]]:
#                     if address_json[key]["flag"] == 1 and len(addset) > 0:
#                         count = count + 1
#                         mv_mappedOPColsList.append(addset)
#                         flag = True
#                         addset = []
#                     ostr=i["Tag"] + '@' + key + '@' + str(count)
#                     mv_mapped_op_col_list.append(key)
#                     mv_mappedOPColsDict[col.lower()] = ostr
#                     # identifiedColumns.append(i["Tag"] + '@' + key + '@' + str(count))
#                     addset.append(ostr)
#                     flag = False
#                     # loc = inputDF.columns.get_loc(col)
#                     break_innermost=1
#                     break
#             if break_innermost:
#                 break
#     mv_unmappedOPCols=[col for col in address_json.keys() if col not in mv_mapped_op_col_list]
#     mv_unmappedInpCols=[col for col in input_columns  if col not in mv_mappedOPColsDict.keys()]
#     print 'mv_mappedOPColsDict:',mv_mappedOPColsDict
#
#     if not flag and len(addset) > 0:
#         mv_mappedOPColsList.append(addset)
#
#     return mv_mappedOPColsDict,mv_mappedOPColsList,mv_unmappedOPCols,mv_unmappedInpCols


# def mv_mapColumns(input_columns,mv_mapping_json):
#     return checkAddressColumns(input_columns,mv_mapping_json)

def splitFrames1(frames, sv_splitFrameCols, mv_splitFrameCols,*args):
    errDictDupValInSvColName = "errDictDupValInSv"
    import mvDerivation_Common as MV

    rejectedRecordsLog = " *** RejectedRecordsLog *** \n "
    splittedFrames = []
    rejectedFramesLog = ""
    for idx, frame in enumerate(frames):
        ttype=frame['VALIDATED_ACTION'][0]
        errDictDupValInSv = {}
        rowDF = pd.DataFrame()
        skipFrame = False
        droppedRows = frame.drop_duplicates(subset=list(sv_splitFrameCols.keys()))[
                      1:]  # drop duplicates and check if got multiple rows(single valued case having multiple values exception.)
        if not droppedRows.empty:
            failedCols = []
            rejectedRecordsLog += "--------------------------Single Valued Rejected Frame Errors---------------------------------\n"
            for key, value in list(sv_splitFrameCols.items()):
                if len(frame[key].unique()) > 1:
                    print("Log: ERROR Duplicate records For 'Tax-Id': '" + str(
                        frame.TAX_ID.unique()) + "' 'NPI': '" + str(frame.FINAL_NPI.unique()) + "'" + key)
                    failedCols.append(key)
                    errDictDupValInSv[key] = MV.ErrorMessage([],
                                                             ["c", "Multiple values found, Picking up the first value",
                                                              str(list(frame[key])[0])])
                rowDF.loc[0, key] = frame[key].unique()[0]
            log = "Log: ERROR Duplicate records For 'Tax-Id': '" + str(frame.TAX_ID.unique()[0]) + "' 'NPI': '" + str(
                frame.FINAL_NPI.unique()[0]) + "' - [" + ", ".join(failedCols) + "]"
            rejectedFramesLog += log + "\n"

        else:
            rowDF = frame[:1][list(sv_splitFrameCols.keys())]
        # rowDF[errDictDupValInSvColName] = str(errDictDupValInSv)
        diAssign = {errDictDupValInSvColName: str(errDictDupValInSv)}
        rowDF = rowDF.assign(**diAssign)
        # skip the record which has error, for later error suppressions.
        if skipFrame == True:
            continue

        extractCols = []

        for key, value in list(mv_splitFrameCols.items()):
            for col in frame.columns:
                colnmList = col.split("@")
                if len(colnmList) > 1:
                    if key == colnmList[1].split('#')[0]:
                        extractCols.append(col)
                else:
                    if key == col:
                        extractCols.append(col)

        mvRowDF = frame[extractCols]
        frameTuple = (rowDF, mvRowDF, idx,args[0],args[1],args[2],ttype)
        splittedFrames.append(frameTuple)
    return splittedFrames, rejectedFramesLog

# def mv_address_normalization(frame,mv_address_sets):
#     fdf = pd.DataFrame()
#     for address_set in mv_address_sets:
#         tdf = frame[address_set].drop_duplicates()
#         # print tdf
#         for index, value in tdf.iterrows():
#             tdict = {}
#             for col in tdf.columns:  # Todo: include include all columns in output.
#                 # print value[col]
#                 x = col.split('@')
#                 # print x[1]
#                 tdict['ADDRESS_TYPE'] = x[0]
#                 tdict[x[1]] = value[col]
#                 # print x
#             fdf = fdf.append(tdict, ignore_index=True)
#             #            for k in tdict.keys():
#             #                fdf.loc[index,k] = tdict[k]
#
#             #    print fdf
#     return fdf

# def getExcelFieldNames(ExcelLocation, sheetIdx, HeaderIdx):
#     def getrow(ws, hdidx):
#         count = 0
#         for row in ws.iter_rows():
#             count = count + 1
#             if count == hdidx:
#                 return row
#
#     def CleanColumnName(row):
#         col = []
#         dstnctCol = []
#         dupCol = []
#         for i in [x for x in row if x.value is not None]:
#             val = i.value
#             val = val.replace('\n', '|').replace('\r', '|')
#             val = [x for x in val.split("|") if x][0]
#             val = val.strip().lower()
#             col.append(val)
#             if val not in dstnctCol:
#                 dstnctCol.append(val)
#             else:
#                 dupCol.append(val)
#         return col, dupCol
#
#     def getFinalRowHeader(rowCol, dupCol):
#         process = []
#         for i in dupCol:
#             count = 1
#             for j, x in enumerate(rowCol):
#                 if x == i:
#                     if x in process:
#                         count = count + 1
#                     else:
#                         process.append(x)
#                     v = x + str(count)
#                     rowCol[j] = v
#         return rowCol
#     workbook = openpyxl.load_workbook(ExcelLocation)
#     workSheet = workbook[str(workbook.sheetnames[sheetIdx])]
#     count = 0
#     HeaderRow = getrow(workSheet, HeaderIdx)
#     rowCol, dupCol = CleanColumnName(HeaderRow)
#     return getFinalRowHeader(rowCol, dupCol)

def mvAddressNormalization(frame,mvAddressSets,parentCategory,lstAsIs,mvMappings):
    fdf = pd.DataFrame()
    # drvColList =['DERIVED_'+ lval for lval in lstAsIs]
    drvColList =['DERIVED_ADDRESS_INDICATOR','DERIVED_ADDRESS_TYPE']
    # drvTypeCol='DERIVED_'+parentCategory.upper()+'_TYPE'
    # indicatorCol = parentCategory.upper() + '_INDICATOR'
    # TypeCol = parentCategory.upper() + '_TYPE'
    groupCols = [k for k, v in list(mvMappings.items()) if v.get('GROUP_DETAILS')]
    diGroup = getDiGroupDefault(parentCategory,mvMappings)

    def keepOnlyNonBlankDi(lstDi):
        nonBlank = [elm for elm in lstDi if any(elm.values())]
        if nonBlank:
            return nonBlank
        else:
            return lstDi[0:1]
    for address_set in mvAddressSets:
        diGroupAdrSet=copy.deepcopy(diGroup)
        # tdf = frame[address_set].drop_duplicates()
        # tdf = frame[['ADDRESS_TYPE','ADDRESS_INDICATOR']+address_set]
        tdf = frame[address_set]
        tdf = tdf.replace(r'^\s+|\s+$', '', regex=True).replace('', np.nan)
        tdf = tdf.replace(r'\s\s+', ' ', regex=True).replace('', np.nan)
        sbst = [col for col in frame[address_set].columns if
                'ADDRESS_LINE_1' in col or 'ADDRESS_LINE_2' in col or 'ADDRESS_STATE' in col or 'ADDRESS_CITY' in col or 'ADDRESS_ZIP' in col]
        tdf = tdf.dropna(how='all')

        # print tdf
        diGroupAdrSet=updateDiGroup(diGroupAdrSet,address_set,mvMappings)
        for index, value in tdf.iterrows():
            # tdict = {}
            tdict = {key: None for key, val in list(mvMappings.items()) if val['PARENT_CATEGORY'] == parentCategory and not val.get('GROUP_DETAILS')}


            for col in tdf.columns:  # Todo: include include all columns in output.
                # if col in lstAsIs:
                #     tdict[col]=value[col]
                #     continue
                # print value[col]
                x = col.split('@')
                tdict['order']=x[-1]
                # print x[1]
                x1=x[0].split('_')
                outColParts=x[1].split('#')


                if len(x1)==2 and len(drvColList) == 2:
                    tdict[drvColList[0]] = x1[0]
                    tdict[drvColList[1]] = x1[1]
                elif len(x[0].split('_'))==1 and len(drvColList) == 2:
                    tdict[drvColList[0]] = ''
                    tdict[drvColList[1]] = x1[0]
                if len(outColParts)==2:
                    outCol = outColParts[0]
                    grpCounter=int(outColParts[1])
                    grpName=mvMappings[outCol]['GROUP_DETAILS']['GROUP_NAME']
                    diGroupAdrSet[grpName][grpCounter][outCol]='' if pd.isnull(value[col]) else value[col]
                else:
                    tdict[x[1]] = value[col]
                # print x
            diGroupTrans = {}


            for grpNm, di in list(diGroupAdrSet.items()):
                diGroupTrans[grpNm] = str(keepOnlyNonBlankDi([tup[1] for tup in sorted(di.items())]))
            tdict.update(copy.deepcopy(diGroupTrans))
            fdf = fdf._append(tdict, ignore_index=True)
            #            for k in tdict.keys():
            #                fdf.loc[index,k] = tdict[k]

            #    print fdf
    # if TypeCol in fdf.columns:
    fdf = fdf.drop_duplicates()
    if fdf.empty:
        diGroupTrans = {}
        for grpNm, di in list(diGroup.items()):
            diGroupTrans[grpNm] = str(keepOnlyNonBlankDi([tup[1] for tup in sorted(di.items())]))
        tdict = {key: None for key, val in list(mvMappings.items()) if val['PARENT_CATEGORY'] == parentCategory and key not in groupCols}
        tdict.update({item: None for item in drvColList+['order']})
        tdict.update(copy.deepcopy(diGroupTrans))
        fdf = fdf._append(tdict, ignore_index=True)
        pass
    fdf = fdf.reset_index()
    return fdf

def mvGeneralNormalization(frame,mvAddressSets,parentCategory,lstAsIs,mvMappings):
    fdf = pd.DataFrame()
    # drvColList =['DERIVED_'+ lval for lval in lstAsIs]
    drvColList = ['DERIVED_'+parentCategory.upper()+'_INDICATOR']
    # drvTypeCol = 'DERIVED_' + parentCategory.upper() + '_TYPE'
    # indicatorCol = parentCategory.upper() + '_INDICATOR'
    # TypeCol = parentCategory.upper() + '_TYPE'
    for address_set in mvAddressSets:
        # tdf = frame[address_set].drop_duplicates()
        tdf=frame[address_set]
        tdf = tdf.replace(r'^\s+|\s+$', '', regex=True).replace('', np.nan)
        tdf = tdf.dropna(how='all')

        # print tdf
        for index, value in tdf.iterrows():
            tdict = {key: None for key, val in list(mvMappings.items()) if val['PARENT_CATEGORY'] == parentCategory}
            for col in tdf.columns:  # Todo: include include all columns in output.
                # if col in lstAsIs:
                #     tdict[col]=value[col]
                #     continue
                # print value[col]
                x = col.split('@')
                tdict[parentCategory+'_order'] = x[-1]
                # print x[1]
                if len(drvColList) >= 1:
                    tdict[drvColList[0]] = x[0]
                tdict[x[1]] = value[col]
                # print x
            fdf = fdf._append(tdict, ignore_index=True)
            #            for k in tdict.keys():
            #                fdf.loc[index,k] = tdict[k]

            #    print fdf
    fdf = fdf.drop_duplicates()
    if fdf.empty:
        tdict = {key: None for key, val in list(mvMappings.items()) if val['PARENT_CATEGORY'] == parentCategory}
        tdict.update({item:None for item in drvColList+[parentCategory+'_order']})
        fdf=fdf._append(tdict, ignore_index=True)
        pass
    #todo: logic for updating original address_type and address_indicator column based on the values of the derived one
    fdf = fdf.reset_index()
    return fdf

def mvNormalizeColumns(splittedFrame,functionMappings,dictMVMappings,mvMappings):
    # normalizedMVFrames=[]
    # tupple = splittedFrame
        # print type(tupple),len(tupple)
    dictNdf={}
    for parentCategory in list(functionMappings.keys()):
        # if parentCategory!='Address':
        #     continue
        lstAsIs = [(val["order"], key) for key, val in list(mvMappings.items()) if
                   val['PARENT_CATEGORY'] == parentCategory and val['DATAFRAME_TYPE'] == 'AsIs' and key in
                   list(dictMVMappings[parentCategory][0].values())]
        lstAsIs.sort()
        lstAsIs = [l[1] for l in lstAsIs]

        ndf = eval(functionMappings[parentCategory]['normalization'])(splittedFrame, dictMVMappings[parentCategory][1],parentCategory, lstAsIs,mvMappings)
        ndf["ROW_NUM"] = ndf.index
        ndf["ROW_COUNT"] = len(ndf.index)
        dictNdf.update({parentCategory: ndf})
        # normalizedMVFrames.append(dictNdf)
        # dictNdf = concatdictNdf(dictNdf)

    dictNdf = getFinalDegreeDF(dictNdf)
    return dictNdf

def getFinalDegreeDF(dictNdf):
    degreeDF = dictNdf["Degree"]
    credDF = dictNdf["Credential"]
    titleDF = dictNdf["Title"]

    if not degreeDF.empty and "DEGREE" in degreeDF.columns and not degreeDF["DEGREE"].isnull().all():
        dictNdf["FinalDegree"] = degreeDF
        return dictNdf
    elif not titleDF.empty and "TITLE" in titleDF.columns and not titleDF["TITLE"].isnull().all():
        dictNdf["FinalDegree"] = titleDF.rename(columns={'TITLE': 'DEGREE', 'PRIMARY_TITLE_IND': 'DEGREE_INDICATOR', 'DERIVED_TITLE_INDICATOR': 'DERIVED_DEGREE_INDICATOR','Title_order':'Degree_order'})
        dictNdf["FinalDegree"]["ROW_NUM"] = titleDF["ROW_NUM"]
        dictNdf["FinalDegree"]["ROW_COUNT"] = titleDF["ROW_COUNT"]
        return dictNdf
    elif not credDF.empty and "CREDENTIAL" in credDF.columns and not credDF["CREDENTIAL"].isnull().all():
        dictNdf["FinalDegree"] =  credDF.rename(columns={'CREDENTIAL': 'DEGREE', 'PRIMARY_CREDENTIAL_IND': 'DEGREE_INDICATOR', 'DERIVED_CREDENTIAL_INDICATOR': 'DERIVED_DEGREE_INDICATOR','Credential_order':'Degree_order'})
        dictNdf["FinalDegree"]["ROW_NUM"] = credDF["ROW_NUM"]
        dictNdf["FinalDegree"]["ROW_COUNT"] = credDF["ROW_COUNT"]
        return dictNdf
    else:
        dictNdf["FinalDegree"] = pd.DataFrame(data={"DEGREE":[''], "DEGREE_INDICATOR":[''], "DERIVED_DEGREE_INDICATOR": [''], "ROW_NUM":[0], "ROW_COUNT":[1],'Degree_order':[0]}).reset_index()
        return dictNdf

def updateDiGroup(diGroup,address_set,mvMappings):
    for elm in address_set:
        mappingParts = elm.split('@')
        outColParts = mappingParts[1].split('#')
        # print(outColParts)
        if len(outColParts) == 2:
            outCol = outColParts[0]
            grpCounter = int(outColParts[1])
            grpName = mvMappings[outCol]['GROUP_DETAILS']['GROUP_NAME']
            groupDi={k:"" for k,v in list(mvMappings.items()) if v.get('GROUP_DETAILS') and v['GROUP_DETAILS']['GROUP_NAME']==grpName}
            # print(grpName)
            if grpCounter not in diGroup[grpName]:
                diGroup[grpName].update({grpCounter: groupDi})
    return diGroup

def getDiGroupDefault(pCat,mvMappings):
    diGroup = {}
    mvMappingsFiltered = {k: v for k, v in list(mvMappings.items()) if v.get('PARENT_CATEGORY')==pCat and v.get('GROUP_DETAILS')}
    for k,v in list(mvMappingsFiltered.items()):
        grpName=v['GROUP_DETAILS']['GROUP_NAME']
        if grpName not in diGroup:
            diGroup.update({grpName:{0:{k1:"" for k1,v1 in list(mvMappingsFiltered.items()) if v1['GROUP_DETAILS']['GROUP_NAME']==grpName}}})
    return diGroup

def mapMVUnorderedColumns(inputColumns, mvMappingJson,parentCategory, mappSotCols):
    mvMappedOPColsDict = {}  # renameColumns expects dictionary
    mvMappedOPCols = []
    filterdMappingJson={key:value for key,value in list(mvMappingJson.items()) if value['PARENT_CATEGORY'] == parentCategory and value['DATAFRAME_TYPE'] == "Normalize"}
    asIsMappingJson={key:value for key,value in list(mvMappingJson.items()) if value['PARENT_CATEGORY'] == parentCategory and value['DATAFRAME_TYPE'] == "AsIs"}
    mapDi={}
    for col in inputColumns:
        if col in mappSotCols:
            continue
        for key, value in list(filterdMappingJson.items()):
            for i in value["Column_Type"]:
                if col in [item.lower() for item in i["Input_Column"]]:
                    diKey = key
                    if diKey not in mapDi:
                        mapDi.update({diKey:[]})
                    mapDi[diKey].append((i["Tag"], col))
                    break
            else:
                continue
            break


        for k , v in list(asIsMappingJson.items()):
            if col in [item for item in v["Input_Column"] if item not in list(mvMappedOPColsDict.keys())]:
                mvMappedOPCols.append(k)
                mvMappedOPColsDict[col.lower()] = k
    print(mapDi)
    setDi = {}
    for k, v in list(mapDi.items()):
        for idx, item in enumerate(v):
            map = item[0] + '@' + k + '@' + str(idx)
            mvMappedOPColsDict[item[1]] = map
            if idx not in setDi:
                setDi.update({idx: []})
            setDi[idx].append(map)
    print(setDi)
    mvMappedOPColsList = [v for k,v in sorted(setDi.items())]
    print('mvMappedOPColsList:',mvMappedOPColsList)
    mvMappedOPCols=[key for key in list(mapDi.keys())]
    mvUnmappedOPCols=[col for col in list(filterdMappingJson.keys()) if col not in mvMappedOPCols]
    mvUnmappedInpCols=[col for col in inputColumns  if col not in list(mvMappedOPColsDict.keys())]
    print('mvMappedOPColsDict:',mvMappedOPColsDict)

    return mvMappedOPColsDict,mvMappedOPColsList,mvUnmappedOPCols,mvUnmappedInpCols

def mapMVGeneralColumns(inputColumns, mvMappingJson, parentCategory, mappedInpCols):
    mvMappedOPColsDict = {}  # renameColumns expects dictionary
    missingColumns = []  # missing columns in the input file after checking with columns.json
    identifiedColumns = []
    mvMappedOPCols = []
    mvMappedOPColsList = []
    addset = []
    count = 0
    flag = False
    filterdMappingJson = {key: value for key, value in list(mvMappingJson.items()) if
                          value['PARENT_CATEGORY'] == parentCategory and value['DATAFRAME_TYPE'] == "Normalize"}
    asIsMappingJson = {key: value for key, value in list(mvMappingJson.items()) if
                       value['PARENT_CATEGORY'] == parentCategory and value['DATAFRAME_TYPE'] == "AsIs"}
    tag = ''
    groupTracker = {}
    prev_group_name = None
    for col in inputColumns:
        # if col in ['languages spoken at this location temp1','languages spoken at this location temp2','languages spoken at this location temp3','language spoken by temp1','language spoken by temp2','language spoken by temp3']:
        #     pass
        if col in mappedInpCols:
            # if prev_group_name:
            #     groupTracker[prev_group_name]['keys'] = []
            #     groupTracker[prev_group_name]['counter'] += 1
            continue
        break_innermost = 0

        for key, value in list(filterdMappingJson.items()):
            for i in value["Column_Type"]:
                if col in [item.lower() for item in i["Input_Column"]]:
                    if filterdMappingJson[key]["flag"] == 1 and len(addset) > 0:
                        groupTracker = {}
                        prev_group_name = None
                        count = count + 1
                        mvMappedOPColsList.append(addset)
                        tag = i["Tag"]
                        flag = True
                        addset = []

                    # change from Raj <
                    if value.get('GROUP_DETAILS'):
                        group_name = value['GROUP_DETAILS']['GROUP_NAME']
                    else:
                        group_name=None
                    # if prev_group_name:
                    #     if prev_group_name != group_name:
                    #         prev_group_name = group_name
                    #         if group_name and group_name in groupTracker:
                    #             groupTracker[group_name]['keys'] = []
                    #             #groupTracker[group_name]['counter'] += 1 #todo this step should be verified
                    # elif not prev_group_name and group_name:
                    #     prev_group_name=group_name
                    if prev_group_name != group_name:
                       if prev_group_name:
                           groupTracker[prev_group_name]['keys'] = []
                           groupTracker[prev_group_name]['counter'] +=1
                       prev_group_name = group_name

                    if group_name and group_name in groupTracker:
                        if key in groupTracker[group_name]['keys']:
                            groupTracker[group_name]['keys'] = [key]
                            groupTracker[group_name]['counter'] += 1
                        else:
                            groupTracker[group_name]['keys'].append(key)
                    elif group_name and group_name not in groupTracker:
                        groupTracker.update({group_name: {'counter': 0, 'keys': [key]}})

                    # >change from Raj

                    if tag == '':
                        tag = i["Tag"]
                    if group_name:
                        ostr = tag + '@' + key + '#' + str(groupTracker[group_name]['counter']) + '@' + str(count)
                    else:
                        ostr = tag + '@' + key + '@' + str(count)
                    if key not in [item.split('@')[1] for item in addset]: #['PLSV@ADDRESS_LINE_1@0','PLSV@ADDRESS_CITY@0','PLSV@lANGUAGE#0@0','PLSV@lANGUAGE#1@0']
                        mvMappedOPCols.append(key)
                        mvMappedOPColsDict[col.lower()] = ostr
                        # identifiedColumns.append(i["Tag"] + '@' + key + '@' + str(count))
                        addset.append(ostr)
                        flag = False
                        # loc = inputDF.columns.get_loc(col)
                        break_innermost = 1
                        break
            if break_innermost:
                break
        # else:
        #     if prev_group_name:
        #         groupTracker[prev_group_name]['keys'] = []
        #         groupTracker[prev_group_name]['counter'] += 1

        for k, v in list(asIsMappingJson.items()):
            if col in [item for item in v["Input_Column"] if item not in list(mvMappedOPColsDict.keys())]:
                mvMappedOPCols.append(k)
                mvMappedOPColsDict[col.lower()] = k
    mvUnmappedOPCols = [col for col in list(filterdMappingJson.keys()) if col not in mvMappedOPCols]
    mvUnmappedInpCols = [col for col in inputColumns if col not in list(mvMappedOPColsDict.keys())]
    print('mvMappedOPColsDict:', mvMappedOPColsDict)

    if not flag and len(addset) > 0:
        mvMappedOPColsList.append(addset)

    return mvMappedOPColsDict, mvMappedOPColsList, mvUnmappedOPCols, mvUnmappedInpCols

def mvMapColumns(inputColumns,mvMappingJson,functionMappings):
    dictMVMappings={}
    mappedSotCols = []
    for parentCategory in list(functionMappings.keys()):
        # if parentCategory != 'Address':
        #     continue
        pCatMapping={parentCategory: list(
            eval(functionMappings[parentCategory]['mapping'])(inputColumns, mvMappingJson, parentCategory,mappedSotCols))}
        dictMVMappings.update(pCatMapping)
        mappedSotCols = mappedSotCols + list(pCatMapping[parentCategory][0].keys())
    # dictMVMappings = adoptOrphanSets(dictMVMappings,'Address',1)
    diOrphanColumns={'Address':[('DIR_IND',),('ELECTRONIC_COMM',),('AGE_LIMIT', 'MIN_AGE', 'MAX_AGE'),('WORKING_DAYS','WORKING_HOURS','BREAK_HOURS'),('MONDAY_START_TIME', 'MONDAY_END_TIME', 'TUESDAY_START_TIME', 'TUESDAY_END_TIME', 'WEDNESDAY_START_TIME', 'WEDNESDAY_END_TIME', 'THURSDAY_START_TIME', 'THURSDAY_END_TIME', 'FRIDAY_START_TIME', 'FRIDAY_END_TIME', 'SATURDAY_START_TIME', 'SATURDAY_END_TIME', 'SUNDAY_START_TIME', 'SUNDAY_END_TIME')]}
    dictMVMappings = adoptOrphanColumns(dictMVMappings,'Address',diOrphanColumns['Address'])
    return dictMVMappings

def combineMVMapping(dictMVMappings):
    mvMappedOPColsDict = {}
    mvMappedOPColsList = []
    mvUnmappedOPCols = []
    mvUnmappedInpCols = []

    for key, lst in list(dictMVMappings.items()):
        mvMappedOPColsDict.update(lst[0])
        mvMappedOPColsList += lst[1]
        mvUnmappedOPCols += lst[2]
        mvUnmappedInpCols += lst[3]
    mvUnmappedOPCols=list(set(mvUnmappedOPCols))
    mvUnmappedInpCols=[col for col in list(set(mvUnmappedInpCols)) if col not in list(mvMappedOPColsDict.keys())]

    return mvMappedOPColsDict,mvMappedOPColsList,mvUnmappedOPCols,mvUnmappedInpCols

def adoptOrphanColumns(dictMVMappings,pCat,lstOrphCols):
    dictMVMappingsCopy = copy.deepcopy(dictMVMappings)
    di = dictMVMappingsCopy[pCat][0]
    lol = dictMVMappingsCopy[pCat][1]
    diOrphCol={}
    for tup in lstOrphCols:
        flag=0
        tempDiOrphCol={}
        for col in tup:
            lstMatch=list({item.split('@')[0]+'#'+item.split('@')[1].split('#')[1]+'@'+item.split('@')[2] if len(item.split('@')[1].split('#'))==2 else item.split('@')[0]+'@'+item.split('@')[2] for sublist in lol for item in sublist if col==item.split('@')[1].split('#')[0]})
            # lstMatch=['BILL@1','MAILING@2']
            if re.search('PLSV|GENERAL',' '.join(lstMatch),re.IGNORECASE):
                flag=1
                break
            if lstMatch:
                tempDiOrphCol.update({col:lstMatch})
        if not flag:
            diOrphCol.update(tempDiOrphCol)
    lstTemp=[(lst[0].split('@')[0],int(lst[0].split('@')[2]),idx) for idx,lst in enumerate(lol) if re.search('PLSV|GENERAL',lst[0].split('@')[0],re.IGNORECASE)]
    sorted(lstTemp,key=lambda x: x[1],reverse=True)
    if len(lstTemp) == 0:
        return dictMVMappings
    else:
        adopterSet = lstTemp[0]

    deltaDiValues={}
        # deltaDiValues.update({item.split('@')[0]+k+item.split('@')[1]:adopterSet[0]+k+str(adopterSet[1]) for item in v})
    for k,v in list(diOrphCol.items()):
        for item in v:
            splt=item.split('@')
            lstSplt0=splt[0].split('#')
            if len(lstSplt0)==2:
                oldName=lstSplt0[0]+'@'+k+'#'+lstSplt0[1]+'@'+splt[1]
                newName=adopterSet[0]+'@'+k+'#'+lstSplt0[1]+'@'+str(adopterSet[1])
            else:
                oldName = splt[0] + '@' + k + '@' + splt[1]
                newName = adopterSet[0] + '@' + k + '@' + str(adopterSet[1])
            deltaDiValues.update({oldName:newName})
            lol[adopterSet[2]].append(newName)
            idxOrphanLst=[idx for idx,lst in enumerate(lol) if lst[0].split('@')[2]==item.split('@')[1]]
            if idxOrphanLst:
                lol[idxOrphanLst[0]].remove(oldName)
    # lol=[lst for lst in lol if lst]
    if [] in lol:
        lol.remove([])
    for k,v in list(di.items()):
        if v in deltaDiValues:
            di[k]=deltaDiValues[v]
    return dictMVMappingsCopy

def adoptOrphanSets(dictMVMappings,pCat,adopter):
    dictMVMappingsCopy=copy.deepcopy(dictMVMappings)
    di=dictMVMappingsCopy[pCat][0]
    lol=dictMVMappingsCopy[pCat][1]
    idxOrphan=determineOrpanSet(lol)
    counterAdopter=lol[adopter][0].split('@')[2]
    typeAdopter=lol[adopter][0].split('@')[0]
    outpuColsAdopter=[elem.split('@')[1] for elem in lol[adopter]]
    modifiedOrphan=[(e,typeAdopter+'@'+e.split('@')[1]+'@'+counterAdopter) for e in lol[idxOrphan] if e.split('@')[1] not in outpuColsAdopter]
    if len(lol)>=2 and idxOrphan is not None:
        lol[adopter]=lol[adopter]+ [elem1 for elem0,elem1 in modifiedOrphan]
    for k,v in list(di.items()):
        if v in dict(modifiedOrphan):
            di[k]=dict(modifiedOrphan)[v]
            print(k,dict(modifiedOrphan)[v])
    lol.pop(idxOrphan)
    return dictMVMappingsCopy


def determineOrpanSet(lol):
    for idx,l in enumerate(lol):
        if not len([item for item in l if 'ADDRESS_LINE_1' in item]):
            return idx

def ErrorMessage(lst, Input):
    if not len(lst) >= 1:
        Type = Input[0]
        Message = Input[1]
        Inp = Input[2]
    else:
        Type = lst[0].populate(Input)
        Message = lst[1].populate(Input)
        Inp = lst[2].populate(Input)

    intrimStr = str(Message) + "|" + "Input: " + str(Inp)
    if Type == "R":
        return "Reject: " + intrimStr
    elif Type == "E":
        return "Error: " + intrimStr
    else:
        return "Clarify: " + intrimStr

def isErrorVal(str):
    return bool(re.search("reject\:|error\:|clarify\:", str, re.IGNORECASE))

def fetchResponse(url,params={}):
    headers = {"Accept": "text/html/json,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
               "User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36"}
    res=requests.get(url=url,params=params,headers=headers,verify=False)
    return res.json()
    # responseMap = {
    #     "https://plmi-stage.uhc.com/api/GetAll/GetDelegates": "https_plmi-stage_uhc_com_api_GetAll_GetDelegates",
    #     "https://plmi-stage.uhc.com/api/DLIStatusAPI/GET": "https_plmi-stage_uhc_com_api_DLIStatusAPI_GET",
    #     "https://plmi-stage.uhc.com/api/Delegates/System": "https_plmi-stage_uhc_com_api_Delegates_System",
    #     "https://plmi-stage.uhc.com/api/TaxIdInfo/Get": "https_plmi-stage_uhc_com_api_TaxIdInfo_Get_DelegateId=116",
    #     "https://plmi-stage.uhc.com/api/PlatformOverview/Get": "https_plmi-stage_uhc_com_api_PlatformOverview_Get_DelegateId=116"}
    # # res=requests.get(url=url,params=params,headers=headers)
    # dirPath='C:\Users\RA373432\Downloads\uhg market info Updated\uhg market info\output\\'
    # filePath=dirPath+responseMap[url]+'.json'
    # with open(filePath,'rb') as fp:
    #     res=json.load(fp)
    # return res

def createOneRowDf(cols):
    df = pd.DataFrame([{col: "" for col in cols}])
    return df

def renameColsTovalidNames(df):
    tmp=copy.deepcopy(df)
    tmp.columns = [re.sub('[^\w]', '_', val) for val in list(tmp.columns)] #renames columns to view tmp dataframe in pycharm
    return tmp

def fetchProviderCondition(*args):
    fields=("Provider","Type")
    return dict(list(zip(fields,args)))


def splitFrames2(frames, svCols, mvCols,*args):
    splittedFrames = []
    for idx, frame in enumerate(frames):
        errDictDupValInSv={}
        ttype=frame['VALIDATED_ACTION'][0]
        svDf = frame.filter(items=svCols)
        mvDf = frame.filter(items=mvCols)
        svDf1=svDf.iloc[:1]
        uniqRowCount=1 if len(svDf)==1 else svDf.apply(lambda row: ''.join(row.values), axis=1).nunique()
        if uniqRowCount>1:
            msg="Multiple values found, Picking up the first value"
            mvSer = svDf1.iloc[0][svDf.apply(lambda cl: len(np.unique(cl)) > 1)]
            errDictDupValInSv=dict(mvSer.apply(lambda val: ErrorMessage([], ["c",msg , val])))

        svDf1["errDictDupValInSv"]=str(errDictDupValInSv)
        frameTuple = (svDf1, mvDf, idx,args[0],args[1],args[2],ttype)
        splittedFrames.append(frameTuple)
    return splittedFrames, ""

def splitFrames3(frame, svCols, mvCols):
    errDictDupValInSv={}
    svDf = frame.filter(items=svCols)
    mvDf = frame.filter(items=mvCols)
    svDf1=svDf.iloc[:1]
    uniqRowCount=1 if len(svDf)==1 else svDf.apply(lambda row: ''.join(row.values), axis=1).nunique()
    if uniqRowCount>1:
        msg="Multiple values found, Picking up the first value"
        mvSer = svDf1.iloc[0][svDf.apply(lambda cl: len(np.unique(cl)) > 1)]
        errDictDupValInSv=dict(mvSer.apply(lambda val: ErrorMessage([], ["c",msg , val])))

    svDf1["errDictDupValInSv"]=str(errDictDupValInSv)
    return svDf1, mvDf

def getNlpLookup(svMap,mvMap):
    def func(fieldType):
        if fieldType == 'sv':
            diLookup = {key: val['Input_Column'] for key, val in list(svMap.items())}
        elif fieldType == 'mv':
            diLookup = getMvLookup(mvMap)
        diTokenizedLookup = {}
        for pCat, looup in list(diLookup.items()):
            diTokens = {}
            for word in looup:
                dict_word_combination = nlp.make_word_combination(word)
                diTokens[word] = dict_word_combination
            diTokenizedLookup.update({pCat: diTokens})
        return diTokenizedLookup

    sv_nlp_lookup=func('sv')
    mv_nlp_lookup=func('mv')
    return sv_nlp_lookup,mv_nlp_lookup

def getMvLookup(mvMap):
    lstParentCat = []
    diParentCat = {}
    for key, val in list(mvMap.items()):
        # if key not in excluded and key in outputCols:
        if val['PARENT_CATEGORY'] not in lstParentCat:
            lstParentCat.append(val['PARENT_CATEGORY'])
        for lookup in val.get('Column_Type', []):
            if val['PARENT_CATEGORY'] not in diParentCat:
                diParentCat.update({val['PARENT_CATEGORY']: []})
            if lookup['Input_Column'] != [""]:
                diParentCat[val['PARENT_CATEGORY']] += lookup['Input_Column']
    return diParentCat